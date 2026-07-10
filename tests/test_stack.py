"""Tests for the natural-language markdown authoring layer.

Everything here runs offline: the loader is structural, and the loaded
runtime exercises its deterministic fallbacks, so the whole stacked
authoring model -- skills.md -> persona.md -> workflow.md -> process.md ->
policy.md -> memory.md -> Runtime -- is testable with no LLM configured.
"""

from __future__ import annotations

import json
import os
import shutil
import time
from pathlib import Path

import pytest

from ear import (
    Exchange,
    Intent,
    KnowledgeSource,
    Loader,
    Persona,
    Runtime,
    SessionStore,
    Spawner,
    Strategy,
    load_runtime,
)
from ear.section import parse_document

EXAMPLE_STACK = Path(__file__).resolve().parent.parent / "examples" / "credit_risk_stack"

requires_anthropic_key = pytest.mark.skipif(
    os.environ.get("EAR_LIVE_TESTS") != "1" or not os.environ.get("ANTHROPIC_API_KEY"),
    reason="live-LLM tests are opt-in: set EAR_LIVE_TESTS=1 (and ANTHROPIC_API_KEY) to run them -- they bill real model calls",
)


# ---------------------------------------------------------------------------
# The markdown section parser.
# ---------------------------------------------------------------------------


def test_parse_document_splits_title_preamble_and_sections():
    document = parse_document(
        "# Skills\n\nPrompts stacked into skills.\n\n"
        "## banding\nBand the profile.\n\n## grading\nGrade the bands.\n"
    )
    assert document.title == "Skills"
    assert document.preamble == "Prompts stacked into skills."
    assert [section.name for section in document.sections] == ["banding", "grading"]
    assert document.section_named("Banding") is not None


def test_section_body_extracts_only_known_fields():
    document = parse_document(
        "## Cap\nThe cap is firm: never exceed it.\n\nFallback: amount <= 100\nApplies to: runtime\n"
    )
    body = document.sections[0].body(field_keys=("fallback", "applies to"))
    assert body.field("fallback") == "amount <= 100"
    assert body.field("applies to") == "runtime"
    # The colon inside prose is not swallowed as a field.
    assert "The cap is firm: never exceed it." in body.prose


def test_section_body_collects_bullets_and_numbered_items():
    document = parse_document("## W\n1. First step.\n2. Second step.\n- a bullet\n")
    body = document.sections[0].body()
    assert body.numbered == ["First step.", "Second step."]
    assert body.bullets == ["a bullet"]


def test_section_body_folds_wrapped_items_into_the_item_above():
    """Authors wrap long bullets and numbered steps at the column limit,
    indenting the continuation. Per-line parsing silently truncated every
    wrapped item -- a workflow step lost the '(Persona)' delegation at its
    end (so the step ran with no persona and half its instruction), and a
    deliverable bullet lost most of its declared meaning. An indented
    follow-on line folds into the item above; a blank or flush-left line
    ends it."""
    document = parse_document(
        "## W\n"
        "1. Load the raw source data -- extracts staged as the daily\n"
        "   sales workbook -- and run the sanity check that flags anomalies\n"
        "   as they land. (Sales MIS Guru)\n"
        "2. Second step, single line. (Controller)\n"
        "- status: exactly one of validated, blocked or pending -- validated\n"
        "  only when this cycle's own step completed\n"
        "\n"
        "Flush-left prose after a blank line stays prose.\n"
    )
    body = document.sections[0].body()
    assert body.numbered[0] == (
        "Load the raw source data -- extracts staged as the daily "
        "sales workbook -- and run the sanity check that flags anomalies "
        "as they land. (Sales MIS Guru)"
    )
    assert body.numbered[1] == "Second step, single line. (Controller)"
    assert body.bullets == [
        "status: exactly one of validated, blocked or pending -- validated "
        "only when this cycle's own step completed"
    ]
    assert "Flush-left prose after a blank line stays prose." in body.prose


def test_wrapped_workflow_steps_still_delegate_to_their_persona(tmp_path):
    """End-to-end through the Loader: a workflow step wrapped across lines
    keeps its trailing persona delegation and full instruction text."""
    stack = write_stack(
        tmp_path / "stack",
        skills="# Skills\n\n## banding\nBand the profile.\n",
        persona="# Personas\n\n## Credit Risk Guru\n\nBand carefully.\n\nSkills: banding\n",
        workflow=(
            "# Workflows\n\n## W\n\n"
            "1. Band the applicant's credit profile against every rule the\n"
            "   manual declares, never a shortcut. (Credit Risk Guru)\n"
        ),
        process="# Runtime\n\n## P\n\nRuns W.\n\nWorkflows: W\n",
    )
    runtime = load_runtime(stack)
    step = runtime.processes[0].workflows[0].steps[0]
    assert step.persona is not None and step.persona.name == "Credit Risk Guru"
    assert "never a shortcut" in step.instruction
    assert "(Credit Risk Guru)" not in step.instruction


def test_parse_document_handles_crlf_and_missing_title():
    document = parse_document("## Only Section\r\nSome prose.\r\n")
    assert document.title == ""
    assert document.sections[0].name == "Only Section"
    assert document.sections[0].body().prose == "Some prose."


# ---------------------------------------------------------------------------
# Stacking a runtime from markdown, layer by layer.
# ---------------------------------------------------------------------------


def write_stack(directory: Path, **files: str) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    for filename, text in files.items():
        (directory / f"{filename}.md").write_text(text, encoding="utf-8")
    return directory


MINIMAL_STACK = dict(
    skills=(
        "# Skills\n\n## risk_grade\nCombine the score band and DTI band into a grade A-E.\n\n"
        "## decide\nDecide approve or decline against the grade.\n"
    ),
    persona=(
        "# Personas\n\n## Credit Risk Guru\nUnderwrite conservatively.\n\n"
        "Skills: risk_grade, decide\n"
    ),
    workflow=(
        "# Workflows\n\n## Underwriting Workflow\n\n"
        "1. Band the profile and assign a risk grade. (Credit Risk Guru)\n"
        "2. Decide approve or decline against the grade. (Credit Risk Guru)\n\n"
        "Policies: Loan Amount Cap\n"
    ),
    process=(
        "# Credit Risk Runtime\n\n## Underwrite Consumer Loan\n"
        "Evaluates a consumer loan application end to end.\n\n"
        "Workflows: Underwriting Workflow\n"
    ),
    policy=(
        "# Policies\n\n## Loan Amount Cap\nThe loan must not exceed $75,000.\n\n"
        "Fallback: loan_amount <= 75000\nApplies to: Underwriting Workflow\n\n"
        "## DTI Ceiling\nDebt-to-income must not exceed 0.43.\n\n"
        "Fallback: debt_to_income <= 0.43\nApplies to: runtime\n"
    ),
)


def test_load_runtime_stacks_every_layer(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))

    assert runtime.name == "Credit Risk Runtime"
    assert [process.name for process in runtime.processes] == ["Underwrite Consumer Loan"]

    workflow = runtime.processes[0].workflows[0]
    assert workflow.name == "Underwriting Workflow"
    assert [step.instruction for step in workflow.steps] == [
        "Band the profile and assign a risk grade.",
        "Decide approve or decline against the grade.",
    ]
    assert all(step.persona.name == "Credit Risk Guru" for step in workflow.steps)

    persona = workflow.steps[0].persona
    assert [skill.name for skill in persona.skills] == ["risk_grade", "decide"]
    assert persona.skills[0].prompt == "Combine the score band and DTI band into a grade A-E."

    # Policy mapping: the cap governs the workflow, the ceiling the runtime.
    assert [policy.name for policy in workflow.policies] == ["Loan Amount Cap"]
    assert [policy.name for policy in runtime.policies] == ["DTI Ceiling"]


def test_loaded_runtime_reasons_offline_with_stacked_capabilities(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    decision = runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))
    assert "Credit Risk Runtime" in decision
    assert "Credit Risk Guru" in decision


def test_loaded_runtime_enforces_runtime_policy_from_markdown(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    with pytest.raises(PermissionError, match="DTI Ceiling"):
        runtime.reason(Intent(text="Underwrite a consumer loan", context={"debt_to_income": 0.60}))


def test_loaded_runtime_enforces_workflow_policy_from_markdown(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    with pytest.raises(PermissionError, match="Loan Amount Cap"):
        runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 90000}))


def test_unknown_skill_reference_fails_loudly(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["persona"] = "# Personas\n\n## Guru\nBe careful.\n\nSkills: not_a_skill\n"
    with pytest.raises(ValueError, match="Unknown skill 'not_a_skill'.*risk_grade"):
        load_runtime(write_stack(tmp_path / "stack", **stack))


def test_unknown_persona_delegation_stays_in_the_instruction(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = (
        "# Workflows\n\n## Underwriting Workflow\n\n1. Check the profile. (Nobody Known)\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    step = runtime.processes[0].workflows[0].steps[0]
    assert step.instruction == "Check the profile. (Nobody Known)"
    assert step.persona is None


def test_persona_can_stack_skills_as_bullets_and_define_inline(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["persona"] = (
        "# Personas\n\n## Credit Risk Guru\nUnderwrite conservatively.\n\n"
        "- risk_grade\n- decide\n- escalate: Escalate anything ambiguous to a human reviewer.\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    persona = runtime.processes[0].workflows[0].steps[0].persona
    assert [skill.name for skill in persona.skills] == ["risk_grade", "decide", "escalate"]
    assert persona.skills[2].prompt == "Escalate anything ambiguous to a human reviewer."


def test_workflow_default_persona_field_delegates_every_step(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = (
        "# Workflows\n\n## Underwriting Workflow\n\nPersona: Credit Risk Guru\n\n"
        "1. Band the profile.\n2. Decide the application.\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    workflow = runtime.processes[0].workflows[0]
    assert all(step.persona.name == "Credit Risk Guru" for step in workflow.steps)


def test_workflow_unreferenced_by_any_process_is_wrapped_not_dropped(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["process"] = "# Credit Risk Runtime\n"
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    assert [process.name for process in runtime.processes] == ["Underwriting Workflow"]
    assert runtime.processes[0].workflows[0].name == "Underwriting Workflow"


def test_policy_without_applies_to_governs_the_runtime(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["policy"] = "# Policies\n\n## Blanket Control\nEvery decision must be explainable.\n"
    stack["workflow"] = "# Workflows\n\n## Underwriting Workflow\n\n1. Decide. (Credit Risk Guru)\n"
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    assert [policy.name for policy in runtime.policies] == ["Blanket Control"]


def test_missing_files_load_an_empty_runtime(tmp_path):
    directory = tmp_path / "empty"
    directory.mkdir()
    runtime = load_runtime(directory)
    assert runtime.name == "empty"
    assert runtime.processes == []
    decision = runtime.reason(Intent(text="anything at all"))
    assert "empty" in decision


# ---------------------------------------------------------------------------
# tenant.md: the org a stack belongs to.
# ---------------------------------------------------------------------------


def test_missing_tenant_file_defaults_to_the_default_org(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    assert runtime.tenant.org_id == "default"
    assert runtime.tenant.fiscal_year_start is None
    assert runtime.tenant.fiscal_year_end is None


def test_tenant_file_stacks_org_id_and_fiscal_year(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["tenant"] = (
        "## Acme Capital\n\n"
        "Org id: org_acme_prod\n"
        "Fiscal year start: 2026-04-01\n"
        "Fiscal year end: 2027-03-31\n"
        "Timezone: Asia/Kolkata\n"
        "Secret env var: EAR_ACME_SECRET\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    tenant = runtime.tenant
    assert tenant.org_id == "org_acme_prod"
    assert tenant.name == "Acme Capital"
    assert tenant.fiscal_year_start.isoformat() == "2026-04-01"
    assert tenant.fiscal_year_end.isoformat() == "2027-03-31"
    assert tenant.timezone == "Asia/Kolkata"
    assert tenant.secret_env_var == "EAR_ACME_SECRET"


def test_tenant_file_without_org_id_fails_loudly(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["tenant"] = "## Acme Capital\n\nFiscal year start: 2026-04-01\n"
    with pytest.raises(ValueError, match="Org id"):
        load_runtime(write_stack(tmp_path / "stack", **stack))


def test_tenant_file_with_unreadable_date_fails_loudly(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["tenant"] = "## Acme Capital\n\nOrg id: org_acme_prod\nFiscal year start: sometime soon\n"
    with pytest.raises(ValueError, match="Fiscal year start"):
        load_runtime(write_stack(tmp_path / "stack", **stack))


def test_tenant_fiscal_year_bounds_falls_back_to_calendar_year():
    from datetime import date

    from ear.tenant import Tenant

    tenant = Tenant()
    start, end = tenant.fiscal_year_bounds(today=date(2026, 7, 5))
    assert (start, end) == (date(2026, 1, 1), date(2026, 12, 31))


# ---------------------------------------------------------------------------
# ear.identity.Claim: who is calling, and which Tenant they may act as.
# ---------------------------------------------------------------------------


def test_claim_may_act_as_the_org_ids_it_carries():
    from ear import Claim

    claim = Claim(subject="alice", org_ids=("org_acme_prod",))
    assert claim.may_act_as("org_acme_prod") is True
    assert claim.may_act_as("org_other") is False


def test_claim_require_raises_tenant_boundary_violation_for_an_unauthorized_org():
    from ear import Claim, TenantBoundaryViolation

    claim = Claim(subject="alice", org_ids=("org_acme_prod",))
    with pytest.raises(TenantBoundaryViolation, match="alice.*org_other"):
        claim.require("org_other")
    claim.require("org_acme_prod")  # authorized -- no raise


def test_runtime_reason_with_no_claim_behaves_exactly_as_before(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    decision = runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))
    assert "Credit Risk Runtime" in decision


def test_runtime_reason_refuses_a_claim_not_authorized_for_its_tenant(tmp_path):
    from ear import Claim, TenantBoundaryViolation

    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    assert runtime.tenant.org_id == "default"
    claim = Claim(subject="bob", org_ids=("org_other",))
    with pytest.raises(TenantBoundaryViolation, match="bob"):
        runtime.reason(Intent(text="Underwrite", context={"loan_amount": 5000}), claim=claim)


def test_runtime_reason_admits_a_claim_authorized_for_its_tenant(tmp_path):
    from ear import Claim

    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    claim = Claim(subject="alice", org_ids=("default", "org_other"))
    decision = runtime.reason(Intent(text="Underwrite", context={"loan_amount": 5000}), claim=claim)
    assert "Credit Risk Runtime" in decision


# ---------------------------------------------------------------------------
# The memory.md strategy: every setting read from natural language.
# ---------------------------------------------------------------------------

STRATEGY_MARKDOWN = """# Memory & Strategy

## Context History
Keep the 30 most recent cycles verbatim; compress anything older.

## Cross-Session Data
Persist memory, experience and adaptations to `.ear/session.json` between sessions.

## Subagent Spawning
Allow spawning up to 4 subagents, each scoped to a single persona.

## Model Selection
Reason with anthropic/claude-opus-4-8, reading the credential from
ANTHROPIC_API_KEY, at a temperature of 0.2.

## MCP
- credit_bureau: pulls credit reports and score history, via `bureau-mcp-server`
- core_banking: reads balances and repayment history, via `corebank-mcp-server`

## Tools
- amortization_calculator: computes the monthly payment for an amount, rate and term

## Skills Discovery
Rank processes by reading their descriptions against the intent, most relevant first.

## Ontological Settings
- risk grade: a letter from A to E, where A is the strongest credit
- decision: exactly one of approve or decline, never a hedge
"""


def test_strategy_reads_every_section_from_natural_language():
    strategy = Strategy.from_markdown(STRATEGY_MARKDOWN)

    assert strategy.history_capacity == 30
    assert strategy.session_enabled and strategy.session_path == ".ear/session.json"
    assert strategy.subagents_enabled and strategy.max_subagents == 4
    assert strategy.provider == "anthropic"
    assert strategy.model == "anthropic/claude-opus-4-8"
    assert strategy.api_key_env_var == "ANTHROPIC_API_KEY"
    assert strategy.temperature == 0.2
    assert [server.name for server in strategy.mcp_servers] == ["credit_bureau", "core_banking"]
    assert strategy.mcp_servers[0].command == "bureau-mcp-server"
    assert "pulls credit reports" in strategy.mcp_servers[0].description
    assert [tool.name for tool in strategy.tools] == ["amortization_calculator"]
    assert "most relevant first" in strategy.skills_discovery
    assert strategy.ontology.meaning_of("risk grade").startswith("a letter from A to E")


def test_strategy_reads_model_from_prose_without_a_slash():
    strategy = Strategy.from_markdown(
        "## Model Selection\nUse Anthropic's claude-haiku-4-5 with the key in MY_SECRET_KEY.\n"
    )
    assert strategy.provider == "anthropic"
    assert strategy.model == "claude-haiku-4-5"
    assert strategy.api_key_env_var == "MY_SECRET_KEY"


def test_strategy_prose_slashes_are_not_mistaken_for_models():
    strategy = Strategy.from_markdown(
        "## Model Selection\nDecide approve/decline offline; no model is configured.\n"
    )
    assert strategy.model == ""
    assert strategy.model_binding() is None


def test_strategy_disables_subagent_spawning_from_prose():
    strategy = Strategy.from_markdown("## Subagent Spawning\nNever spawn subagents in this runtime.\n")
    assert strategy.subagents_configured
    assert strategy.subagents_enabled is False


def test_strategy_narrative_carries_ontology_tools_and_mcp():
    narrative = Strategy.from_markdown(STRATEGY_MARKDOWN).narrative()
    assert "risk grade: a letter from A to E" in narrative
    assert "amortization_calculator" in narrative
    assert "credit_bureau" in narrative
    assert "Discovery guidance" in narrative


def test_loader_applies_strategy_to_memory_spawner_and_store(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = STRATEGY_MARKDOWN
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)

    assert runtime.memory.capacity == 30
    assert runtime.spawner.enabled and runtime.spawner.limit == 4
    assert runtime.session_store is not None
    assert runtime.session_store.path == str(directory / ".ear" / "session.json")
    assert runtime.strategy.ontology.meaning_of("decision") != ""


def test_loader_attaches_model_binding_only_when_credential_resolves(tmp_path, monkeypatch):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Model Selection\nReason with anthropic/claude-opus-4-8; key in FAKE_STACK_KEY.\n"
    directory = write_stack(tmp_path / "stack", **stack)

    monkeypatch.delenv("FAKE_STACK_KEY", raising=False)
    assert load_runtime(directory).model_binding is None

    monkeypatch.setenv("FAKE_STACK_KEY", "not-a-real-key")
    binding = load_runtime(directory).model_binding
    assert binding is not None
    assert binding.model_id == "anthropic/claude-opus-4-8"
    assert binding.api_key_env_var == "FAKE_STACK_KEY"


# ---------------------------------------------------------------------------
# Cross-session data: memory survives into a fresh session.
# ---------------------------------------------------------------------------


def test_session_store_round_trips_memory_experience_and_adaptations(tmp_path):
    store = SessionStore(str(tmp_path / "state" / "session.json"))
    first = Runtime(name="Persistent-Runtime", session_store=store)
    first.reason(Intent(text="Underwrite the first loan", context={"loan_amount": 1000}))
    first.adaptations.learn_from(first.experience)
    store.save(first)

    second = Runtime(name="Persistent-Runtime")
    assert store.restore(second) is True
    assert second.memory.working[0].intent_text == "Underwrite the first loan"
    assert second.memory.working[0].evidence.basis.startswith("Resolved via")
    assert second.experience.observations == 1
    assert len(second.adaptations.impressions) == 1


def test_runtime_saves_session_after_every_cycle(tmp_path):
    path = tmp_path / "session.json"
    runtime = Runtime(name="Saving-Runtime", session_store=SessionStore(str(path)))
    runtime.reason(Intent(text="first cycle"))
    payload = json.loads(path.read_text(encoding="utf-8"))
    assert payload["runtime"] == "Saving-Runtime"
    assert len(payload["memory"]["working"]) == 1


def test_new_session_loaded_from_markdown_picks_up_prior_memory(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Cross-Session Data\nPersist everything to `state.json` between sessions.\n"
    directory = write_stack(tmp_path / "stack", **stack)

    first = load_runtime(directory)
    first.reason(Intent(text="Underwrite the first loan", context={"loan_amount": 1000}))

    second = load_runtime(directory)
    assert len(second.memory.working) == 1
    assert second.memory.working[0].intent_text == "Underwrite the first loan"


def test_session_store_restore_survives_a_corrupt_file(tmp_path):
    path = tmp_path / "session.json"
    path.write_text("not json at all", encoding="utf-8")
    runtime = Runtime(name="Robust-Runtime")
    assert SessionStore(str(path)).restore(runtime) is False
    assert runtime.memory.working == []


# ---------------------------------------------------------------------------
# Markdown as the system-native input and output format.
# ---------------------------------------------------------------------------


def test_intent_reads_from_markdown_with_typed_context():
    intent = Intent.from_markdown(
        "# Underwrite a $18,500 personal loan for Priya Raman\n\n"
        "Requested for a kitchen renovation.\n\n"
        "## Context\n\n"
        "- loan_amount: 18500\n- debt_to_income: 0.28\n- existing_defaults: 0\n"
        "- first_time_borrower: yes\n- purpose: kitchen renovation\n"
    )
    assert intent.text.startswith("Underwrite a $18,500 personal loan")
    assert "kitchen renovation" in intent.text
    assert intent.context["loan_amount"] == 18500
    assert intent.context["debt_to_income"] == 0.28
    assert intent.context["existing_defaults"] == 0
    assert intent.context["first_time_borrower"] is True
    assert intent.context["purpose"] == "kitchen renovation"


def test_intent_round_trips_through_markdown():
    original = Intent(text="Underwrite a loan", context={"loan_amount": 5000, "purpose": "solar panels"})
    recovered = Intent.from_markdown(original.to_markdown())
    assert recovered.text == original.text
    assert recovered.context == original.context


def test_session_store_round_trips_through_markdown(tmp_path):
    store = SessionStore(str(tmp_path / "session.md"))
    first = Runtime(name="Markdown-Persistent-Runtime")
    first.reason(Intent(text="Underwrite the first loan", context={"loan_amount": 1000, "purpose": "a bike"}))
    first.memory.compressed.append("3 earlier cycles (e.g. approved twice)")
    first.adaptations.learn_from(first.experience)
    store.save(first)

    second = Runtime(name="Markdown-Persistent-Runtime")
    assert store.restore(second) is True
    entry = second.memory.working[0]
    assert entry.intent_text == "Underwrite the first loan"
    assert entry.context == {"loan_amount": 1000, "purpose": "a bike"}
    assert entry.decision == str(first.memory.working[0].decision)
    assert entry.evidence.basis.startswith("Resolved via")
    assert second.memory.compressed == ["3 earlier cycles (e.g. approved twice)"]
    assert second.experience.observations == 1
    assert second.experience.decision_counts == first.experience.decision_counts
    assert second.adaptations.impressions[0].insight == first.adaptations.impressions[0].insight


def test_reasoning_log_flushes_markdown_trail(tmp_path):
    path = tmp_path / "reasoning.md"
    runtime = Runtime(name="Markdown-Audited-Runtime")
    runtime.reasoning_log.path = str(path)
    runtime.reason(Intent(text="first cycle"))
    runtime.reason(Intent(text="second cycle"))

    trail = path.read_text(encoding="utf-8")
    assert "## Cycle 1 --" in trail and "## Cycle 2 --" in trail
    assert "### deliberation" in trail and "### explanation" in trail
    # Free text is blockquoted, so it can never be mistaken for structure.
    assert "> " in trail


def test_exchange_answers_intent_documents_with_decision_documents(tmp_path):
    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    (directory / "intents").mkdir()
    (directory / "intents" / "priya-raman.md").write_text(
        "# Underwrite a $18,500 personal loan for Priya Raman\n\n"
        "## Context\n\n- loan_amount: 18500\n- debt_to_income: 0.28\n",
        encoding="utf-8",
    )
    runtime = load_runtime(directory)
    written = Exchange(directory).run(runtime)

    assert [path.name for path in written] == ["priya-raman.md"]
    decision = (directory / "decisions" / "priya-raman.md").read_text(encoding="utf-8")
    assert decision.startswith("# Decision -- Underwrite a $18,500 personal loan for Priya Raman")
    assert "Status: decided" in decision
    assert "## Completion summary" in decision
    assert "- Status: decided" in decision
    assert "Credit Risk Guru" in decision
    assert "## Policy judgments" in decision

    # Idempotent: already-answered intents are not replayed.
    assert Exchange(directory).run(runtime) == []


def test_exchange_completion_summary_reports_tool_outputs(tmp_path):
    directory = tmp_path / "stack"
    directory.mkdir()
    (directory / "intent.md").write_text("# Check a file\n", encoding="utf-8")
    runtime = Runtime(name="Tool-Summary-Runtime")

    def reason(intent, approval=None):
        runtime.reasoning_log.begin_cycle(intent)
        runtime.reasoning_log.record(
            stage="tool",
            inputs={"tool": "read_file", "arguments": {"path": "missing.xlsx"}},
            output="FAILED -- Tool 'read_file' failed: no such file",
        )
        runtime.reasoning_log.record(
            stage="tool",
            inputs={"tool": "run_shell", "arguments": {"command": "python3 probe.py"}},
            output="rows: 907\n[exit 0, 10 ms]",
        )
        runtime.reasoning_log.record(
            stage="summarize",
            inputs={"tool": "run_shell"},
            output="probe.py ran ok, exit 0, observed 907 rows",
        )
        return "Status: validated\n\nDone."

    runtime.reason = reason
    written = Exchange(directory).run(runtime)

    assert [path.name for path in written] == ["decision.md"]
    decision = (directory / "decision.md").read_text(encoding="utf-8")
    assert "## Completion summary" in decision
    assert "- Status: decided (decision reports: validated)" in decision
    assert "probe.py ran ok, exit 0, observed 907 rows" in decision


def test_exchange_writes_blocked_decision_documents(tmp_path):
    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    (directory / "intent.md").write_text(
        "# Underwrite an oversized loan\n\n## Context\n\n- loan_amount: 90000\n",
        encoding="utf-8",
    )
    runtime = load_runtime(directory)
    written = Exchange(directory).run(runtime)

    assert [path.name for path in written] == ["decision.md"]
    decision = (directory / "decision.md").read_text(encoding="utf-8")
    assert "Status: BLOCKED" in decision
    assert "Loan Amount Cap" in decision
    assert "VIOLATED" in decision


def test_exchange_turns_a_provider_failure_into_a_blocked_decision(tmp_path, monkeypatch):
    """An infrastructure stop -- the provider call itself failing (billing,
    auth, network) -- lands as a BLOCKED decision document, exactly like a
    governance stop. The live failure this closes: a billing 400 mid-cycle
    raised straight through Exchange.run, crashing the driver and erasing
    the whole cycle from the exchange -- every tool call the model had
    already made simply vanished."""
    from ear.llm import LMError

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    (directory / "intent.md").write_text("# Underwrite a loan\n\n## Context\n\n- loan_amount: 100\n", encoding="utf-8")
    runtime = load_runtime(directory)

    def fail(intent, approval=None):
        raise LMError("LLM call to https://api.anthropic.com/v1/messages failed (400): credit balance too low")

    monkeypatch.setattr(runtime, "reason", fail)
    written = Exchange(directory).run(runtime)

    assert [path.name for path in written] == ["decision.md"]
    decision = (directory / "decision.md").read_text(encoding="utf-8")
    assert "Status: BLOCKED" in decision
    assert "credit balance too low" in decision


def test_loader_defaults_session_and_audit_to_markdown(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = (
        "## Cross-Session Data\nPersist everything between sessions.\n\n"
        "## Reasoning Audit Trail\nLog every reasoning step for review.\n"
    )
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    assert runtime.session_store.path == str(directory / ".ear" / "session.md")
    assert runtime.reasoning_log.path == str(directory / ".ear" / "reasoning.md")

    runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))
    second = load_runtime(directory)
    assert len(second.memory.working) == 1


def test_cycle_numbering_continues_across_sessions_in_one_trail(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Reasoning Audit Trail\nLog every reasoning step for review.\n"
    directory = write_stack(tmp_path / "stack", **stack)

    first = load_runtime(directory)
    first.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))

    second = load_runtime(directory)
    second.reason(Intent(text="Underwrite another loan", context={"loan_amount": 6000}))

    trail = (directory / ".ear" / "reasoning.md").read_text(encoding="utf-8")
    assert "## Cycle 1 --" in trail and "## Cycle 2 --" in trail
    assert second.reasoning_log.cycle == 2


def test_declared_path_ignores_prose_mentions_of_stack_files():
    strategy = Strategy.from_markdown(
        "## Cross-Session Data\nAs declared here in memory.md, persist everything to .ear/session.md between sessions.\n"
    )
    assert strategy.session_path == ".ear/session.md"


# ---------------------------------------------------------------------------
# The reasoning audit trail.
# ---------------------------------------------------------------------------


def test_reasoning_log_records_every_stage_of_a_cycle(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))

    stages = [record.stage for record in runtime.reasoning_log.records]
    assert stages[0] == "intent"
    assert "policy" in stages and "discovery" in stages
    assert "deliberation" in stages and "explanation" in stages

    deliberation = runtime.reasoning_log.for_stage("deliberation")[0]
    # The full stacked prompt material is on the record for prompt review.
    assert "Credit Risk Guru" in deliberation.inputs["capabilities"]
    assert "risk_grade" in deliberation.inputs["capabilities"]
    assert deliberation.model == "deterministic-fallback"

    policy = runtime.reasoning_log.for_stage("policy")[0]
    assert policy.output == "complies"
    assert policy.rationale  # the "why" is never dropped


def test_reasoning_log_records_blocked_cycles_too(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    with pytest.raises(PermissionError):
        runtime.reason(Intent(text="Underwrite a consumer loan", context={"debt_to_income": 0.60}))
    violated = [record for record in runtime.reasoning_log.for_stage("policy") if record.output == "VIOLATED"]
    assert [record.inputs["policy"] for record in violated] == ["DTI Ceiling"]
    assert "0.43" in violated[0].rationale


def test_reasoning_log_flushes_jsonl_after_each_cycle(tmp_path):
    path = tmp_path / "trail" / "reasoning.jsonl"
    runtime = Runtime(name="Audited-Runtime")
    runtime.reasoning_log.path = str(path)
    runtime.reason(Intent(text="first cycle"))
    runtime.reason(Intent(text="second cycle"))

    lines = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]
    assert {line["cycle"] for line in lines} == {1, 2}
    assert {line["stage"] for line in lines} >= {"intent", "discovery", "deliberation", "explanation"}
    # Append-only: nothing was double-written.
    assert len(lines) == len(runtime.reasoning_log.records)


def test_reasoning_log_render_groups_by_cycle():
    runtime = Runtime(name="Rendered-Runtime")
    runtime.reason(Intent(text="one cycle"))
    rendered = runtime.reasoning_log.render()
    assert "=== Cycle 1" in rendered
    assert "[deliberation]" in rendered


def test_strategy_declares_the_audit_trail_in_natural_language():
    strategy = Strategy.from_markdown(
        "## Reasoning Audit Trail\nLog every reasoning step to `audit/trail.jsonl` for review.\n"
    )
    assert strategy.audit_enabled
    assert strategy.audit_path == "audit/trail.jsonl"


def test_loader_points_the_reasoning_log_at_the_declared_path(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Reasoning Audit Trail\nLog every reasoning step to `.ear/reasoning.jsonl`.\n"
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    assert runtime.reasoning_log.path == str(directory / ".ear" / "reasoning.jsonl")

    runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))
    assert (directory / ".ear" / "reasoning.jsonl").exists()


def test_policy_judge_returns_the_rationale_offline():
    from ear import Policy

    policy = Policy(name="Cap", fallback_expression="amount <= 10")
    complies, rationale = policy.judge(amount=5)
    assert complies is True and "amount <= 10" in rationale
    complies, rationale = policy.judge(amount=50)
    assert complies is False


# ---------------------------------------------------------------------------
# Contracts: a workflow's Deliverable, declared in natural language,
# extracted and judged by the model at runtime, delivered as ## Data.
# ---------------------------------------------------------------------------

CONTRACT_WORKFLOW = (
    "# Workflows\n\n## Underwriting Workflow\n\n"
    "1. Band the profile and assign a risk grade. (Credit Risk Guru)\n"
    "2. Decide approve or decline against the grade. (Credit Risk Guru)\n\n"
    "Policies: Loan Amount Cap\n\n"
    "### Deliverable\n\nThe decision as structured facts.\n\n"
    "- decision: exactly one of approve or decline\n"
    "- risk grade: the letter grade from A to E\n"
)


def test_deliverable_section_becomes_the_workflow_contract(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = CONTRACT_WORKFLOW
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))

    workflow = runtime.processes[0].workflows[0]
    contract = workflow.contract
    assert contract is not None
    assert contract.name == "Underwriting Workflow Deliverable"
    assert [f.name for f in contract.fields] == ["decision", "risk grade"]
    assert contract.fields[1].identifier == "risk_grade"
    assert "structured facts" in contract.description
    # The Deliverable section is the workflow's contract, never a workflow
    # or a process of its own.
    assert [process.name for process in runtime.processes] == ["Underwrite Consumer Loan"]
    assert len(workflow.steps) == 2


def test_deliverable_with_no_workflow_above_fails_loudly(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = "# Workflows\n\n### Deliverable\n\n- decision: approve or decline\n"
    with pytest.raises(ValueError, match="no workflow above"):
        load_runtime(write_stack(tmp_path / "stack", **stack))


def test_deliverable_field_without_a_meaning_fails_loudly(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = CONTRACT_WORKFLOW.replace("- decision: exactly one of approve or decline", "- decision")
    with pytest.raises(ValueError, match="must be written as 'name: meaning'"):
        load_runtime(write_stack(tmp_path / "stack", **stack))


def test_contract_judge_falls_back_to_structural_presence():
    from ear import Contract

    contract = Contract(name="Deliverable").add_field("decision", "approve or decline")
    conforms, rationale = contract.judge({"decision": "approve"})
    assert conforms is True and "structural" in rationale
    conforms, rationale = contract.judge({"decision": "  "})
    assert conforms is False and "decision" in rationale


def test_offline_contract_extraction_is_skipped_on_the_record(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = CONTRACT_WORKFLOW
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))

    record = runtime.reasoning_log.for_stage("contract")[0]
    assert "skipped" in record.output and "no model" in record.output
    # No fabricated data reaches the evidence.
    assert "data" not in runtime.memory.working[-1].evidence.sources


def test_data_section_round_trips_typed_values():
    from ear.exchange import data_from_decision_document

    document = (
        "# Decision -- x\n\n## Data\n\n"
        "- decision: approve\n- risk grade: B\n- monthly_payment: 372.5\n- defaults: 0\n"
    )
    data = data_from_decision_document(document)
    assert data == {"decision": "approve", "risk grade": "B", "monthly_payment": 372.5, "defaults": 0}


# ---------------------------------------------------------------------------
# The Examiner: markdown-native evaluation with honest offline grading.
# ---------------------------------------------------------------------------


def test_intent_from_markdown_skips_named_sections():
    intent = Intent.from_markdown(
        "# Underwrite a loan\n\n## Expected\n\nIt should be approved.\n\n## Context\n\n- loan_amount: 1000\n",
        skip_sections=("expected",),
    )
    assert "approved" not in intent.text
    assert intent.context == {"loan_amount": 1000}


def test_examiner_grades_structurally_offline_and_reports(tmp_path):
    from ear import Examiner

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    evaluations = directory / "evaluations"
    evaluations.mkdir()
    (evaluations / "resolves.md").write_text(
        "# Underwrite a small consumer loan\n\n## Context\n\n- loan_amount: 5000\n\n"
        "## Expected\n\n- decision: resolved\n",
        encoding="utf-8",
    )
    (evaluations / "blocked.md").write_text(
        "# Underwrite an oversized loan\n\n## Context\n\n- loan_amount: 90000\n\n"
        "## Expected\n\n- decision: Loan Amount Cap\n",
        encoding="utf-8",
    )
    (evaluations / "prose-only.md").write_text(
        "# Underwrite anything\n\n## Expected\n\nThe decision must be conservative.\n",
        encoding="utf-8",
    )

    runtime = load_runtime(directory)
    examination = Examiner().examine(runtime, evaluations)

    assert examination.counts() == {
        "examined": 3,
        "passed": 2,
        "failed": 0,
        "ungraded": 1,
        "criteria": 0,
        "criteria_passed": 0,
        "criteria_failed": 0,
    }
    assert examination.passed is True
    ungraded = next(result for result in examination.results if result.passed is None)
    assert "need a model" in ungraded.rationale

    report = (evaluations / "report.md").read_text(encoding="utf-8")
    assert "Passed: 2" in report and "Ungraded: 1" in report
    assert [record.output for record in runtime.reasoning_log.for_stage("evaluation")] == [
        "passed",
        "ungraded",
        "passed",
    ]


def test_examiner_failed_expectation_fails_the_suite(tmp_path):
    from ear import Examiner

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    evaluations = directory / "evaluations"
    evaluations.mkdir()
    (evaluations / "impossible.md").write_text(
        "# Underwrite a loan\n\n## Context\n\n- loan_amount: 5000\n\n"
        "## Expected\n\n- decision: an outcome that cannot possibly appear\n",
        encoding="utf-8",
    )
    examination = Examiner().examine(load_runtime(directory), evaluations)
    assert examination.passed is False
    assert examination.results[0].verdict == "FAILED"


# ---------------------------------------------------------------------------
# Optimizer: the trail as the training corpus, reviews as the labels.
# ---------------------------------------------------------------------------


def test_optimizer_builds_trainset_from_jsonl_trail(tmp_path):
    from ear import Optimizer

    runtime = Runtime(name="Trained-Runtime")
    runtime.reasoning_log.path = str(tmp_path / "trail.jsonl")
    runtime.reason(Intent(text="first cycle", context={"loan_amount": 100}))
    runtime.reason(Intent(text="second cycle"))

    trainset = Optimizer().trainset_from_trail(runtime.reasoning_log.path)
    assert len(trainset) == 2
    assert trainset[0].intent == "first cycle"
    assert trainset[0].decision.startswith("[Trained-Runtime]")


def test_optimizer_builds_trainset_from_markdown_trail(tmp_path):
    from ear import Optimizer

    runtime = Runtime(name="MD-Runtime")
    runtime.reasoning_log.path = str(tmp_path / "trail.md")
    runtime.reason(Intent(text="first cycle"))

    trainset = Optimizer().trainset_from_trail(runtime.reasoning_log.path)
    assert len(trainset) == 1
    assert trainset[0].intent == "first cycle"
    assert trainset[0].decision.startswith("[MD-Runtime]")


def test_optimizer_reads_reviewer_verdicts_and_excludes_the_unlabelled(tmp_path):
    from ear import Optimizer

    decisions = tmp_path / "decisions"
    decisions.mkdir()
    (decisions / "reviewed.md").write_text(
        "# Decision -- Underwrite a loan\n\n## Intent\n\n> Underwrite a loan\n\n"
        "## Decision\n\n> Approved at grade B.\n\n"
        "## Review\n\nVerdict: correct\n\n> Well reasoned and within policy.\n",
        encoding="utf-8",
    )
    (decisions / "unreviewed.md").write_text(
        "# Decision -- Another loan\n\n## Decision\n\n> Declined.\n",
        encoding="utf-8",
    )
    (decisions / "ambiguous.md").write_text(
        "# Decision -- A third loan\n\n## Decision\n\n> Approved.\n\n## Review\n\nVerdict: maybe\n",
        encoding="utf-8",
    )

    labelled = Optimizer().verdicts_from_documents(decisions)
    assert len(labelled) == 1
    assert labelled[0].verdict is True
    assert labelled[0].intent == "Underwrite a loan"
    assert labelled[0].decision == "Approved at grade B."
    assert "Well reasoned" in labelled[0].note


def test_optimizer_metric_offline_uses_normalized_containment():
    from ear import Optimizer

    metric = Optimizer().metric(None)
    assert metric("approve the loan", "I would approve the loan today.") == 1.0
    assert metric("approve the loan", "Declined outright.") == 0.0


def test_optimizer_refine_is_a_no_op_without_a_model():
    from ear import Optimizer
    from ear.optimizer import Example
    from ear.signatures import ReasonAboutIntent

    before = ReasonAboutIntent.instruction
    result = Optimizer().refine(ReasonAboutIntent, [Example(intent="x", decision="y")], model_binding=None)
    assert result == before  # reflection is a judgment; no model, no change


def test_native_judgment_parses_markdown_sections_from_a_stub_lm():
    from ear.judgment import Field, Judgment

    class StubLM:
        def complete(self, prompt, system=""):
            return "## complies\n\nyes\n\n## rationale\n\nThe amount is within the limit.\n"

    judgment = Judgment(
        instruction="Judge it.",
        inputs=[Field("context")],
        outputs=[Field("complies", kind="bool"), Field("rationale", kind="text")],
    )
    result = judgment.run(StubLM(), context="amount 5")
    assert result.complies is True
    assert "within the limit" in result.rationale


# ---------------------------------------------------------------------------
# Live: contracts extracted and judged, evaluations graded, by a real model.
# ---------------------------------------------------------------------------

LIVE_MEMORY = (
    "# Memory & Strategy\n\n## Model Selection\n\n"
    "Reason with anthropic/claude-haiku-4-5, reading the credential from\n"
    "ANTHROPIC_API_KEY, at a temperature of 0.2.\n"
)


@requires_anthropic_key
def test_contract_extraction_delivers_typed_data_live(tmp_path):
    from ear.exchange import data_from_decision_document

    stack = dict(MINIMAL_STACK)
    stack["workflow"] = CONTRACT_WORKFLOW
    stack["memory"] = LIVE_MEMORY
    directory = write_stack(tmp_path / "stack", **stack)
    (directory / "intent.md").write_text(
        "# Underwrite a $12,000 personal loan for a strong applicant\n\n"
        "## Context\n\n- loan_amount: 12000\n- credit_score: 785\n- debt_to_income: 0.2\n"
        "- existing_defaults: 0\n",
        encoding="utf-8",
    )

    runtime = load_runtime(directory)
    Exchange(directory).run(runtime)

    record = runtime.reasoning_log.for_stage("contract")[-1]
    assert record.output == "conformant"
    data = data_from_decision_document((directory / "decision.md").read_text(encoding="utf-8"))
    assert "approve" in str(data.get("decision", "")).lower()
    assert data.get("risk grade")


@requires_anthropic_key
def test_examiner_grades_the_example_suite_live(tmp_path, monkeypatch):
    from ear import Examiner

    directory = tmp_path / "credit_risk_stack"
    shutil.copytree(EXAMPLE_STACK, directory)
    memory = (directory / "memory.md").read_text(encoding="utf-8")
    (directory / "memory.md").write_text(
        memory.replace("anthropic/claude-opus-4-8", "anthropic/claude-haiku-4-5"), encoding="utf-8"
    )

    runtime = load_runtime(directory)
    examination = Examiner().examine(runtime, directory / "evaluations")

    assert examination.counts()["ungraded"] == 0
    assert examination.passed is True
    assert (directory / "evaluations" / "report.md").exists()


# ---------------------------------------------------------------------------
# Panels: multi-persona deliberation, native -- the pattern is prose in
# workflow.md, turns and synthesis land on the trail, budgets are code.
# ---------------------------------------------------------------------------

PANEL_STACK = dict(
    MINIMAL_STACK,
    persona=(
        "# Personas\n\n## Credit Risk Guru\nUnderwrite conservatively.\n\n"
        "Skills: risk_grade\n\n"
        "## Customer Advocate\nArgue the applicant's side plainly.\n\nSkills: decide\n"
    ),
    workflow=(
        "# Workflows\n\n## Underwriting Workflow\n\n"
        "Pattern: adversarial debate; the Credit Risk Guru has the last word\n\n"
        "1. Assess the risk. (Credit Risk Guru)\n"
        "2. Make the applicant's case. (Customer Advocate)\n"
    ),
)


def test_pattern_is_authored_as_prose_on_the_workflow(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **PANEL_STACK))
    workflow = runtime.processes[0].workflows[0]
    assert workflow.pattern == "adversarial debate; the Credit Risk Guru has the last word"


def test_patterned_workflow_convenes_a_panel_offline(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **PANEL_STACK))
    decision = runtime.reason(Intent(text="Underwrite a marginal loan", context={"loan_amount": 5000}))

    turns = runtime.reasoning_log.for_stage("conversation")
    assert len(turns) == 4  # two rounds around a two-persona table
    assert {record.inputs["speaker"] for record in turns} == {"Credit Risk Guru", "Customer Advocate"}
    assert all("no model bound" in record.output for record in turns)

    assert "Credit Risk Guru" in decision and "Customer Advocate" in decision
    assert "no model bound" in decision  # the offline panel never fakes a judgment
    deliberation = runtime.reasoning_log.for_stage("deliberation")[0]
    assert deliberation.inputs["style"] == "adversarial debate; the Credit Risk Guru has the last word"
    assert "[Credit Risk Guru]" in deliberation.inputs["transcript"]


def test_pattern_with_a_single_persona_stays_single_voiced(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = (
        "# Workflows\n\n## Underwriting Workflow\n\nPattern: debate\n\n"
        "1. Decide approve or decline. (Credit Risk Guru)\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    runtime.reason(Intent(text="Underwrite a loan", context={"loan_amount": 5000}))
    assert runtime.reasoning_log.for_stage("conversation") == []


def test_panel_budget_is_enforced_in_code():
    from ear import Panel

    runtime = Runtime(name="Budgeted-Runtime")
    personas = [Persona(name="A"), Persona(name="B")]
    Panel(rounds=50, max_turns=5).convene(runtime, personas, Intent(text="anything"))
    assert len(runtime.reasoning_log.for_stage("conversation")) == 5


@requires_anthropic_key
def test_panel_debates_live(tmp_path):
    stack = dict(PANEL_STACK)
    stack["memory"] = LIVE_MEMORY
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    decision = runtime.reason(
        Intent(
            text="Underwrite a $9,000 personal loan for a marginal applicant",
            context={"loan_amount": 9000, "credit_score": 668, "debt_to_income": 0.41},
        )
    )
    turns = runtime.reasoning_log.for_stage("conversation")
    # Speakers are judged now, and consensus may conclude the panel before
    # the budget of four -- but never before both sides have spoken.
    assert 2 <= len(turns) <= 4
    assert {record.inputs["speaker"] for record in turns} == {"Credit Risk Guru", "Customer Advocate"}
    assert all(record.model == "anthropic/claude-haiku-4-5" for record in turns)
    assert decision and "no model bound" not in decision


# ---------------------------------------------------------------------------
# Journeys: durable, resumable, step-wise execution -- every leg a full
# governed cycle, the state a markdown record, native.
# ---------------------------------------------------------------------------


def test_journey_completes_and_is_settled(tmp_path):
    from ear import Journey

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    runtime = load_runtime(directory)
    journey = Journey(directory / "journeys" / "loan.md")

    status = journey.run(runtime, Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))
    assert status == "completed"
    assert [leg.status for leg in journey.legs] == ["decided", "decided"]
    assert "Credit Risk Runtime" in journey.decision
    record = (directory / "journeys" / "loan.md").read_text(encoding="utf-8")
    assert "Status: completed" in record and "## Leg 2" in record
    assert runtime.reasoning_log.cycle == 2  # every leg was a full governed cycle

    # Settled: running again replays nothing.
    again = Journey(directory / "journeys" / "loan.md")
    assert again.run(runtime) == "completed"
    assert runtime.reasoning_log.cycle == 2


def test_journey_survives_a_crash_and_resumes_where_the_record_ends(tmp_path):
    from ear import Journey

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    runtime = load_runtime(directory)
    calls = {"count": 0}
    orchestrate = runtime.orchestrator.orchestrate

    def crash_on_second(*args, **kwargs):
        calls["count"] += 1
        if calls["count"] == 2:
            raise RuntimeError("worker died mid-journey")
        return orchestrate(*args, **kwargs)

    runtime.orchestrator.orchestrate = crash_on_second
    journey = Journey(directory / "journeys" / "loan.md")
    with pytest.raises(RuntimeError, match="worker died"):
        journey.run(runtime, Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))

    record = (directory / "journeys" / "loan.md").read_text(encoding="utf-8")
    assert "Status: in progress" in record and "## Leg 1" in record and "## Leg 2" not in record

    # A fresh runtime resumes exactly where the record ends.
    recovered = load_runtime(directory)
    resumed = Journey(directory / "journeys" / "loan.md")
    assert resumed.run(recovered) == "completed"
    assert recovered.reasoning_log.cycle == 1  # only the lost leg was walked
    assert resumed.legs[0].decision  # leg one's decision came from the record, not a replay


def test_journey_blocks_hard_and_parks_for_approval(tmp_path):
    from ear import Approval, Journey

    directory = approval_stack(tmp_path)
    runtime = load_runtime(directory)
    journey = Journey(directory / "journeys" / "big-loan.md")
    status = journey.run(runtime, Intent(text="Underwrite a large loan", context={"loan_amount": 60000}))
    assert status == "PENDING APPROVAL"
    cycles_when_parked = runtime.reasoning_log.cycle

    # Without a verdict the journey stays parked and replays nothing.
    assert Journey(journey.path).run(runtime) == "PENDING APPROVAL"
    assert runtime.reasoning_log.cycle == cycles_when_parked

    released = Journey(journey.path)
    assert released.run(runtime, approval=Approval(verdict=True, approver="lakshmi@gigkri.com")) == "completed"
    assert "approved by lakshmi@gigkri.com" in [r.output for r in runtime.reasoning_log.for_stage("approval")]

    blocked = Journey(directory / "journeys" / "huge-loan.md")
    assert blocked.run(runtime, Intent(text="Underwrite a huge loan", context={"loan_amount": 90000})) == "BLOCKED"
    assert "Status: BLOCKED" in blocked.path.read_text(encoding="utf-8")


def test_journey_refuses_to_resume_over_a_changed_stack(tmp_path):
    from ear import Journey

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    runtime = load_runtime(directory)
    journey = Journey(directory / "journeys" / "loan.md")
    journey.run(runtime, Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))

    changed = (directory / "workflow.md").read_text(encoding="utf-8").replace(
        "Band the profile and assign a risk grade.", "Do something entirely different."
    )
    (directory / "workflow.md").write_text(changed, encoding="utf-8")
    with pytest.raises(ValueError, match="forge the record"):
        Journey(journey.path).run(load_runtime(directory))


def test_journey_first_run_needs_an_intent_and_a_stack(tmp_path):
    from ear import Journey

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    runtime = load_runtime(directory)
    with pytest.raises(ValueError, match="needs an intent"):
        Journey(directory / "journeys" / "empty.md").run(runtime)
    with pytest.raises(ValueError, match="no workflow steps"):
        Journey(tmp_path / "nowhere.md").run(Runtime(name="Empty-Runtime"), Intent(text="x"))



# ---------------------------------------------------------------------------
# Executable tools: declarations meet handlers in the ToolBinder, the model
# decides when to call, and every invocation lands on the trail.
# ---------------------------------------------------------------------------

TOOLS_MEMORY = (
    "# Memory & Strategy\n\n## Tools\n\n"
    "- amortization_calculator: computes the monthly payment for an amount, an annual rate in percent, and a term in months\n"
)


def monthly_payment(amount: float, annual_rate_percent: float, months: int) -> float:
    """Compute a fixed monthly payment for a loan."""
    monthly_rate = annual_rate_percent / 100 / 12
    if monthly_rate == 0:
        return round(amount / months, 2)
    factor = (1 + monthly_rate) ** months
    return round(amount * monthly_rate * factor / (factor - 1), 2)


def test_tool_binder_resolves_declared_tools_and_rejects_strays(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = TOOLS_MEMORY
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    runtime.bind_tool("amortization_calculator", monthly_payment)

    bound = {tool.name: tool for tool in runtime.tool_binder.bound_tools(runtime)}
    assert "monthly payment" in bound["amortization_calculator"].description  # the declared description, not the docstring
    assert bound["amortization_calculator"].identifier == "amortization_calculator"
    # the runtime's own basic tools (list/view/create/retire) ride along by default
    assert {"list_tools", "view_tool", "create_tool", "retire_tool"} <= bound.keys()

    runtime.bind_tool("undeclared_gadget", monthly_payment)
    with pytest.raises(ValueError, match="matches nothing the stack declares.*amortization_calculator"):
        runtime.tool_binder.bound_tools(runtime)


def test_handler_skills_bind_automatically_and_bindings_override():
    from ear import Persona, Skill, ToolBinder, Workflow

    persona = Persona(name="Analyst")
    persona.add_skill(Skill(name="fetch_score", prompt="Fetch the credit score.", handler=lambda applicant: 700))
    workflow = Workflow(name="W")
    workflow.add_persona(persona)
    runtime = Runtime(name="Skilled-Runtime")

    binder = ToolBinder()
    bound = binder.bound_tools(runtime, plan=[workflow])
    assert [tool.name for tool in bound] == ["fetch_score"]
    assert bound[0].handler("anyone") == 700

    # An explicit binding for the same skill overrides its own handler;
    # explicit bindings resolve against the runtime's whole stack, so the
    # workflow joins a process first.
    from ear import Process

    holder = Process(name="P")
    holder.add_workflow(workflow)
    runtime.add_process(holder)
    binder.bind("fetch_score", lambda applicant: 750)
    bound = binder.bound_tools(runtime, plan=[workflow])
    assert bound[0].handler("anyone") == 750


def test_offline_deliberation_lists_tools_without_invoking_them(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = TOOLS_MEMORY
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    runtime.bind_tool("amortization_calculator", monthly_payment)
    runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))

    deliberation = runtime.reasoning_log.for_stage("deliberation")[0]
    assert "amortization_calculator" in deliberation.inputs["tools"]
    assert runtime.reasoning_log.for_stage("tool") == []


def test_logged_tool_wrapper_contains_failures_on_the_record():
    from ear.tool_binder import BoundTool, ToolBinder

    def broken(query: str) -> str:
        raise RuntimeError("bureau offline")

    runtime = Runtime(name="Contained-Runtime")
    wrapped = ToolBinder._logged(runtime, BoundTool(name="credit_bureau_check", description="checks", handler=broken))
    result = wrapped("anything")
    assert "failed: bureau offline" in str(result)
    record = runtime.reasoning_log.for_stage("tool")[0]
    assert record.output.startswith("FAILED")
    assert record.inputs["tool"] == "credit_bureau_check"
    assert "duration_ms" in record.inputs


# ---------------------------------------------------------------------------
# Acquirer: a runtime lists, views, and grows its own toolset, natively.
# ---------------------------------------------------------------------------


def test_acquirer_lists_and_views_declared_tools_and_skills(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = TOOLS_MEMORY
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))

    listing = runtime.acquirer.list_tools(runtime)
    assert "amortization_calculator (authored, declared)" in listing
    assert "risk_grade (skill, prompt-only)" in listing

    view = runtime.acquirer.view_tool(runtime, "amortization_calculator")
    assert "Origin: authored" in view
    assert "computes the monthly payment" in view

    assert runtime.acquirer.view_tool(runtime, "nope") == "No tool or skill named 'nope' is declared."


def test_acquirer_create_tool_declares_and_persists(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = TOOLS_MEMORY
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)

    result = runtime.acquirer.create_tool(runtime, "fetch_rate_sheet", "Look up today's published rate sheet.")
    assert "Declared tool 'fetch_rate_sheet'" in result
    assert any(tool.name == "fetch_rate_sheet" for tool in runtime.strategy.tools)

    persisted = (directory / ".ear" / "tools.md").read_text(encoding="utf-8")
    assert "## fetch_rate_sheet" in persisted
    assert "Status: active" in persisted

    # A duplicate name is refused rather than shadowing the first.
    refusal = runtime.acquirer.create_tool(runtime, "fetch_rate_sheet", "Something else entirely.")
    assert "already declared" in refusal

    # It is context only until a handler binds it.
    bound_names = {tool.name for tool in runtime.tool_binder.bound_tools(runtime)}
    assert "fetch_rate_sheet" not in bound_names


def test_a_tool_declared_only_in_tools_md_is_unreachable_by_the_native_tool_loop(tmp_path):
    """The invariant `create_tool` promises in its own return message
    ('context to the model until a handler binds it') held end to end: a
    live deliberation naming a tools.md-only entry gets the same
    hallucinated-tool recovery a made-up name would, never an execution --
    because `bound_tools()` never auto-promotes a declared Tool into the
    executable set, no matter how it was declared (memory.md or
    `.ear/tools.md`)."""
    from ear.reasoner import Reasoner

    stack = dict(MINIMAL_STACK)
    stack["memory"] = TOOLS_MEMORY
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    runtime.acquirer.create_tool(runtime, "fetch_rate_sheet", "Look up today's rate sheet.")

    bound = runtime.tool_binder.bound_tools(runtime)
    assert "fetch_rate_sheet" not in {tool.name for tool in bound}  # declared, never bound

    lm = ScriptedLM(
        [
            _tool_action(tool="fetch_rate_sheet", args="- date: today"),  # the acquired-but-unbound name
            _tool_action(decision="Approved without the rate sheet."),
        ]
    )
    decision = Reasoner._reason_with_tools(
        Intent(text="Underwrite a consumer loan"), runtime, lm, context={}, capabilities="none", tools=bound, max_iterations=4
    )
    assert decision == "Approved without the rate sheet."

    records = runtime.reasoning_log.for_stage("tool")
    assert records[0].inputs["recovery"] is True
    assert records[0].output.startswith("RECOVERED")
    assert "no tool named 'fetch_rate_sheet'" in records[0].output
    # Never actually invoked -- there is no non-recovery record for it.
    assert not any(not record.inputs.get("recovery") and record.inputs.get("tool") == "fetch_rate_sheet" for record in records)


def test_acquirer_retire_tool_refuses_authored_and_retires_acquired(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = TOOLS_MEMORY
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)

    refusal = runtime.acquirer.retire_tool(runtime, "amortization_calculator")
    assert "authored in memory.md" in refusal
    assert any(tool.name == "amortization_calculator" for tool in runtime.strategy.tools)

    runtime.acquirer.create_tool(runtime, "scratch_tool", "A throwaway tool.")
    result = runtime.acquirer.retire_tool(runtime, "scratch_tool", reason="no longer needed")
    assert result == "Retired tool 'scratch_tool'."
    assert not any(tool.name == "scratch_tool" for tool in runtime.strategy.tools)

    persisted = (directory / ".ear" / "tools.md").read_text(encoding="utf-8")
    assert "Status: retired -- no longer needed" in persisted

    assert runtime.acquirer.retire_tool(runtime, "not_a_tool") == "No tool named 'not_a_tool' is declared."


def test_load_runtime_merges_acquired_tools_and_memory_md_wins_on_clash(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = TOOLS_MEMORY
    directory = write_stack(tmp_path / "stack", **stack)
    (directory / ".ear").mkdir()
    (directory / ".ear" / "tools.md").write_text(
        "# Acquired Tools\n\n"
        "## fetch_rate_sheet\n\nDescription: Look up today's published rate sheet.\nStatus: active\n\n"
        "## retired_gadget\n\nDescription: An old one.\nStatus: retired -- superseded\n\n"
        "## amortization_calculator\n\nDescription: A shadow that must lose to memory.md.\nStatus: active\n",
        encoding="utf-8",
    )

    runtime = load_runtime(directory)
    names = {tool.name: tool for tool in runtime.strategy.tools}
    assert "fetch_rate_sheet" in names
    assert names["fetch_rate_sheet"].origin == "acquired"
    assert "retired_gadget" not in names
    # memory.md's own declaration is never shadowed by the acquired file.
    assert names["amortization_calculator"].origin == "authored"
    assert "computes the monthly payment" in names["amortization_calculator"].description


def test_tool_acquisition_can_be_disabled_from_prose(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = (
        "# Memory & Strategy\n\n## Tools\n\n"
        "- amortization_calculator: computes the monthly payment\n\n"
        "This is a fixed toolset -- never create new tools.\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    assert runtime.strategy.tool_acquisition is False
    bound_names = {tool.name for tool in runtime.tool_binder.bound_tools(runtime)}
    assert bound_names.isdisjoint({"list_tools", "view_tool", "create_tool", "retire_tool"})


def test_acquirer_exposes_itself_as_bound_tools_in_a_live_toolloop():
    runtime = Runtime(name="Self-Extending-Runtime")
    tools = runtime.acquirer.as_tools(runtime)
    names = {tool.name for tool in tools}
    assert names == {"list_tools", "view_tool", "create_tool", "retire_tool"}
    assert "no tools or skills declared" in tools[0].handler()


SANDBOXED_TOOLS_MEMORY = (
    "# Memory & Strategy\n\n## Sandbox\n\nEach runtime runs in an isolated workspace "
    "under `.ear/box`. Expose a shell.\n\n## Tools\n\n"
    "- amortization_calculator: computes the monthly payment\n"
)


def test_acquirer_confines_a_created_tool_inside_the_sandbox(tmp_path):
    """The whole point of running an agent in a k8s pod: its blast radius
    stays inside the Sandbox. A self-declared tool's own file must land
    there too, never at the stack-level `.ear/tools.md` a plain host path
    would name outside the box."""
    directory = write_stack(tmp_path / "stack", **{**MINIMAL_STACK, "memory": SANDBOXED_TOOLS_MEMORY})
    runtime = load_runtime(directory)
    assert runtime.sandbox is not None

    result = runtime.acquirer.create_tool(runtime, "fetch_rate_sheet", "Look up today's rate sheet.")
    assert "inside the sandbox" in result

    assert runtime.sandbox.exists(".ear/tools.md")
    assert "fetch_rate_sheet" in runtime.sandbox.read_text(".ear/tools.md")
    # Never leaked to the stack-level file outside the box.
    assert not (directory / ".ear" / "tools.md").exists()


def test_load_runtime_reloads_a_sandboxed_acquired_tool_from_inside_the_box(tmp_path):
    directory = write_stack(tmp_path / "stack", **{**MINIMAL_STACK, "memory": SANDBOXED_TOOLS_MEMORY})
    first = load_runtime(directory)
    first.acquirer.create_tool(first, "fetch_rate_sheet", "Look up today's rate sheet.")

    second = load_runtime(directory)
    assert any(tool.name == "fetch_rate_sheet" for tool in second.strategy.tools)


def test_acquirer_retire_tool_writes_the_rotation_note_inside_the_sandbox(tmp_path):
    directory = write_stack(tmp_path / "stack", **{**MINIMAL_STACK, "memory": SANDBOXED_TOOLS_MEMORY})
    runtime = load_runtime(directory)
    runtime.acquirer.create_tool(runtime, "scratch_tool", "A throwaway tool.")

    result = runtime.acquirer.retire_tool(runtime, "scratch_tool", reason="no longer needed")
    assert result == "Retired tool 'scratch_tool'."
    assert "Status: retired -- no longer needed" in runtime.sandbox.read_text(".ear/tools.md")
    assert not (directory / ".ear" / "tools.md").exists()


def test_discover_then_acquire_a_new_language_recipe_with_no_hardcoded_table(tmp_path):
    """No per-language table lives anywhere in EAR's own code: running an
    unfamiliar toolchain is the model's own reasoning over `run_shell`'s
    raw stdout/stderr/exit-code (`SandboxResult.render()`), not a canned
    lookup. Once a working recipe is found, `create_tool` persists it --
    tool acquisition is a runtime, one-time activity, never a static dump
    shipped in code. Persisting is still not binding: the recipe comes
    back as declared context (`view_tool`), replayed through the same
    generic `run_shell` next time, never auto-executed."""
    from ear.reasoner import Reasoner

    directory = write_stack(tmp_path / "stack", **{**MINIMAL_STACK, "memory": SANDBOXED_TOOLS_MEMORY})
    runtime = load_runtime(directory)
    bound = runtime.tool_binder.bound_tools(runtime)

    lm = ScriptedLM(
        [
            # First attempt: an unfamiliar toolchain EAR has no built-in
            # knowledge of -- the failure is raw text, not a templated error.
            _tool_action(tool="run_shell", args="- command: widgetlang --version"),
            # The model reasons over that raw failure and tries a different
            # command -- still nothing EAR's code chose for it.
            _tool_action(tool="run_shell", args="- command: echo widgetlang built ok"),
            # Having found what works, the model persists it for reuse.
            _tool_action(
                tool="create_tool",
                args=(
                    "- name: run_widgetlang\n"
                    "- description: Build and run a WidgetLang source file.\n"
                    "- command: echo widgetlang built ok"
                ),
            ),
            _tool_action(decision="WidgetLang toolchain confirmed; recipe saved as run_widgetlang."),
        ]
    )
    decision = Reasoner._reason_with_tools(
        Intent(text="Run a WidgetLang snippet"), runtime, lm, context={}, capabilities="none", tools=bound, max_iterations=6
    )
    assert decision == "WidgetLang toolchain confirmed; recipe saved as run_widgetlang."

    records = runtime.reasoning_log.for_stage("tool")
    assert "could not launch 'widgetlang'" in records[0].output  # raw, unmediated failure
    assert "widgetlang built ok" in records[1].output

    # Acquired -- persisted inside the sandbox, reviewable next load.
    assert any(tool.name == "run_widgetlang" for tool in runtime.strategy.tools)
    assert "run_widgetlang" in runtime.sandbox.read_text(".ear/tools.md")
    # Still not auto-executable: declaring a recipe never binds a handler.
    assert "run_widgetlang" not in {tool.name for tool in runtime.tool_binder.bound_tools(runtime)}
    view = runtime.acquirer.view_tool(runtime, "run_widgetlang")
    assert "Command: echo widgetlang built ok" in view  # what the model would replay via run_shell next time


# ---------------------------------------------------------------------------
# Basic toolsets: mechanical capabilities (fetch a URL, search, send mail)
# that ship ready -- enabled/disabled by declaration, never re-derived by
# the model each time it needs one.
# ---------------------------------------------------------------------------


def test_toolsets_default_when_no_section_is_declared():
    from ear.strategy import Strategy

    strategy = Strategy.from_markdown("# Memory & Strategy\n\nNo toolsets section at all.\n")
    assert strategy.toolsets["internet_access"] is True
    assert strategy.toolsets["internet_search"] is False
    assert strategy.toolsets["read_documents"] is True
    assert strategy.toolsets["write_documents"] is False
    assert strategy.toolsets["code_executor"] is True
    assert strategy.toolsets["browser_automation"] is False
    assert strategy.toolsets["terminal"] is False
    assert strategy.toolsets["email_sender"] is False
    assert strategy.toolsets["mcp_connector"] is False
    assert strategy.toolsets["environment_admin"] is True


def test_toolsets_section_overrides_defaults_and_folds_loose_names():
    from ear.strategy import Strategy

    memory = (
        "# Memory & Strategy\n\n## Toolsets\n\n"
        "- Internet Search: enabled, provider tavily, key env var LENS_SEARCH_API_KEY\n"
        "- Code Executor/Writer: disabled\n"
        "- Terminal / Shell: enabled\n"
        "- Email Sender: enabled, smtp host smtp.example.com, port 2525, "
        "user env var SMTP_USER, password env var SMTP_PASS\n"
    )
    strategy = Strategy.from_markdown(memory)
    assert strategy.toolsets["internet_search"] is True
    assert strategy.search_provider == "tavily"
    assert strategy.search_api_key_env_var == "LENS_SEARCH_API_KEY"
    assert strategy.toolsets["code_executor"] is False
    assert strategy.toolsets["terminal"] is True
    assert strategy.toolsets["email_sender"] is True
    assert strategy.smtp_host == "smtp.example.com"
    assert strategy.smtp_port == 2525
    assert strategy.smtp_user_env_var == "SMTP_USER"
    assert strategy.smtp_password_env_var == "SMTP_PASS"
    # untouched defaults survive alongside the declared overrides
    assert strategy.toolsets["internet_access"] is True
    assert strategy.toolsets["environment_admin"] is True


def test_loader_binds_only_the_enabled_basic_toolsets_by_default(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    bound_names = {tool.name for tool in runtime.tool_binder.bound_tools(runtime)}
    assert "fetch_url" in bound_names  # internet_access defaults on
    assert "web_search" not in bound_names  # internet_search defaults off
    assert "send_email" not in bound_names  # email_sender defaults off


def test_loader_binds_web_search_and_email_once_declared_enabled(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = (
        "# Memory & Strategy\n\n## Toolsets\n\n"
        "- Internet Search: enabled, provider tavily, key env var LENS_SEARCH_API_KEY\n"
        "- Email Sender: enabled, smtp host smtp.example.com, user env var SMTP_USER, "
        "password env var SMTP_PASS\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    bound = {tool.name: tool for tool in runtime.tool_binder.bound_tools(runtime)}
    assert "web_search" in bound
    assert "send_email" in bound
    assert "fetch_url" in bound  # internet_access still on by default alongside it


class _FakeOpener:
    """Stands in for `ipv4_opener(context)`'s return value -- an object
    exposing `.open(request, timeout=...)`, the real seam `fetch_url`/
    `web_search`/`LM._post` call through. Patching `ipv4_opener` itself
    (rather than `urllib.request.urlopen`) mocks the actual boundary the
    code uses now that outbound HTTPS is forced over IPv4 (see
    `ear.llm.ipv4_opener`) -- a plain `urlopen` patch would never be
    reached, since nothing calls that name directly any more."""

    def __init__(self, handler):
        self._handler = handler

    def open(self, request, timeout=None):
        return self._handler(request, timeout)


def test_fetch_url_fetches_and_records_on_the_trail(monkeypatch):
    from ear import WebAccess

    class _FakeResponse:
        def read(self, *_args):
            return b"hello from the web"

        def __enter__(self):
            return self

        def __exit__(self, *exc):
            return False

    monkeypatch.setattr("ear.web.ipv4_opener", lambda context: _FakeOpener(lambda request, timeout: _FakeResponse()))
    result = WebAccess().fetch_url("https://example.com")
    assert result == "hello from the web"


def test_fetch_url_fails_loudly_and_never_crashes_the_cycle(monkeypatch):
    import urllib.error

    from ear import WebAccess
    from ear.tool_binder import BoundTool, ToolBinder

    def broken(request, timeout):
        raise urllib.error.URLError("no route to host")

    monkeypatch.setattr("ear.web.ipv4_opener", lambda context: _FakeOpener(broken))
    runtime = Runtime(name="Web-Runtime")
    tool = BoundTool(name="fetch_url", description="fetch", handler=WebAccess().fetch_url)
    invoke = ToolBinder().logged_handler(runtime, tool)
    result = invoke("https://example.com")
    assert "failed" in result and "no route to host" in result
    record = runtime.reasoning_log.for_stage("tool")[0]
    assert record.output.startswith("FAILED")


def test_web_search_refuses_without_a_declared_provider_and_key():
    from ear import WebAccess, WebError

    with pytest.raises(WebError, match="Toolsets: internet_search"):
        WebAccess().web_search("credit risk news")


def test_web_search_calls_tavily_with_the_declared_key(monkeypatch):
    from ear import WebAccess

    captured = {}

    class _FakeResponse:
        def read(self):
            return json.dumps({"results": [{"title": "T", "url": "https://x", "content": "c"}]}).encode("utf-8")

        def __enter__(self):
            return self

        def __exit__(self, *exc):
            return False

    def fake_transport(request, timeout):
        captured["body"] = json.loads(request.data.decode("utf-8"))
        return _FakeResponse()

    monkeypatch.setattr("ear.web.ipv4_opener", lambda context: _FakeOpener(fake_transport))
    monkeypatch.setenv("TAVILY_KEY", "secret-key")
    web = WebAccess(search_provider="tavily", search_api_key_env_var="TAVILY_KEY")
    result = web.web_search("credit risk news")
    assert "T -- https://x" in result
    assert captured["body"]["api_key"] == "secret-key"


def test_send_email_refuses_without_a_declared_host():
    from ear import Mail, MailError

    with pytest.raises(MailError, match="Toolsets: email_sender"):
        Mail().send_email("a@b.com", "subject", "body")


def test_send_email_sends_through_the_declared_smtp_host(monkeypatch):
    from ear import Mail

    sent = {}

    class _FakeSMTP:
        def __init__(self, host, port, timeout=None):
            sent["host"], sent["port"] = host, port

        def __enter__(self):
            return self

        def __exit__(self, *exc):
            return False

        def starttls(self):
            sent["starttls"] = True

        def login(self, user, password):
            sent["login"] = (user, password)

        def send_message(self, message):
            sent["to"] = message["To"]
            sent["subject"] = message["Subject"]

    monkeypatch.setattr("ear.mail.smtplib.SMTP", _FakeSMTP)
    monkeypatch.setenv("SMTP_USER", "bot@example.com")
    monkeypatch.setenv("SMTP_PASS", "hunter2")
    mail = Mail(host="smtp.example.com", user_env_var="SMTP_USER", password_env_var="SMTP_PASS")
    result = mail.send_email("customer@example.com", "Your loan decision", "Approved.")
    assert result == "sent to customer@example.com"
    assert sent == {
        "host": "smtp.example.com",
        "port": 587,
        "starttls": True,
        "login": ("bot@example.com", "hunter2"),
        "to": "customer@example.com",
        "subject": "Your loan decision",
    }


def test_code_executor_and_environment_admin_are_the_same_physical_shell_by_different_names(tmp_path):
    """Terminal, Code Executor and Environment Admin are three access
    grants over one physical capability (Sandbox.run) -- not three
    different implementations. Toolsets defaults code_executor and
    environment_admin on, terminal off; proving the first two actually
    run commands (not a restricted subset) shows the access control is
    the name's reachability, never a fake per-name command filter."""
    directory = write_stack(tmp_path / "stack", **{**MINIMAL_STACK, "memory": SANDBOXED_TOOLS_MEMORY})
    runtime = load_runtime(directory)
    bound = {tool.name: tool for tool in runtime.tool_binder.bound_tools(runtime)}

    assert "code_executor" in bound
    assert "environment_admin" in bound
    assert "terminal" not in bound  # off by default -- but see run_shell below

    result_a = bound["code_executor"].handler("echo from code executor")
    result_b = bound["environment_admin"].handler("echo from code executor")
    assert "from code executor" in result_a
    assert "from code executor" in result_b  # identical capability, different name


def test_terminal_toolset_grants_the_name_once_declared_enabled(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = SANDBOXED_TOOLS_MEMORY + "\n## Toolsets\n\n- Terminal / Shell: enabled\n"
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    bound = {tool.name: tool for tool in runtime.tool_binder.bound_tools(runtime)}
    assert "terminal" in bound
    assert "echo hi" in bound["terminal"].handler("echo echo hi") or "hi" in bound["terminal"].handler("echo hi")


def test_disabling_every_shell_toolset_name_still_leaves_the_sandbox_physically_capable(tmp_path):
    """Access control governs reachability, never existence: with every
    shell-backed toolset name off, none of them are bound -- but the
    sandbox itself can still run a command directly, because the
    physical capability was never something a toggle could remove."""
    stack = dict(MINIMAL_STACK)
    stack["memory"] = (
        SANDBOXED_TOOLS_MEMORY
        + "\n## Toolsets\n\n- Terminal / Shell: disabled\n- Code Executor/Writer: disabled\n"
        "- Environment Admin (Stack Setup): disabled\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    bound_names = {tool.name for tool in runtime.tool_binder.bound_tools(runtime)}
    assert bound_names.isdisjoint({"terminal", "code_executor", "environment_admin"})

    # Still physically there -- just not reachable by name for the model.
    result = runtime.sandbox.run("echo still physically here")
    assert result.ok
    assert "still physically here" in result.stdout


# ---------------------------------------------------------------------------
# Optimizer: extended to Skills -- the same reflection primitive N1.4 uses
# for a Judgment's instruction, aimed at a Skill's prompt instead.
# ---------------------------------------------------------------------------


def test_optimizer_refine_skill_is_a_no_op_without_a_model():
    from ear import Optimizer, Skill
    from ear.optimizer import Example

    skill = Skill(name="fetch_score", prompt="Fetch the credit score.")
    result = Optimizer().refine_skill(skill, [Example(intent="x", decision="y")], model_binding=None)
    assert result == "Fetch the credit score."
    assert skill.prompt == "Fetch the credit score."


def test_optimizer_refine_skill_rewrites_the_prompt_with_a_scripted_model():
    from ear import ModelBinding, Optimizer, Skill
    from ear.optimizer import Example

    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM(["## improved instruction\n\nFetch the applicant's credit score from the bureau.\n"])
    skill = Skill(name="fetch_score", prompt="Fetch the credit score.")

    result = Optimizer().refine_skill(
        skill, [Example(intent="check credit", decision="score fetched")], model_binding=binding
    )
    assert result == "Fetch the applicant's credit score from the bureau."
    assert skill.prompt == "Fetch the applicant's credit score from the bureau."


def test_optimizer_persist_skill_replaces_only_its_own_section(tmp_path):
    from ear import Optimizer, Skill

    skills_md = tmp_path / "skills.md"
    skills_md.write_text(
        "# Skills\n\n## risk_grade\nCombine the score band and DTI band into a grade A-E.\n\n"
        "## fetch_score\nFetch the credit score.\n",
        encoding="utf-8",
    )
    refined = Skill(name="fetch_score", prompt="Fetch the applicant's credit score from the bureau.")
    Optimizer().persist_skill(skills_md, refined)

    text = skills_md.read_text(encoding="utf-8")
    assert "Fetch the applicant's credit score from the bureau." in text
    assert "Combine the score band and DTI band into a grade A-E." in text  # untouched section survives
    assert text.count("## fetch_score") == 1


def test_deliberator_backend_seam_stays_on_the_trail():
    class TypedBackend:
        def deliberate(self, runtime, intent, plan=None, research=None):
            return "typed decision from the backend"

    runtime = Runtime(name="Seamed-Runtime")
    runtime.orchestrator.executor.performer.deliberator.backend = TypedBackend()
    decision = runtime.reason(Intent(text="anything"))
    assert decision == "typed decision from the backend"
    record = runtime.reasoning_log.for_stage("deliberation")[0]
    assert record.model == "backend:TypedBackend"
    assert record.output == decision


@requires_anthropic_key
def test_react_invokes_the_bound_calculator_live(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = TOOLS_MEMORY + (
        "\n## Model Selection\n\nReason with anthropic/claude-haiku-4-5, reading the credential from\n"
        "ANTHROPIC_API_KEY, at a temperature of 0.2.\n"
    )
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    runtime.bind_tool("amortization_calculator", monthly_payment)

    decision = runtime.reason(
        Intent(
            text=(
                "Underwrite a $12,000 personal loan at 6 percent annual interest over 36 months; "
                "compute the exact monthly payment with the amortization calculator before deciding."
            ),
            context={"loan_amount": 12000, "credit_score": 760, "debt_to_income": 0.2},
        )
    )
    assert decision
    invocations = runtime.reasoning_log.for_stage("tool")
    assert invocations, "the model never called the bound tool"
    assert invocations[0].inputs["tool"] == "amortization_calculator"
    assert not invocations[0].output.startswith("FAILED")


# ---------------------------------------------------------------------------
# Approval gates: a violated Approval-required policy parks the cycle for
# a human verdict instead of blocking it; the verdict is a markdown
# document, and everything lands on the trail.
# ---------------------------------------------------------------------------

APPROVAL_POLICY = (
    "# Policies\n\n## Loan Amount Cap\nThe loan must not exceed $75,000.\n\n"
    "Fallback: loan_amount <= 75000\nApplies to: runtime\n\n"
    "## Large Loan Human Approval\nLoans above $50,000 need a human approver.\n\n"
    "Fallback: loan_amount <= 50000\nApproval: required\nApplies to: runtime\n"
)


def approval_stack(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["policy"] = APPROVAL_POLICY
    stack["workflow"] = (
        "# Workflows\n\n## Underwriting Workflow\n\n"
        "1. Decide approve or decline. (Credit Risk Guru)\n"
    )
    return write_stack(tmp_path / "stack", **stack)


def test_loader_reads_the_approval_field_and_rejects_the_unreadable(tmp_path):
    runtime = load_runtime(approval_stack(tmp_path))
    by_name = {policy.name: policy for policy in runtime.policies}
    assert by_name["Large Loan Human Approval"].approval_required is True
    assert by_name["Loan Amount Cap"].approval_required is False

    stack = dict(MINIMAL_STACK)
    stack["workflow"] = "# Workflows\n\n## Underwriting Workflow\n\n1. Decide. (Credit Risk Guru)\n"
    stack["policy"] = "# Policies\n\n## Odd Gate\nSomething.\n\nApproval: perhaps\n"
    with pytest.raises(ValueError, match="unreadable Approval field"):
        load_runtime(write_stack(tmp_path / "unreadable", **stack))

    stack["policy"] = "# Policies\n\n## Ungated\nSomething.\n\nApproval: not required\n"
    runtime = load_runtime(write_stack(tmp_path / "negated", **stack))
    assert runtime.policies[0].approval_required is False


def test_violated_approval_gate_parks_the_cycle_on_the_record(tmp_path):
    from ear import ApprovalRequired

    runtime = load_runtime(approval_stack(tmp_path))
    with pytest.raises(ApprovalRequired, match="Large Loan Human Approval") as parked:
        runtime.reason(Intent(text="Underwrite a large loan", context={"loan_amount": 60000}))

    assert isinstance(parked.value, PermissionError)  # existing handlers keep working
    policy_record = next(
        record for record in runtime.reasoning_log.for_stage("policy")
        if record.inputs["policy"] == "Large Loan Human Approval"
    )
    assert policy_record.output == "PENDING APPROVAL"
    pending = runtime.reasoning_log.for_stage("approval")[0]
    assert "PENDING -- human approval required" in pending.output
    assert runtime.reasoning_log.for_stage("usage")  # a parked cycle is accounted too
    assert len(runtime.memory) == 0  # nothing was decided, nothing is remembered


def test_approved_verdict_releases_the_gate_and_is_on_the_record(tmp_path):
    from ear import Approval

    runtime = load_runtime(approval_stack(tmp_path))
    approval = Approval(verdict=True, approver="lakshmi@gigkri.com", note="Collateral covers it.")
    decision = runtime.reason(Intent(text="Underwrite a large loan", context={"loan_amount": 60000}), approval=approval)

    assert decision
    released = runtime.reasoning_log.for_stage("approval")[0]
    assert released.output == "approved by lakshmi@gigkri.com"
    assert "Collateral covers it." in released.rationale
    assert len(runtime.memory) == 1


def test_rejected_verdict_blocks_like_any_violation(tmp_path):
    from ear import Approval, ApprovalRequired

    runtime = load_runtime(approval_stack(tmp_path))
    approval = Approval(verdict=False, approver="lakshmi@gigkri.com")
    with pytest.raises(PermissionError, match="Large Loan Human Approval") as blocked:
        runtime.reason(Intent(text="Underwrite a large loan", context={"loan_amount": 60000}), approval=approval)
    assert not isinstance(blocked.value, ApprovalRequired)  # a rejection is a block, not a park
    assert runtime.reasoning_log.for_stage("approval")[0].output == "REJECTED by lakshmi@gigkri.com"


def test_hard_block_wins_over_a_pending_gate(tmp_path):
    runtime = load_runtime(approval_stack(tmp_path))
    from ear import ApprovalRequired

    with pytest.raises(PermissionError, match="Loan Amount Cap") as blocked:
        runtime.reason(Intent(text="Underwrite a huge loan", context={"loan_amount": 90000}))
    assert not isinstance(blocked.value, ApprovalRequired)


def test_approval_document_round_trips():
    from ear import Approval

    text = Approval(verdict=True, approver="lakshmi@gigkri.com", note="Fine by me.").to_markdown("A big loan")
    read = Approval.from_markdown(text)
    assert read.verdict is True
    assert read.approver == "lakshmi@gigkri.com"
    assert read.note == "Fine by me."
    assert Approval.from_markdown("# Approval\n\nVerdict: perhaps\n").verdict is None


def test_exchange_parks_releases_and_stays_idempotent(tmp_path):
    directory = approval_stack(tmp_path)
    (directory / "intents").mkdir()
    (directory / "intents" / "big-loan.md").write_text(
        "# Underwrite a $60,000 loan\n\n## Context\n\n- loan_amount: 60000\n",
        encoding="utf-8",
    )
    runtime = load_runtime(directory)
    exchange = Exchange(directory)

    exchange.run(runtime)
    parked = (directory / "decisions" / "big-loan.md").read_text(encoding="utf-8")
    assert "Status: PENDING APPROVAL" in parked
    assert "## Awaiting approval" in parked and "Large Loan Human Approval" in parked

    # No approval document yet: nothing to do.
    assert exchange.run(runtime) == []

    (directory / "approvals").mkdir()
    (directory / "approvals" / "big-loan.md").write_text(
        "# Approval -- Underwrite a $60,000 loan\n\n"
        "Verdict: approved\nApprover: lakshmi@gigkri.com\n\n> Collateral covers it.\n",
        encoding="utf-8",
    )
    written = exchange.run(runtime)
    assert [path.name for path in written] == ["big-loan.md"]
    released = (directory / "decisions" / "big-loan.md").read_text(encoding="utf-8")
    assert "Status: decided" in released
    assert "## Approval" in released and "lakshmi@gigkri.com" in released

    # Released cycles are settled: a third run replays nothing.
    assert exchange.run(runtime) == []


def test_exchange_writes_blocked_document_on_a_rejected_verdict(tmp_path):
    directory = approval_stack(tmp_path)
    (directory / "intent.md").write_text(
        "# Underwrite a $60,000 loan\n\n## Context\n\n- loan_amount: 60000\n", encoding="utf-8"
    )
    runtime = load_runtime(directory)
    exchange = Exchange(directory)
    exchange.run(runtime)
    (directory / "approval.md").write_text("# Approval\n\nVerdict: rejected\nApprover: lakshmi\n", encoding="utf-8")
    exchange.run(runtime)
    final = (directory / "decision.md").read_text(encoding="utf-8")
    assert "Status: BLOCKED" in final and "Large Loan Human Approval" in final


def test_exchange_refuses_an_unreadable_approval_verdict(tmp_path):
    directory = approval_stack(tmp_path)
    (directory / "intent.md").write_text(
        "# Underwrite a $60,000 loan\n\n## Context\n\n- loan_amount: 60000\n", encoding="utf-8"
    )
    runtime = load_runtime(directory)
    exchange = Exchange(directory)
    exchange.run(runtime)
    (directory / "approval.md").write_text("# Approval\n\nVerdict: perhaps\n", encoding="utf-8")
    with pytest.raises(ValueError, match="no readable Verdict"):
        exchange.run(runtime)


@requires_anthropic_key
def test_approval_gate_parks_and_releases_live(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["policy"] = APPROVAL_POLICY
    stack["memory"] = LIVE_MEMORY
    directory = write_stack(tmp_path / "stack", **stack)
    (directory / "intent.md").write_text(
        "# Underwrite a $62,000 personal loan for a strong applicant\n\n"
        "## Context\n\n- loan_amount: 62000\n- credit_score: 790\n- debt_to_income: 0.15\n",
        encoding="utf-8",
    )
    runtime = load_runtime(directory)
    exchange = Exchange(directory)
    exchange.run(runtime)
    assert "Status: PENDING APPROVAL" in (directory / "decision.md").read_text(encoding="utf-8")

    (directory / "approval.md").write_text(
        "# Approval\n\nVerdict: approved\nApprover: lakshmi@gigkri.com\n\n> Reviewed; proceed.\n",
        encoding="utf-8",
    )
    exchange.run(runtime)
    released = (directory / "decision.md").read_text(encoding="utf-8")
    assert "Status: decided" in released and "## Approval" in released
    approvals = [record.output for record in runtime.reasoning_log.for_stage("approval")]
    assert "approved by lakshmi@gigkri.com" in approvals


# ---------------------------------------------------------------------------
# N1: LM hardening, per-judgment accounting, pricing, instruction search,
# demos and persisted instructions -- reasoning & optimization depth.
# ---------------------------------------------------------------------------


class FakeReply:
    def __init__(self, payload: dict):
        self._payload = payload

    def read(self):
        return json.dumps(self._payload).encode("utf-8")

    def __enter__(self):
        return self

    def __exit__(self, *args):
        return False


ANTHROPIC_REPLY = {
    "content": [{"type": "text", "text": "## decision\n\nfine\n"}],
    "usage": {"input_tokens": 12, "output_tokens": 5},
}


def test_lm_retries_transient_failures_and_records_them(monkeypatch):
    import io
    import urllib.error

    from ear.llm import LM

    attempts = {"count": 0}

    def flaky(request, timeout):
        attempts["count"] += 1
        if attempts["count"] <= 2:
            # A real HTTPError always carries the response body as `fp` --
            # `ear.llm._post` reads it via `error.read()` to surface the
            # provider's detail, so the mock must supply a real file-like
            # object here too, not None.
            raise urllib.error.HTTPError(request.full_url, 529, "overloaded", {}, io.BytesIO(b"overloaded"))
        return FakeReply(ANTHROPIC_REPLY)

    monkeypatch.setattr("ear.llm.ipv4_opener", lambda context: _FakeOpener(flaky))
    monkeypatch.setattr("ear.llm.time.sleep", lambda seconds: None)

    lm = LM(model="anthropic/test-model", api_key="k")
    reply = lm.complete("anything")
    assert "fine" in reply
    assert attempts["count"] == 3
    entry = lm.history[-1]
    assert entry["retries"] == 2
    assert entry["usage"] == {
        "prompt_tokens": 12,
        "completion_tokens": 5,
        "cache_read_tokens": 0,
        "cache_write_tokens": 0,
    }
    assert entry["latency_ms"] >= 0


def test_lm_fails_fast_on_non_retryable_errors(monkeypatch):
    import io
    import urllib.error

    from ear.llm import LM, LMError

    attempts = {"count": 0}

    def unauthorized(request, timeout):
        attempts["count"] += 1
        raise urllib.error.HTTPError(request.full_url, 401, "unauthorized", {}, io.BytesIO(b"unauthorized"))

    monkeypatch.setattr("ear.llm.ipv4_opener", lambda context: _FakeOpener(unauthorized))
    with pytest.raises(LMError, match="401"):
        LM(model="anthropic/test-model", api_key="bad").complete("anything")
    assert attempts["count"] == 1  # auth errors never retry


class ScriptedLM:
    """A stub LM answering with fixed markdown sections and recording
    history the way the real one does."""

    def __init__(self, replies=None):
        self.replies = list(replies or [])
        self.history = []
        self.prompts = []

    def complete(self, prompt, system="", cache_prefix=""):
        self.prompts.append(prompt)
        self.cache_prefixes = getattr(self, "cache_prefixes", [])
        self.cache_prefixes.append(cache_prefix)
        reply = self.replies.pop(0) if self.replies else "## decision\n\nok\n"
        self.history.append(
            {"usage": {"prompt_tokens": 10, "completion_tokens": 3}, "latency_ms": 7, "retries": 0}
        )
        return reply


def test_per_judgment_usage_rides_the_stage_record():
    from ear import ModelBinding, Policy

    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM(["## complies\n\nyes\n\n## rationale\n\nWithin the cap.\n"])
    runtime = Runtime(name="Accounted-Runtime", model_binding=binding)
    runtime.add_policy(Policy(name="Cap", statement="Stay under the cap."))

    runtime.governor.govern(runtime, Intent(text="check", context={"amount": 5}))
    record = runtime.reasoning_log.for_stage("policy")[0]
    assert record.input_tokens == 10 and record.output_tokens == 3
    assert record.latency_ms == 7
    assert "10+3 tok" in record.render()


def test_fallback_judgments_are_never_billed(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))
    for record in runtime.reasoning_log.records:
        assert record.input_tokens == 0 and record.output_tokens == 0


def test_pricing_is_declared_in_prose_and_prices_usage():
    strategy = Strategy.from_markdown(
        "## Pricing\n\nInput tokens cost $3 per million; output tokens cost $15 per million.\n"
    )
    assert strategy.input_rate_per_million == 3.0
    assert strategy.output_rate_per_million == 15.0
    assert strategy.dollars(1_000_000, 200_000) == pytest.approx(6.0)
    assert Strategy().dollars(1000, 1000) is None  # undeclared -> never invented


def test_dollars_prices_cached_input_at_read_and_write_multipliers():
    strategy = Strategy.from_markdown(
        "## Pricing\n\nInput tokens cost $3 per million; output tokens cost $15 per million.\n"
    )
    # 1M uncached input ($3) + 1M cache reads (0.1x -> $0.30) + 1M cache writes
    # (1.25x -> $3.75) + 200k output ($3) = $10.05. The three input counts are
    # distinct, so nothing double-bills.
    cost = strategy.dollars(1_000_000, 200_000, 1_000_000, 1_000_000)
    assert cost == pytest.approx(3.0 + 0.30 + 3.75 + 3.0)
    # A cache read is an order of magnitude cheaper than the same tokens uncached.
    assert strategy.dollars(0, 0, 1_000_000, 0) == pytest.approx(0.30)
    # Defaulted cache args keep the old two-arg call byte-for-byte unchanged.
    assert strategy.dollars(1_000_000, 200_000) == pytest.approx(6.0)


def test_reasoning_record_round_trips_cache_tokens():
    from ear.reasoning_log import ReasoningLog

    log = ReasoningLog()
    log.record(
        stage="selection",
        output="chose a tool",
        model="anthropic/claude-sonnet-5",
        usage={"input_tokens": 40, "output_tokens": 8, "cache_read_tokens": 900, "cache_write_tokens": 60},
    )
    entry = log.records[-1]
    assert entry.cache_read_tokens == 900 and entry.cache_write_tokens == 60

    # Survives the JSONL trail round-trip (to_json -> from_trail).
    import json
    import tempfile

    with tempfile.NamedTemporaryFile("w", suffix=".jsonl", delete=False) as handle:
        handle.write(entry.to_json() + "\n")
        path = handle.name
    restored = ReasoningLog.from_trail(path).records[-1]
    assert restored.cache_read_tokens == 900 and restored.cache_write_tokens == 60
    assert "cache_read_tokens" in json.loads(entry.to_json())


def test_usage_record_carries_dollars_only_when_priced():
    from ear import ModelBinding

    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM()
    binding.lm.history.append({"usage": {"prompt_tokens": 1_000_000, "completion_tokens": 0}, "retries": 1})
    runtime = Runtime(name="Priced-Runtime", model_binding=binding)
    runtime.strategy = Strategy.from_markdown("## Pricing\n\nInput tokens cost $3 per million.\n")

    runtime._record_usage(started=time.monotonic(), calls_before=0)
    usage = runtime.reasoning_log.for_stage("usage")[0]
    assert usage.inputs["cost"] == pytest.approx(3.0)
    assert "~$3.0" in usage.output and "1 retried" in usage.output


def test_search_refuses_without_a_model():
    from ear import Optimizer
    from ear.judgment import Field, Judgment
    from ear.optimizer import Example

    judgment = Judgment(instruction="Decide.", inputs=[Field("intent")], outputs=[Field("decision")])
    with pytest.raises(ValueError, match="optimization is judgment"):
        Optimizer().search(judgment, [Example(intent="x", decision="y")], model_binding=None)


def test_demos_render_into_the_prompt_in_answer_shape():
    from ear.judgment import Field, Judgment

    judgment = Judgment(
        instruction="Decide.",
        inputs=[Field("intent")],
        outputs=[Field("decision")],
        demos=[{"intent": "a small loan", "decision": "approve"}],
    )
    prompt = judgment.render_prompt({"intent": "a big loan"})
    assert "Worked example 1:" in prompt
    assert prompt.index("a small loan") < prompt.index("Now the task at hand:") < prompt.index("a big loan")


def test_select_demos_prefers_reviewed_examples_within_budget():
    from ear import Optimizer
    from ear.judgment import Field, Judgment
    from ear.optimizer import Example
    from ear.signatures import ReasonAboutIntent

    judgment = Judgment(
        instruction=ReasonAboutIntent.instruction,
        inputs=list(ReasonAboutIntent.inputs),
        outputs=list(ReasonAboutIntent.outputs),
    )
    examples = [
        Example(intent="unreviewed", decision="maybe"),
        Example(intent="approved one", decision="approve", verdict=True),
        Example(intent="wrong one", decision="nonsense", verdict=False),
    ]
    demos = Optimizer().select_demos(judgment, examples, budget_chars=200)
    assert demos[0]["intent"] == "approved one"  # reviewed first
    assert all(demo["intent"] != "wrong one" for demo in demos)  # never the judged-wrong
    assert judgment.demos == demos


def test_instructions_persist_and_reload_with_demos(tmp_path):
    from ear import Optimizer
    from ear.judgment import Field, Judgment

    judgment = Judgment(
        instruction="The refined instruction.",
        inputs=[Field("intent")],
        outputs=[Field("decision")],
        demos=[{"intent": "a loan", "decision": "approve"}],
    )
    optimizer = Optimizer()
    path = optimizer.save_instructions(tmp_path / ".ear" / "instructions.md", {"MyJudgment": judgment})

    fresh = Judgment(instruction="The shipped default.", inputs=[Field("intent")], outputs=[Field("decision")])
    applied = optimizer.load_instructions(path, {"MyJudgment": fresh})
    assert applied == ["MyJudgment"]
    assert fresh.instruction == "The refined instruction."
    assert fresh.demos == [{"intent": "a loan", "decision": "approve"}]


def test_load_runtime_applies_persisted_instructions_from_dot_ear(tmp_path):
    """N1.6's actual promise: optimization survives restarts because the
    Loader applies `.ear/instructions.md` on load -- not just that
    Optimizer.load_instructions works when called directly."""
    from ear import Optimizer
    from ear.judgment import Judgment
    from ear.signatures import REGISTRY

    target = REGISTRY["ExplainDecision"]
    original_instruction, original_demos = target.instruction, target.demos
    try:
        directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
        Optimizer().save_instructions(
            directory / ".ear" / "instructions.md",
            {"ExplainDecision": Judgment(
                instruction="A persisted, reload-surviving explanation instruction.",
                inputs=list(target.inputs),
                outputs=list(target.outputs),
                demos=[{"decision": "approved", "explanation": "within every declared limit"}],
            )},
        )

        load_runtime(directory)  # the Loader call under test; discards its own runtime

        assert REGISTRY["ExplainDecision"].instruction == (
            "A persisted, reload-surviving explanation instruction."
        )
        assert REGISTRY["ExplainDecision"].demos == [
            {"decision": "approved", "explanation": "within every declared limit"}
        ]
    finally:
        target.instruction, target.demos = original_instruction, original_demos


@requires_anthropic_key
def test_search_improves_or_keeps_the_baseline_live():
    from ear import ModelBinding, Optimizer
    from ear.judgment import Field, Judgment
    from ear.optimizer import Example

    binding = ModelBinding(provider="anthropic", model=os.environ.get("ANTHROPIC_TEST_MODEL", "claude-haiku-4-5"))
    binding.activate()
    # A scratch copy so the search never mutates the shared registry.
    judgment = Judgment(
        instruction="Decide approve or decline for the loan intent. Answer with one word.",
        inputs=[Field("intent", "The loan request")],
        outputs=[Field("decision", "approve or decline, one word", "str")],
    )
    examples = [
        Example(intent="A $5,000 loan, credit score 790, no debts", decision="approve", verdict=True),
        Example(intent="A $9,000 loan for a defaulted borrower with no income", decision="decline", verdict=True),
    ]
    outcome = Optimizer().search(
        judgment, examples, model_binding=binding, generations=1, candidates=2, holdout=1.0
    )
    assert outcome.best_score >= outcome.baseline_score
    assert outcome.evaluations >= 2
    assert judgment.instruction  # the winner (or the kept baseline) is in place


# ---------------------------------------------------------------------------
# Observability: the trail fans out to exporters (a native protocol --
# anything with export(record)) and every cycle carries usage accounting.
# ---------------------------------------------------------------------------


class CollectingExporter:
    def __init__(self, fail: bool = False):
        self.records = []
        self.flushes = 0
        self.fail = fail

    def export(self, record):
        if self.fail:
            raise RuntimeError("exporter down")
        self.records.append(record)

    def flush(self):
        self.flushes += 1


def test_reasoning_log_fans_out_to_exporters_exactly_once():
    exporter = CollectingExporter()
    runtime = Runtime(name="Exported-Runtime")
    runtime.reasoning_log.exporters.append(exporter)

    runtime.reason(Intent(text="first cycle"))
    assert [record.stage for record in exporter.records][0] == "intent"
    assert exporter.records[-1].stage == "usage"
    assert exporter.flushes == 1

    seen = len(exporter.records)
    runtime.reasoning_log.flush()  # nothing pending -- no double export
    assert len(exporter.records) == seen


def test_exporter_failure_never_breaks_a_cycle(tmp_path):
    runtime = Runtime(name="Resilient-Runtime")
    runtime.reasoning_log.path = str(tmp_path / "trail.md")
    runtime.reasoning_log.exporters.append(CollectingExporter(fail=True))

    decision = runtime.reason(Intent(text="first cycle"))
    assert decision  # the cycle completed
    assert runtime.reasoning_log.export_errors
    assert "exporter down" in runtime.reasoning_log.export_errors[0]
    # The file on disk stays the canonical record.
    assert "## Cycle 1" in (tmp_path / "trail.md").read_text(encoding="utf-8")


def test_usage_record_closes_every_cycle_offline():
    runtime = Runtime(name="Accounted-Runtime")
    runtime.reason(Intent(text="a cycle"))
    usage = runtime.reasoning_log.for_stage("usage")[0]
    assert "0 model calls" in usage.output and "ms" in usage.output
    assert usage.inputs["model_calls"] == 0
    assert isinstance(usage.inputs["latency_ms"], int)


def test_usage_is_recorded_on_blocked_cycles_too(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    with pytest.raises(PermissionError):
        runtime.reason(Intent(text="Underwrite", context={"debt_to_income": 0.60}))
    assert runtime.reasoning_log.for_stage("usage")


# ---------------------------------------------------------------------------
# Knowledge and the Librarian: declared sources, retrieval on the record,
# citations in evidence and decision documents.
# ---------------------------------------------------------------------------

KNOWLEDGE_MEMORY = (
    "# Memory & Strategy\n\n## Knowledge\n\n"
    "The reference material the Librarian may consult and cite.\n\n"
    "- underwriting manual: `knowledge/manual.md`\n"
)

MANUAL = (
    "# Underwriting Manual\n\n## Section 4.2 -- Marginal applicants\n\n"
    "A grade C applicant whose debt-to-income band is not low must be declined.\n\n"
    "## Section 9.9 -- Office plants\n\nWater the plants on Fridays.\n"
)


def knowledge_stack(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = KNOWLEDGE_MEMORY
    directory = write_stack(tmp_path / "stack", **stack)
    (directory / "knowledge").mkdir()
    (directory / "knowledge" / "manual.md").write_text(MANUAL, encoding="utf-8")
    return directory


def test_strategy_reads_file_and_url_knowledge_sources():
    strategy = Strategy.from_markdown(KNOWLEDGE_MEMORY)
    assert strategy.knowledge_sources == [
        KnowledgeSource(name="underwriting manual", pattern="knowledge/manual.md")
    ]
    assert "reference material" in strategy.knowledge

    declared = Strategy.from_markdown(
        "## Knowledge\n\n- market brief: https://example.com/brief.md, refetch weekly\n"
    ).knowledge_sources
    assert declared == [
        KnowledgeSource(name="market brief", url="https://example.com/brief.md", refresh_days=7.0)
    ]

    with pytest.raises(ValueError, match="declares no path"):
        Strategy.from_markdown("## Knowledge\n\n- manual\n")


def test_loader_builds_knowledge_chunked_by_section(tmp_path):
    runtime = load_runtime(knowledge_stack(tmp_path))
    knowledge = runtime.librarian.knowledge
    assert knowledge is not None and len(knowledge) == 2
    assert knowledge.passages[0].source == "underwriting manual -- manual.md § Section 4.2 -- Marginal applicants"
    assert "must be declined" in knowledge.passages[0].text


def test_loader_fails_loudly_on_a_missing_knowledge_source(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Knowledge\n\n- manual: `knowledge/absent.md`\n"
    with pytest.raises(ValueError, match="matched no files"):
        load_runtime(write_stack(tmp_path / "stack", **stack))


def test_librarian_retrieves_structurally_offline_with_citations(tmp_path):
    directory = knowledge_stack(tmp_path)
    (directory / "intent.md").write_text(
        "# Underwrite a marginal grade C applicant whose debt-to-income band is not low\n\n"
        "## Context\n\n- loan_amount: 9000\n",
        encoding="utf-8",
    )
    runtime = load_runtime(directory)
    Exchange(directory).run(runtime)

    retrieval = runtime.reasoning_log.for_stage("retrieval")[0]
    assert "structural retrieval" in retrieval.rationale
    assert "Section 4.2" in retrieval.output

    evidence = runtime.memory.working[-1].evidence
    assert any("Section 4.2" in citation for citation in evidence.sources["citations"])
    decision_document = (directory / "decision.md").read_text(encoding="utf-8")
    assert "## Sources" in decision_document and "underwriting manual" in decision_document

    deliberation = runtime.reasoning_log.for_stage("deliberation")[0]
    assert "never as" in deliberation.inputs["knowledge"]
    assert "must be declined" in deliberation.inputs["knowledge"]


def test_librarian_retriever_seam_accepts_any_passage_source():
    from ear import Passage

    class CustomRetriever:
        def retrieve(self, query):
            return [Passage(source="manual.md", text="Section one text.")]

    runtime = Runtime(name="Retriever-Runtime")
    runtime.librarian.retriever = CustomRetriever()
    research = runtime.librarian.research(runtime, Intent(text="anything"))
    assert research is not None and research.citations[0] == "manual.md"
    assert runtime.reasoning_log.for_stage("retrieval")


@requires_anthropic_key
def test_retrieval_cites_the_manual_live(tmp_path):
    directory = tmp_path / "credit_risk_stack"
    shutil.copytree(EXAMPLE_STACK, directory)
    memory = (directory / "memory.md").read_text(encoding="utf-8")
    (directory / "memory.md").write_text(
        memory.replace("anthropic/claude-opus-4-8", "anthropic/claude-haiku-4-5"), encoding="utf-8"
    )
    (directory / "intents").mkdir()
    # A unique reference defeats the LM's prompt cache, so this test
    # genuinely exercises live usage accounting, not a cached replay.
    reference = int(time.time())
    (directory / "intents" / "marginal.md").write_text(
        f"# Underwrite a $9,500 personal loan for a marginal applicant (reference {reference})\n\n"
        "## Context\n\n- loan_amount: 9500\n- credit_score: 665\n- debt_to_income: 0.41\n"
        "- annual_income: 52000\n- existing_defaults: 0\n",
        encoding="utf-8",
    )
    runtime = load_runtime(directory)
    Exchange(directory).run(runtime)

    retrieval = runtime.reasoning_log.for_stage("retrieval")[-1]
    assert retrieval.model == "anthropic/claude-haiku-4-5"
    decision_document = (directory / "decisions" / "marginal.md").read_text(encoding="utf-8")
    assert "## Sources" in decision_document and "underwriting manual" in decision_document

    usage = runtime.reasoning_log.for_stage("usage")[-1]
    assert usage.inputs["model_calls"] > 0
    assert usage.inputs["input_tokens"] > 0


# ---------------------------------------------------------------------------
# Dynamic-at-runtime stages: selection, scheduling and delegation are
# judgments the LLM makes per cycle, with deterministic fallbacks offline
# and an audit record either way.
# ---------------------------------------------------------------------------


def test_selector_falls_back_to_dedupe_and_logs_the_choice_offline():
    from ear import Process

    runtime = Runtime(name="Choosing-Runtime")
    first, second = Process(name="Underwrite Loan"), Process(name="Close Account")
    selected = runtime.selector.select(runtime, [first, second, first], intent=Intent(text="underwrite"))

    assert selected == [first, second]
    record = runtime.reasoning_log.for_stage("selection")[0]
    assert record.output == "Underwrite Loan, Close Account"
    assert record.model == "deterministic-fallback"


def test_selector_stays_silent_when_there_is_no_choice():
    from ear import Process

    runtime = Runtime(name="Quiet-Runtime")
    runtime.selector.select(runtime, [Process(name="Only Process")], intent=Intent(text="x"))
    assert runtime.reasoning_log.for_stage("selection") == []


def test_scheduler_falls_back_to_composition_order_and_logs_offline():
    from ear import Workflow

    runtime = Runtime(name="Ordering-Runtime")
    plan = [Workflow(name="Banding"), Workflow(name="Deciding")]
    scheduled = runtime.scheduler.schedule(plan, runtime=runtime, intent=Intent(text="underwrite"))

    assert scheduled == plan and scheduled is not plan
    record = runtime.reasoning_log.for_stage("scheduling")[0]
    assert record.output == "Banding, Deciding"
    assert record.model == "deterministic-fallback"


def test_delegator_leaves_steps_as_authored_offline():
    from ear import Workflow

    runtime = Runtime(name="Delegating-Runtime")
    persona = Persona(name="Credit Risk Guru")
    workflow = Workflow(name="Underwriting Workflow")
    workflow.add_step("Band the profile.", persona=persona)
    workflow.add_step("Write the customer note.")  # left undelegated by the author

    runtime.delegator.delegate(runtime, Intent(text="underwrite"), [workflow])
    assert workflow.steps[1].persona is None
    assert runtime.reasoning_log.for_stage("delegation") == []


def test_delegator_apply_only_binds_resolvable_assignments():
    from ear.delegator import Delegator
    from ear import Workflow

    persona = Persona(name="Customer Advocate")
    workflow = Workflow(name="W")
    workflow.add_step("Write the customer note.")
    undelegated = [(1, workflow.steps[0])]

    applied = Delegator._apply(undelegated, [persona], ["1: Customer Advocate", "9: Nobody", "garbage"])
    assert applied == [(1, persona)]
    assert workflow.steps[0].persona is persona


def test_runtime_reason_records_selection_between_processes(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["process"] = (
        "# Credit Risk Runtime\n\n## Underwrite Consumer Loan\nEvaluates a loan application.\n\n"
        "Workflows: Underwriting Workflow\n\n## Close Dormant Account\nCloses a dormant account.\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    runtime.reason(Intent(text="anything ambiguous", context={"loan_amount": 5000}))
    assert runtime.reasoning_log.for_stage("selection")


def test_unknown_reference_error_suggests_the_close_match(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["persona"] = "# Personas\n\n## Guru\nBe careful.\n\nSkills: risk_grad\n"
    with pytest.raises(ValueError, match="did you mean 'risk_grade'"):
        load_runtime(write_stack(tmp_path / "stack", **stack))


def test_recaller_and_auditor_stay_deterministic_and_silent_offline():
    runtime = Runtime(name="Offline-Runtime")
    runtime.reason(Intent(text="first cycle"))
    runtime.reason(Intent(text="second cycle"))
    # Offline recall is the full window and audit is the flag -- neither is
    # a judgment, so neither lands in the trail.
    assert runtime.reasoning_log.for_stage("recall") == []
    assert runtime.reasoning_log.for_stage("audit") == []
    entry = runtime.memory.working[-1]
    assert entry.evidence.sources["audited"] is True
    assert "first cycle" in entry.evidence.sources["recalled_memory"]


def test_adaptation_is_logged_when_distilled():
    runtime = Runtime(name="Adapting-Runtime")
    runtime.adapter.adapt_every = 1
    runtime.reason(Intent(text="underwrite something"))
    record = runtime.reasoning_log.for_stage("adaptation")[0]
    assert record.output == runtime.adaptations.impressions[0].insight
    assert record.model == "deterministic-fallback"


@requires_anthropic_key
def test_recall_and_audit_are_llm_judged_and_on_the_trail():
    from ear import ModelBinding

    binding = ModelBinding(provider="anthropic", model=os.environ.get("ANTHROPIC_TEST_MODEL", "claude-haiku-4-5"))
    runtime = Runtime(name="Live-Audited-Runtime", model_binding=binding)
    runtime.memory.record("approved a $5,000 bicycle loan", decision="approved at grade B")
    runtime.memory.record("cancelled a lunch reservation", decision="cancelled")

    runtime.reason(Intent(text="Underwrite another small bicycle loan", context={"loan_amount": 4000}))

    recall = runtime.reasoning_log.for_stage("recall")[0]
    assert recall.model == binding.model_id
    audit = runtime.reasoning_log.for_stage("audit")[0]
    assert audit.output.strip()
    assert runtime.memory.working[-1].evidence.sources["audit_assessment"] == audit.output


@requires_anthropic_key
def test_selector_chooses_the_relevant_process_with_llm():
    from ear import ModelBinding, Process

    binding = ModelBinding(provider="anthropic", model=os.environ.get("ANTHROPIC_TEST_MODEL", "claude-haiku-4-5"))
    binding.activate()
    runtime = Runtime(name="Choosing-Runtime", model_binding=binding)
    loan = Process(name="Underwrite Consumer Loan", description="Reviews a loan applicant's credit to decide approval.")
    lunch = Process(name="Cancel Lunch Reservation", description="Cancels a booked cafeteria lunch reservation.")

    selected = runtime.selector.select(runtime, [loan, lunch], intent=Intent(text="Review this loan applicant's credit"))
    assert loan in selected
    assert lunch not in selected
    record = runtime.reasoning_log.for_stage("selection")[0]
    assert record.model == binding.model_id


@requires_anthropic_key
def test_delegator_assigns_the_best_suited_persona_with_llm():
    from ear import ModelBinding, Skill, Workflow

    binding = ModelBinding(provider="anthropic", model=os.environ.get("ANTHROPIC_TEST_MODEL", "claude-haiku-4-5"))
    binding.activate()
    runtime = Runtime(name="Delegating-Runtime", model_binding=binding)

    guru = Persona(name="Credit Risk Guru", instructions="Underwrite conservatively.")
    guru.add_skill(Skill(name="assign_risk_grade", prompt="Grade the credit profile A-E."))
    advocate = Persona(name="Customer Advocate", instructions="Write plainly and kindly to applicants.")
    advocate.add_skill(Skill(name="write_customer_note", prompt="Draft a courteous decision note."))

    workflow = Workflow(name="Underwriting Workflow")
    workflow.add_step("Assign a risk grade.", persona=guru)
    workflow.add_persona(advocate)
    workflow.add_step("Write the customer note announcing the decision.")  # undelegated

    runtime.delegator.delegate(runtime, Intent(text="Underwrite a loan"), [workflow])
    assert workflow.steps[1].persona is advocate
    record = runtime.reasoning_log.for_stage("delegation")[0]
    assert "Customer Advocate" in record.output


# ---------------------------------------------------------------------------
# Subagent spawning.
# ---------------------------------------------------------------------------


def test_runtime_spawns_a_persona_scoped_subagent():
    runtime = Runtime(name="Parent-Runtime")
    persona = Persona(name="Credit Risk Guru", instructions="Underwrite conservatively.")
    decision = runtime.spawn(persona, "Review this application's risk")
    assert "Credit Risk Guru" in decision
    assert len(runtime.spawner.spawned) == 1
    # The subagent's cycle is recorded in its own memory, not the parent's.
    assert len(runtime.memory) == 0
    assert len(runtime.spawner.spawned[0].memory) == 1


def test_spawner_enforces_the_strategy_limit():
    runtime = Runtime(name="Parent-Runtime", spawner=Spawner(limit=1))
    persona = Persona(name="Guru")
    runtime.spawn(persona, "first")
    with pytest.raises(PermissionError, match="Subagent limit of 1"):
        runtime.spawn(persona, "second")


def test_spawner_disabled_by_strategy_refuses_to_spawn():
    runtime = Runtime(name="Parent-Runtime", spawner=Spawner(enabled=False))
    with pytest.raises(PermissionError, match="disabled"):
        runtime.spawn(Persona(name="Guru"), "anything")


# ---------------------------------------------------------------------------
# The shipped example stack loads and reasons end to end, offline.
# ---------------------------------------------------------------------------


def test_example_credit_risk_stack_loads_and_reasons(tmp_path, monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    directory = tmp_path / "credit_risk_stack"
    shutil.copytree(EXAMPLE_STACK, directory)

    runtime = load_runtime(directory)
    assert runtime.name == "Credit Risk Enterprise Runtime"
    assert runtime.memory.capacity == 30
    assert runtime.spawner.limit == 4
    assert [process.name for process in runtime.processes] == ["Underwrite Consumer Loan"]
    workflow = runtime.processes[0].workflows[0]
    assert len(workflow.steps) == 4
    assert workflow.steps[3].persona.name == "Customer Advocate"
    assert [policy.name for policy in runtime.policies] == ["Debt-to-Income Ceiling", "Fair Lending Control"]
    assert [policy.name for policy in workflow.policies] == ["Loan Amount Cap", "Large Loan Human Approval"]
    assert workflow.policies[1].approval_required is True

    decision = runtime.reason(
        Intent(
            text="Underwrite a $20,000 consumer loan",
            context={"loan_amount": 20000, "debt_to_income": 0.30, "credit_score": 720},
        )
    )
    assert "Credit Risk Guru" in decision

    with pytest.raises(PermissionError, match="Loan Amount Cap"):
        runtime.reason(Intent(text="Underwrite a large loan", context={"loan_amount": 90000}))


# ---------------------------------------------------------------------------
# N2: evaluation & knowledge depth -- BM25 narrowing, the persisted gist
# index, URL knowledge sources, report history with regression diffs,
# rubric criteria, and A/B stack comparison.
# ---------------------------------------------------------------------------


def test_bm25_ranks_rare_terms_above_repeated_common_ones():
    from ear import Knowledge, Passage

    knowledge = Knowledge(
        passages=[
            Passage(
                source="manual § plants",
                text="Loan loan loan loan loan loan loan paperwork is filed on Fridays with the loan clerk.",
            ),
            Passage(source="manual § subordination", text="A subordination agreement reprioritizes the loan lien."),
        ]
    )
    # "subordination" is rare (idf-heavy); the first passage's seven
    # repetitions of the common word "loan" saturate instead of winning.
    ranked = knowledge.candidates("subordination of a loan", limit=2)
    assert ranked[0].source == "manual § subordination"


def test_gist_bridges_the_synonym_phrasing_word_matching_misses():
    from ear import Knowledge, Passage

    debt = Passage(
        source="manual § 4.2",
        text="A grade C applicant whose debt-to-income band is not low must be declined.",
    )
    plants = Passage(source="manual § 9.9", text="Water the office plants on Fridays.")
    knowledge = Knowledge(passages=[plants, debt])
    query = "heavy existing borrowings"

    # Without the gist index the synonym phrasing shares no word with any
    # passage: no signal, so the corpus-order fallback surfaces plants.
    assert knowledge.narrowing() == "BM25 over passage text alone (no gist index)"
    assert knowledge.candidates(query, limit=1)[0].source == "manual § 9.9"

    debt.gist = "whether an applicant with heavy existing borrowings or high debt can be approved for a loan"
    assert knowledge.narrowing() == "BM25 over passage text and index gists"
    assert knowledge.candidates(query, limit=1)[0].source == "manual § 4.2"


def test_gist_index_persists_by_content_hash_and_retires_on_edit(tmp_path):
    from ear import Knowledge

    document = "# M\n\n## Ratios\n\nDebt banding governs declines.\n\n## Plants\n\nWater plants on Fridays.\n"
    knowledge = Knowledge().add_document("manual", "m.md", document)
    lm = ScriptedLM(
        ["## gist\n\nhow heavy borrowings force a decline\n", "## gist\n\ncaring for office greenery\n"]
    )
    assert knowledge.build_gists(lm) == 2

    index_path = tmp_path / ".ear" / "index.md"
    knowledge.write_index(index_path, model_label="anthropic/test")
    index_text = index_path.read_text(encoding="utf-8")
    assert "anthropic/test" in index_text and "office greenery" in index_text

    # A fresh load of the same corpus reuses every gist without a model.
    reloaded = Knowledge().add_document("manual", "m.md", document)
    assert reloaded.load_index(index_path) == 2
    assert not reloaded.missing_gists()

    # Editing one section retires only that entry: the edited passage
    # hashes differently, misses the index, and needs re-gisting.
    edited = document.replace("Water plants", "Water plants twice")
    partially = Knowledge().add_document("manual", "m.md", edited)
    assert partially.load_index(index_path) == 1
    assert [passage.source for passage in partially.missing_gists()] == ["manual -- m.md § Plants"]


def test_url_knowledge_source_fetches_once_and_reuses_the_cache(tmp_path, monkeypatch):
    fetches = []

    def fake_fetch(url, timeout=60):
        fetches.append(url)
        return "# Brief\n\n## Outlook\n\nRates are expected to hold.\n"

    monkeypatch.setattr("ear.loader.fetch_text", fake_fetch)
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Knowledge\n\n- market brief: https://example.com/brief.md\n"
    directory = write_stack(tmp_path / "stack", **stack)

    runtime = load_runtime(directory)
    assert fetches == ["https://example.com/brief.md"]
    assert (directory / ".ear" / "knowledge" / "market-brief.md").exists()
    assert any("Rates are expected to hold" in p.text for p in runtime.librarian.knowledge.passages)

    # No refresh cadence declared: the cache stands indefinitely.
    load_runtime(directory)
    assert len(fetches) == 1


def test_url_knowledge_source_honors_the_declared_refresh_cadence(tmp_path, monkeypatch):
    from ear.llm import LMError

    fetches = []

    def fake_fetch(url, timeout=60):
        fetches.append(url)
        if len(fetches) == 3:
            raise LMError("network down")
        return "# Brief\n\nRates hold.\n"

    monkeypatch.setattr("ear.loader.fetch_text", fake_fetch)
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Knowledge\n\n- market brief: https://example.com/brief.md, refetch weekly\n"
    directory = write_stack(tmp_path / "stack", **stack)

    load_runtime(directory)
    load_runtime(directory)  # fresh cache -> no refetch
    assert len(fetches) == 1

    cached = directory / ".ear" / "knowledge" / "market-brief.md"
    eight_days_ago = time.time() - 8 * 86400
    os.utime(cached, (eight_days_ago, eight_days_ago))
    load_runtime(directory)  # stale by the declared week -> refetch
    assert len(fetches) == 2

    os.utime(cached, (eight_days_ago, eight_days_ago))
    runtime = load_runtime(directory)  # refetch fails -> the stale cache stands
    assert len(fetches) == 3
    assert any("Rates hold" in p.text for p in runtime.librarian.knowledge.passages)


def test_url_knowledge_source_with_no_cache_fails_loudly(tmp_path, monkeypatch):
    from ear.llm import LMError

    def failing_fetch(url, timeout=60):
        raise LMError("unreachable")

    monkeypatch.setattr("ear.loader.fetch_text", failing_fetch)
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Knowledge\n\n- market brief: https://example.com/brief.md\n"
    with pytest.raises(ValueError, match="no cached copy"):
        load_runtime(write_stack(tmp_path / "stack", **stack))


def test_retrieval_record_names_its_narrowing_basis(tmp_path):
    directory = knowledge_stack(tmp_path)
    runtime = load_runtime(directory)
    runtime.reason(Intent(text="Underwrite a marginal grade C applicant", context={"loan_amount": 9000}))
    retrieval = runtime.reasoning_log.for_stage("retrieval")[0]
    assert retrieval.inputs["narrowing"] == "BM25 over passage text alone (no gist index)"


def test_reports_archive_and_diff_newly_failing_and_passing(tmp_path):
    from ear import Examiner

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    evaluations = directory / "evaluations"
    evaluations.mkdir()
    passing = "# Underwrite a loan\n\n## Context\n\n- loan_amount: 5000\n\n## Expected\n\n- decision: resolved\n"
    failing = passing.replace("resolved", "unobtainium")
    (evaluations / "steady.md").write_text(passing, encoding="utf-8")
    (evaluations / "flaky.md").write_text(passing, encoding="utf-8")

    runtime = load_runtime(directory)
    first = Examiner().examine(runtime, evaluations)
    assert first.prior_verdicts is None  # no history yet -> nothing to diff
    assert "Changes Since Last Report" not in (evaluations / "report.md").read_text(encoding="utf-8")
    assert len(list((evaluations / "reports").glob("*.md"))) == 1

    (evaluations / "flaky.md").write_text(failing, encoding="utf-8")
    second = Examiner().examine(runtime, evaluations)
    assert second.changes() == {"newly failing": ["flaky"], "newly passing": [], "still failing": []}

    third = Examiner().examine(runtime, evaluations)
    assert third.changes()["still failing"] == ["flaky"]

    (evaluations / "flaky.md").write_text(passing, encoding="utf-8")
    fourth = Examiner().examine(runtime, evaluations)
    assert fourth.changes() == {"newly failing": [], "newly passing": ["flaky"], "still failing": []}
    report = (evaluations / "report.md").read_text(encoding="utf-8")
    assert "- newly passing: flaky" in report
    assert len(list((evaluations / "reports").glob("*.md"))) == 4


def test_rubric_criteria_are_ungraded_offline_never_faked(tmp_path):
    from ear import Examiner

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    evaluations = directory / "evaluations"
    evaluations.mkdir()
    (evaluations / "rubric.md").write_text(
        "# Underwrite a loan\n\n## Context\n\n- loan_amount: 5000\n\n"
        "## Expected\n\n- decision: resolved\n- names the responsible persona\n- states a concrete amount\n",
        encoding="utf-8",
    )
    runtime = load_runtime(directory)
    examination = Examiner().examine(runtime, evaluations)

    result = examination.results[0]
    assert result.passed is True  # the field expectation still grades structurally
    assert [criterion.passed for criterion in result.criteria] == [None, None]
    assert all("needs a model" in criterion.rationale for criterion in result.criteria)
    report = (evaluations / "report.md").read_text(encoding="utf-8")
    assert "Rubric:" in report and "- ungraded: names the responsible persona" in report


def test_rubric_criteria_grade_separately_and_a_failed_criterion_fails(tmp_path):
    from ear import Examiner, ModelBinding

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    runtime = load_runtime(directory)
    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM()
    runtime.model_binding = binding

    graded, rationale, criteria = Examiner()._grade(
        runtime,
        expected_prose="The loan is approved.",
        expected_fields={},
        criteria=["cites the underwriting manual", "names a concrete amount"],
        outcome="Approved at $5,000 per the manual.",
    )
    # ScriptedLM without a script answers '## decision' sections, which the
    # bool field reads as unparseable -- so script the three judgments.
    binding.lm = ScriptedLM(
        [
            "## passed\n\nyes\n\n## rationale\n\nMatches the expectation.\n",
            "## passed\n\nno\n\n## rationale\n\nNo citation appears.\n",
            "## passed\n\nyes\n\n## rationale\n\n$5,000 is concrete.\n",
        ]
    )
    graded, rationale, criteria = Examiner()._grade(
        runtime,
        expected_prose="The loan is approved.",
        expected_fields={},
        criteria=["cites the underwriting manual", "names a concrete amount"],
        outcome="Approved at $5,000 per the manual.",
    )
    assert graded is False  # the failed criterion fails the evaluation
    assert [criterion.passed for criterion in criteria] == [False, True]
    assert "No citation" in criteria[0].rationale


def test_compare_refuses_without_a_model(tmp_path):
    from ear import Examiner

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    evaluations = directory / "evaluations"
    evaluations.mkdir()
    (evaluations / "any.md").write_text("# Underwrite\n\n## Expected\n\nApproved.\n", encoding="utf-8")
    runtime_a = load_runtime(directory, name="A")
    runtime_b = load_runtime(directory, name="B")
    with pytest.raises(ValueError, match="never written down"):
        Examiner().compare(runtime_a, runtime_b, evaluations)
    assert not (evaluations / "comparison.md").exists()


def test_compare_prefers_on_the_record_with_a_scripted_judge(tmp_path):
    from ear import Examiner, ModelBinding

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    evaluations = directory / "evaluations"
    evaluations.mkdir()
    (evaluations / "small-loan.md").write_text(
        "# Underwrite a loan\n\n## Context\n\n- loan_amount: 5000\n\n## Expected\n\nThe loan is handled.\n",
        encoding="utf-8",
    )
    runtime_a = load_runtime(directory, name="Stack-A")
    runtime_b = load_runtime(directory, name="Stack-B")
    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM(
        ["## preference\n\nB\n\n## rationale\n\nB states the amount plainly.\n"]
    )

    # The referee is a dedicated judge binding, independent of the two
    # contestants -- both runtimes reason with their deterministic
    # pipeline, and the script is consumed by the preference alone.
    comparison = Examiner().compare(runtime_a, runtime_b, evaluations, judge=binding)
    assert comparison.counts() == {"A": 0, "B": 1, "tie": 0}
    assert comparison.results[0].preference == "B"

    rendered = (evaluations / "comparison.md").read_text(encoding="utf-8")
    assert "A: Stack-A vs B: Stack-B" in rendered and "Preferred B: 1" in rendered
    record = runtime_a.reasoning_log.for_stage("comparison")[0]
    assert record.output == "B" and "amount plainly" in record.rationale


@requires_anthropic_key
def test_gist_index_builds_once_and_reloads_live(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = (
        KNOWLEDGE_MEMORY
        + "\n## Model Selection\n\nReason with anthropic/claude-haiku-4-5. "
        "The key lives in ANTHROPIC_API_KEY.\n"
    )
    directory = write_stack(tmp_path / "stack", **stack)
    (directory / "knowledge").mkdir()
    (directory / "knowledge" / "manual.md").write_text(MANUAL, encoding="utf-8")

    runtime = load_runtime(directory)
    index_path = directory / ".ear" / "index.md"
    assert index_path.exists()
    assert index_path.read_text(encoding="utf-8").count("## Passage") == 2
    indexing = runtime.reasoning_log.for_stage("indexing")[0]
    assert indexing.inputs["gists_written"] == 2
    assert indexing.input_tokens and indexing.output_tokens  # indexing is on the record, billed

    # The corpus indexes once: a fresh load reuses every gist by content
    # hash and asks the model for nothing.
    reloaded = load_runtime(directory)
    assert not reloaded.reasoning_log.for_stage("indexing")
    assert not reloaded.librarian.knowledge.missing_gists()
    assert reloaded.librarian.knowledge.narrowing() == "BM25 over passage text and index gists"


def test_model_id_at_sentence_end_keeps_no_period():
    strategy = Strategy.from_markdown(
        "## Model Selection\n\nReason with anthropic/claude-haiku-4-5. The key lives in ANTHROPIC_API_KEY.\n"
    )
    assert strategy.model == "anthropic/claude-haiku-4-5"
    assert strategy.provider == "anthropic"


# ---------------------------------------------------------------------------
# N3: execution depth -- prose-authored routing, leg retry policies, the
# journey runner with escalation, event documents, and dynamic panels.
# ---------------------------------------------------------------------------

ROUTED_WORKFLOW = (
    "# Workflows\n\n## Underwriting Workflow\n\n"
    "Routes: if the risk grade is C or worse, skip straight to the decline note; "
    "otherwise continue in order.\n\n"
    "1. Band the profile and assign a risk grade. (Credit Risk Guru)\n"
    "2. Prepare the approval paperwork. (Credit Risk Guru)\n"
    "3. Write the decline note. (Credit Risk Guru)\n\n"
    "Policies: Loan Amount Cap\n"
)


def test_routes_retries_and_escalation_are_authored_in_prose(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = ROUTED_WORKFLOW.replace(
        "Routes:", "Retries: retry a failed leg twice before giving up.\nRoutes:"
    )
    stack["policy"] = MINIMAL_STACK["policy"].replace(
        "Fallback: loan_amount <= 75000\nApplies to: Underwriting Workflow\n",
        "Fallback: loan_amount <= 75000\nApplies to: Underwriting Workflow\n"
        "Approval: required\nEscalate: after 3 days\n",
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    workflow = runtime.processes[0].workflows[0]
    assert "skip straight to the decline note" in workflow.routes
    assert workflow.retry_budget == 2
    cap = workflow.policies[0]
    assert cap.escalation == "after 3 days" and cap.escalation_days == 3.0

    bad = dict(MINIMAL_STACK)
    bad["workflow"] = MINIMAL_STACK["workflow"].replace(
        "Policies:", "Retries: whenever it feels right\nPolicies:"
    )
    with pytest.raises(ValueError, match="no readable count"):
        load_runtime(write_stack(tmp_path / "bad-retries", **bad))

    bad_policy = dict(MINIMAL_STACK)
    bad_policy["policy"] = MINIMAL_STACK["policy"].replace(
        "Applies to: runtime\n", "Applies to: runtime\nEscalate: someday, surely\n"
    )
    with pytest.raises(ValueError, match="no readable period"):
        load_runtime(write_stack(tmp_path / "bad-escalate", **bad_policy))


def test_strategy_reads_the_leg_retry_budget_from_prose():
    strategy = Strategy.from_markdown(
        "## Execution Resilience\n\nRetry a failed leg twice before giving up.\n"
    )
    assert strategy.leg_retry_budget == 2
    assert "Retry a failed leg" in strategy.execution
    assert Strategy.from_markdown("## Execution\n\nNo retries; fail fast.\n").leg_retry_budget == 0


def test_routed_journey_offline_continues_in_order_and_says_so(tmp_path):
    from ear import Journey

    stack = dict(MINIMAL_STACK)
    stack["workflow"] = ROUTED_WORKFLOW
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    journey = Journey(directory / "journeys" / "loan.md")
    status = journey.run(runtime, Intent(text="Underwrite a loan", context={"loan_amount": 5000}))

    assert status == "completed"
    assert [leg.step for leg in journey.legs] == [1, 2, 3]
    routings = runtime.reasoning_log.for_stage("routing")
    assert len(routings) == 3
    assert all("no model bound" in record.rationale for record in routings)
    assert routings[-1].output == "conclude the journey"


def test_routing_judgment_never_invents_a_step_and_honors_the_revisit_budget(tmp_path):
    from ear import Journey, ModelBinding, Workflow

    binding = ModelBinding(provider="anthropic", model="test")
    runtime = Runtime(name="Routed-Runtime", model_binding=binding)
    workflow = Workflow(name="W", routes="loop until done")
    workflow.add_step("first")
    workflow.add_step("second")
    authored = [(workflow, step) for step in workflow.steps]
    journey = Journey(tmp_path / "j.md", intent_text="x")

    from ear.journey import Leg

    leg = Leg(number=1, workflow="W", instruction="first", decision="d", status="decided", step=1)

    # A step number the stack never authored is refused, never improvised.
    binding.lm = ScriptedLM(["## next step\n\n9\n\n## rationale\n\nJump far.\n"])
    assert journey._route(runtime, workflow, authored, leg, {1: 1}) == 2
    assert "names no authored step" in runtime.reasoning_log.for_stage("routing")[-1].rationale

    # A legal loop is honored -- until the revisit budget refuses it.
    binding.lm = ScriptedLM(["## next step\n\n1\n\n## rationale\n\nDo it again.\n"])
    assert journey._route(runtime, workflow, authored, leg, {1: 1}) == 1
    binding.lm = ScriptedLM(["## next step\n\n1\n\n## rationale\n\nAgain, forever.\n"])
    assert journey._route(runtime, workflow, authored, leg, {1: 3}) == 2
    assert "revisit budget" in runtime.reasoning_log.for_stage("routing")[-1].rationale

    # 'conclude' ends the journey; 'next' continues in order.
    binding.lm = ScriptedLM(["## next step\n\nconclude\n\n## rationale\n\nNothing left.\n"])
    assert journey._route(runtime, workflow, authored, leg, {1: 1}) is None
    binding.lm = ScriptedLM(["## next step\n\nnext\n\n## rationale\n\nIn order.\n"])
    assert journey._route(runtime, workflow, authored, leg, {1: 1}) == 2


def test_journey_retries_a_failing_leg_within_the_declared_budget(tmp_path):
    from ear import Journey

    stack = dict(MINIMAL_STACK)
    stack["workflow"] = MINIMAL_STACK["workflow"].replace(
        "Policies:", "Retries: retry a failed leg twice before giving up.\nPolicies:"
    )
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    attempts = {"count": 0}
    true_reason = runtime.reason

    def flaky_reason(intent, approval=None):
        attempts["count"] += 1
        if attempts["count"] <= 2:
            raise RuntimeError("transient worker failure")
        return true_reason(intent, approval=approval)

    runtime.reason = flaky_reason
    journey = Journey(directory / "journeys" / "loan.md")
    status = journey.run(runtime, Intent(text="Underwrite a loan", context={"loan_amount": 5000}))

    assert status == "completed"
    retries = runtime.reasoning_log.for_stage("retry")
    assert len(retries) == 2
    assert all("retrying" in record.output for record in retries)
    assert retries[0].inputs["error"] == "transient worker failure"


def test_journey_fails_on_the_record_when_the_retry_budget_is_exhausted(tmp_path):
    from ear import Journey

    stack = dict(MINIMAL_STACK)
    stack["workflow"] = MINIMAL_STACK["workflow"].replace(
        "Policies:", "Retries: retry once.\nPolicies:"
    )
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)

    def doomed_reason(intent, approval=None):
        raise RuntimeError("the worker never comes back")

    runtime.reason = doomed_reason
    journey = Journey(directory / "journeys" / "loan.md")
    status = journey.run(runtime, Intent(text="Underwrite a loan", context={"loan_amount": 5000}))

    assert status == "FAILED"
    assert journey.legs[-1].status == "FAILED"
    assert "retry budget exhausted" in journey.legs[-1].decision
    record = (directory / "journeys" / "loan.md").read_text(encoding="utf-8")
    assert "Status: FAILED" in record
    assert "exhausted" in runtime.reasoning_log.for_stage("retry")[-1].output
    # Settled: another run replays nothing.
    assert Journey(journey.path).run(runtime) == "FAILED"


def test_journey_without_a_declared_budget_keeps_crash_semantics(tmp_path):
    from ear import Journey

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    runtime = load_runtime(directory)

    def crashing_reason(intent, approval=None):
        raise RuntimeError("worker died")

    runtime.reason = crashing_reason
    with pytest.raises(RuntimeError, match="worker died"):
        Journey(directory / "journeys" / "loan.md").run(
            runtime, Intent(text="Underwrite", context={"loan_amount": 5000})
        )
    assert runtime.reasoning_log.for_stage("retry") == []


def test_journey_consumes_event_documents_exactly_once(tmp_path):
    from ear import Journey

    directory = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    runtime = load_runtime(directory)
    journeys = directory / "journeys"

    # Crash after the first leg so the journey waits mid-walk.
    calls = {"count": 0}
    true_reason = runtime.reason

    def crash_on_second(intent, approval=None):
        calls["count"] += 1
        if calls["count"] == 2:
            raise RuntimeError("worker died mid-journey")
        return true_reason(intent, approval=approval)

    runtime.reason = crash_on_second
    with pytest.raises(RuntimeError):
        Journey(journeys / "loan.md").run(runtime, Intent(text="Underwrite", context={"loan_amount": 5000}))

    (journeys / "events").mkdir()
    (journeys / "events" / "loan-appraisal.md").write_text(
        "# Event -- appraisal received\n\n## Context\n\n- appraisal_value: 250000\n",
        encoding="utf-8",
    )
    runtime.reason = true_reason
    resumed = Journey(journeys / "loan.md")
    assert resumed.run(runtime) == "completed"
    assert resumed.context["appraisal_value"] == 250000
    event = runtime.reasoning_log.for_stage("event")[0]
    assert event.inputs["event"] == "loan-appraisal.md"
    record = (journeys / "loan.md").read_text(encoding="utf-8")
    assert "## Events" in record and "loan-appraisal.md" in record
    assert "appraisal_value: 250000" in record

    # A fresh run consumes nothing twice.
    again = Journey(journeys / "loan.md")
    again.run(runtime)
    assert len(runtime.reasoning_log.for_stage("event")) == 1


def gated_stack(tmp_path, escalate=""):
    # One step per workflow: an approval waives the gate for one governed
    # cycle, so the gated walk is a single leg -- the same shape as
    # approval_stack above.
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = (
        "# Workflows\n\n## Underwriting Workflow\n\n"
        "1. Decide approve or decline. (Credit Risk Guru)\n"
    )
    stack["policy"] = (
        "# Policies\n\n## Loan Amount Cap\nThe loan must not exceed $75,000.\n\n"
        "Fallback: loan_amount <= 75000\nApplies to: Underwriting Workflow\n\n"
        "## Large Loan Approval\nLoans over $50,000 need a human.\n\n"
        "Fallback: loan_amount <= 50000\nApplies to: Underwriting Workflow\n"
        "Approval: required\n" + (f"Escalate: {escalate}\n" if escalate else "")
    )
    return write_stack(tmp_path / "stack", **stack)


def test_runner_resumes_releases_and_escalates_in_one_pass(tmp_path):
    from ear import Journey, Journeys

    directory = gated_stack(tmp_path, escalate="after 3 days")
    runtime = load_runtime(directory)
    journeys = directory / "journeys"

    # a: crashes on its first leg -- the record already exists, in progress.
    def crash(intent, approval=None):
        raise RuntimeError("died")

    true_reason = runtime.reason
    runtime.reason = crash
    with pytest.raises(RuntimeError):
        Journey(journeys / "a.md").run(runtime, Intent(text="Underwrite", context={"loan_amount": 5000}))
    runtime.reason = true_reason

    # b: parks on the approval gate; a human verdict awaits in approvals/.
    assert (
        Journey(journeys / "b.md").run(runtime, Intent(text="Underwrite big", context={"loan_amount": 60000}))
        == "PENDING APPROVAL"
    )
    (journeys / "approvals").mkdir()
    (journeys / "approvals" / "b.md").write_text(
        "# Approval\n\nVerdict: approved\nApprover: lakshmi@gigkri.com\n\n> Collateral covers it.\n",
        encoding="utf-8",
    )

    # c: parks on the same gate with no approval -- and a declared deadline.
    assert (
        Journey(journeys / "c.md").run(runtime, Intent(text="Underwrite big", context={"loan_amount": 61000}))
        == "PENDING APPROVAL"
    )

    four_days = 4 * 86400
    outcomes = Journeys().run_all(runtime, journeys, now=time.time() + four_days)
    assert outcomes["a.md"] == "completed"
    assert outcomes["b.md"] == "completed"
    assert outcomes["c.md"] == "ESCALATED"

    c_record = (journeys / "c.md").read_text(encoding="utf-8")
    assert "Status: ESCALATED" in c_record and "## Escalation" in c_record
    assert "deadline has passed" in c_record
    escalation = runtime.reasoning_log.for_stage("escalation")[0]
    assert escalation.inputs["journey"] == "c.md"

    # A second pass is idempotent: settled stays settled, escalated stays
    # escalated with no duplicate record, and an approval still releases c.
    again = Journeys().run_all(runtime, journeys, now=time.time() + four_days)
    assert again == {"a.md": "completed", "b.md": "completed", "c.md": "ESCALATED"}
    assert len(runtime.reasoning_log.for_stage("escalation")) == 1

    (journeys / "approvals" / "c.md").write_text(
        "# Approval\n\nVerdict: approved\nApprover: lakshmi@gigkri.com\n", encoding="utf-8"
    )
    released = Journeys().run_all(runtime, journeys, now=time.time() + four_days)
    assert released["c.md"] == "completed"


def test_runner_refuses_an_unreadable_approval_on_the_record(tmp_path):
    from ear import Journey, Journeys

    directory = gated_stack(tmp_path)
    runtime = load_runtime(directory)
    journeys = directory / "journeys"
    Journey(journeys / "big.md").run(runtime, Intent(text="Underwrite big", context={"loan_amount": 60000}))
    (journeys / "approvals").mkdir()
    (journeys / "approvals" / "big.md").write_text("# Approval\n\nVerdict: perhaps\n", encoding="utf-8")

    outcomes = Journeys().run_all(runtime, journeys)
    assert outcomes["big.md"] == "PENDING APPROVAL"
    refusal = runtime.reasoning_log.for_stage("approval")[-1]
    assert "unreadable verdict" in refusal.output


def test_panel_speaker_is_judged_and_concludes_early_after_all_have_spoken():
    from ear import ModelBinding, Panel

    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM(
        [
            # Turn 1: the model tries to conclude before anyone spoke -- refused.
            "## speaker\n\nconclude\n\n## rationale\n\nSeems obvious.\n",
            "## statement\n\nGrade C; decline.\n",
            # Turn 2: the model picks the advocate by name.
            "## speaker\n\nCustomer Advocate\n\n## rationale\n\nThe other side must answer.\n",
            "## statement\n\nAgreed -- decline, with a referral.\n",
            # Turn 3: both have spoken; consensus concludes the panel early.
            "## speaker\n\nconclude\n\n## rationale\n\nBoth agree on decline.\n",
            "## decision\n\nDecline, with a referral to the secured product.\n",
        ]
    )
    runtime = Runtime(name="Panel-Runtime", model_binding=binding)
    personas = [Persona(name="Credit Risk Guru"), Persona(name="Customer Advocate")]
    decision = Panel(rounds=3).convene(runtime, personas, Intent(text="Underwrite a marginal loan"))

    turns = runtime.reasoning_log.for_stage("conversation")
    assert len(turns) == 2  # concluded early, well under the budget of 6
    assert "conclusion refused" in turns[0].inputs["chosen_by"]
    assert turns[1].inputs["chosen_by"] == "model"
    assert turns[1].inputs["choice_rationale"] == "The other side must answer."
    deliberation = runtime.reasoning_log.for_stage("deliberation")[0]
    assert deliberation.inputs["concluded_early"] == "Both agree on decline."
    assert decision == "Decline, with a referral to the secured product."


def test_panel_rotation_stands_when_the_choice_names_nobody():
    from ear import ModelBinding, Panel

    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM(
        [
            "## speaker\n\nThe Moderator\n\n## rationale\n\nA phantom.\n",
            "## statement\n\nFirst word.\n",
            "## speaker\n\nB\n\n## rationale\n\nB's turn.\n",
            "## statement\n\nSecond word.\n",
            "## decision\n\nSettled.\n",
        ]
    )
    runtime = Runtime(name="Panel-Runtime", model_binding=binding)
    personas = [Persona(name="A"), Persona(name="B")]
    Panel(rounds=1).convene(runtime, personas, Intent(text="anything"))
    turns = runtime.reasoning_log.for_stage("conversation")
    assert [record.inputs["speaker"] for record in turns] == ["A", "B"]
    assert "names no listed persona" in turns[0].inputs["chosen_by"]


def test_panel_persona_uses_tools_inside_a_turn_on_the_record(tmp_path):
    from ear import ModelBinding, Panel

    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Tools\n\n- credit lookup: fetch the applicant's bureau score\n"
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    runtime.tool_binder.bind("credit lookup", lambda applicant_id: f"score for {applicant_id}: 668")

    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM(
        [
            "## speaker\n\nAnalyst\n\n## rationale\n\nFacts first.\n",
            # The turn calls the tool, then speaks with the fact in hand.
            "## tool\n\ncredit lookup\n\n## arguments\n\n- applicant_id: 42\n\n## statement\n\n\n",
            "## tool\n\n\n\n## arguments\n\n\n\n## statement\n\nThe bureau score is 668; marginal.\n",
            "## decision\n\nDecline at 668.\n",
        ]
    )
    runtime.model_binding = binding
    personas = [Persona(name="Analyst")]
    decision = Panel(rounds=1).convene(runtime, personas, Intent(text="Underwrite"))

    tool_record = runtime.reasoning_log.for_stage("tool")[0]
    assert tool_record.inputs["tool"] == "credit lookup"
    assert "668" in tool_record.output
    turn = runtime.reasoning_log.for_stage("conversation")[0]
    assert turn.output == "The bureau score is 668; marginal."
    assert decision == "Decline at 668."


@requires_anthropic_key
def test_journey_routes_the_authored_skip_live(tmp_path):
    from ear import Journey

    stack = dict(MINIMAL_STACK)
    stack["workflow"] = ROUTED_WORKFLOW
    stack["memory"] = LIVE_MEMORY
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    journey = Journey(directory / "journeys" / "marginal.md")
    status = journey.run(
        runtime,
        Intent(
            text="Underwrite a $9,000 loan for a marginal applicant",
            context={"loan_amount": 9000, "credit_score": 580, "debt_to_income": 0.42},
        ),
    )

    assert status == "completed"
    steps_walked = [leg.step for leg in journey.legs]
    assert steps_walked[0] == 1
    assert 2 not in steps_walked  # the approval paperwork was skipped
    assert 3 in steps_walked
    routing = runtime.reasoning_log.for_stage("routing")[0]
    assert routing.model == "anthropic/claude-haiku-4-5"


@requires_anthropic_key
def test_panel_concludes_early_on_consensus_live(tmp_path):
    stack = dict(PANEL_STACK)
    stack["memory"] = LIVE_MEMORY
    stack["workflow"] = PANEL_STACK["workflow"].replace(
        "Pattern: adversarial debate; the Credit Risk Guru has the last word",
        "Pattern: one short turn each, then conclude immediately once both have spoken -- "
        "do not keep talking past agreement",
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    runtime.panel.rounds = 3  # budget of six turns the consensus should beat
    decision = runtime.reason(
        Intent(
            text="Underwrite a $2,000 loan for a prime applicant with excellent credit",
            context={"loan_amount": 2000, "credit_score": 810, "debt_to_income": 0.05},
        )
    )
    turns = runtime.reasoning_log.for_stage("conversation")
    assert 2 <= len(turns) < 6  # concluded before the budget
    assert decision and "no model bound" not in decision


# ---------------------------------------------------------------------------
# N4: governance & connectivity depth -- approver allow-lists, tool-scoped
# policies, the hash-chained tamper-evident trail, retention + the usage
# ledger, and the native MCP client.
# ---------------------------------------------------------------------------


def test_policy_reads_approvers_and_gates_by_the_allow_list(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["workflow"] = "# Workflows\n\n## Underwriting Workflow\n\n1. Decide. (Credit Risk Guru)\n"
    stack["policy"] = (
        "# Policies\n\n## Loan Amount Cap\nThe loan must not exceed $75,000.\n\n"
        "Fallback: loan_amount <= 75000\nApplies to: runtime\n\n"
        "## Large Loan Approval\nLoans over $50,000 need a senior underwriter.\n\n"
        "Fallback: loan_amount <= 50000\nApplies to: runtime\n"
        "Approval: required\nApprovers: senior@bank.com, Chief Risk Officer\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    policy = {p.name: p for p in runtime.policies}["Large Loan Approval"]
    assert policy.approvers == ["senior@bank.com", "Chief Risk Officer"]
    assert policy.approver_allowed("SENIOR@bank.com") is True
    assert policy.approver_allowed("chief-risk-officer") is True
    assert policy.approver_allowed("intern@bank.com") is False


def test_off_list_approver_is_refused_and_the_gate_stays_parked(tmp_path):
    from ear import Approval, Journey, Journeys

    stack = dict(MINIMAL_STACK)
    stack["workflow"] = "# Workflows\n\n## Underwriting Workflow\n\n1. Decide. (Credit Risk Guru)\n"
    stack["policy"] = (
        "# Policies\n\n## Loan Amount Cap\nThe loan must not exceed $75,000.\n\n"
        "Fallback: loan_amount <= 75000\nApplies to: Underwriting Workflow\n\n"
        "## Large Loan Approval\nLoans over $50,000 need a senior underwriter.\n\n"
        "Fallback: loan_amount <= 50000\nApplies to: Underwriting Workflow\n"
        "Approval: required\nApprovers: senior@bank.com\n"
    )
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)
    journeys = directory / "journeys"

    parked = Journey(journeys / "big.md")
    assert parked.run(runtime, Intent(text="Underwrite big", context={"loan_amount": 60000})) == "PENDING APPROVAL"

    # An off-list approver waives nothing: the gate stays parked, refused
    # on the record.
    off_list = Approval(verdict=True, approver="intern@bank.com")
    assert Journey(journeys / "big.md").run(runtime, approval=off_list) == "PENDING APPROVAL"
    refusal = [r for r in runtime.reasoning_log.for_stage("approval") if r.output.startswith("REFUSED")]
    assert refusal and "intern@bank.com" in refusal[-1].output

    # The listed approver releases it.
    on_list = Approval(verdict=True, approver="senior@bank.com")
    assert Journey(journeys / "big.md").run(runtime, approval=on_list) == "completed"


def test_tool_scoped_policy_blocks_one_call_and_returns_it_to_the_model(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Tools\n\n- wire transfer: move money out of the ledger\n"
    stack["policy"] = MINIMAL_STACK["policy"] + (
        "\n## Transfer Cap\nThe wire transfer tool must never move more than $10,000.\n\n"
        "Fallback: amount <= 10000\nApplies to: tools\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    assert [p.name for p in runtime.tool_policies] == ["Transfer Cap"]
    runtime.tool_binder.bind("wire transfer", lambda amount: f"moved {amount}")

    from ear.tool_binder import BoundTool

    tool = BoundTool(name="wire transfer", description="move money", handler=lambda amount: f"moved {amount}")
    invoke = runtime.tool_binder.logged_handler(runtime, tool)

    # Over the cap: blocked, and the refusal comes back as text (not a raise).
    blocked = invoke(amount=50000)
    assert "blocked by policy 'Transfer Cap'" in blocked
    record = runtime.reasoning_log.for_stage("tool")[-1]
    assert record.output.startswith("BLOCKED") and record.inputs["policy"] == "Transfer Cap"

    # Within the cap: it runs.
    assert invoke(amount=5000) == "moved 5000"


def test_the_trail_is_hash_chained_and_verify_catches_a_hand_edit(tmp_path):
    from ear.reasoning_log import ReasoningLog

    for extension in (".md", ".jsonl"):
        path = tmp_path / f"trail{extension}"
        runtime = Runtime(name="Chained-Runtime")
        runtime.reasoning_log.path = str(path)
        runtime.reason(Intent(text="first cycle", context={"amount": 1}))
        runtime.reason(Intent(text="second cycle", context={"amount": 2}))

        ok, message = ReasoningLog.verify(str(path))
        assert ok is True and "intact" in message

        # Hand-edit one record's content; the chain breaks at that record.
        text = path.read_text(encoding="utf-8")
        tampered = text.replace("second cycle", "smuggled cycle", 1)
        assert tampered != text
        path.write_text(tampered, encoding="utf-8")
        broken, why = ReasoningLog.verify(str(path))
        assert broken is False and "broken chain" in why


def test_the_chain_continues_across_a_resumed_session(tmp_path):
    from ear.reasoning_log import ReasoningLog

    path = tmp_path / "trail.md"
    first = Runtime(name="Session-One")
    first.reasoning_log.path = str(path)
    first.reasoning_log.resume()
    first.reason(Intent(text="cycle in session one"))

    second = Runtime(name="Session-Two")
    second.reasoning_log.path = str(path)
    second.reasoning_log.resume()  # picks up the chain tip and cycle number
    second.reason(Intent(text="cycle in session two"))

    ok, message = ReasoningLog.verify(str(path))
    assert ok is True, message
    assert "## Cycle 2" in path.read_text(encoding="utf-8")  # numbering continued


def test_retention_rotates_old_cycles_with_a_note_and_stays_verifiable(tmp_path):
    from datetime import datetime, timedelta, timezone

    from ear.reasoning_log import ReasoningLog

    path = tmp_path / "trail.md"
    runtime = Runtime(name="Retained-Runtime")
    runtime.reasoning_log.path = str(path)
    runtime.reason(Intent(text="ancient cycle"))
    for record in runtime.reasoning_log.records:
        record.timestamp = datetime.now(timezone.utc) - timedelta(days=100)
    runtime.reason(Intent(text="fresh cycle"))

    rotated = runtime.reasoning_log.rotate(90.0)
    assert rotated > 0
    stages = [r.stage for r in runtime.reasoning_log.records]
    assert stages[0] == "retention"  # the note stands in for what was rotated
    assert not runtime.reasoning_log.for_cycle(1)  # the ancient cycle is gone
    note = runtime.reasoning_log.records[0]
    assert "rotated" in note.output and note.inputs["rotated_records"] == rotated

    ok, message = ReasoningLog.verify(str(path))  # the rewrite re-chained cleanly
    assert ok is True, message


def test_retention_applies_from_a_plain_cycle_with_no_journey_involved(tmp_path):
    """A declared retention window used to take effect only when the
    Journey runner happened to pass over it -- a plain Runtime.reason()
    cycle silently ignored it. Runtime.reason now applies the same
    rotation itself after every cycle, so retention holds whether or not
    Journeys are ever used."""
    from datetime import datetime, timedelta, timezone

    from ear.strategy import Strategy

    path = tmp_path / "trail.md"
    runtime = Runtime(name="Retained-Runtime", strategy=Strategy(retention_days=90.0))
    runtime.reasoning_log.path = str(path)
    runtime.reason(Intent(text="ancient cycle"))
    for record in runtime.reasoning_log.records:
        record.timestamp = datetime.now(timezone.utc) - timedelta(days=100)

    # No Journeys.run_all(), no manual .rotate() call -- just an ordinary
    # cycle. Retention must apply on its own.
    runtime.reason(Intent(text="fresh cycle"))

    stages = [r.stage for r in runtime.reasoning_log.records]
    assert "retention" in stages
    assert not runtime.reasoning_log.for_cycle(1)  # the ancient cycle is gone


def test_usage_ledger_renders_from_the_trail_with_dollars_when_priced():
    from ear import ModelBinding

    binding = ModelBinding(provider="anthropic", model="test")
    binding.lm = ScriptedLM(["## complies\n\nyes\n\n## rationale\n\nfine\n"])
    runtime = Runtime(name="Ledger-Runtime", model_binding=binding)
    runtime.strategy = Strategy.from_markdown(
        "## Pricing\n\nInput tokens cost $3 per million; output tokens cost $15 per million.\n"
    )
    from ear import Policy

    runtime.add_policy(Policy(name="Cap", statement="Stay under the cap."))
    runtime.reason(Intent(text="a priced cycle", context={"amount": 1}))

    report = runtime.reasoning_log.usage_report(strategy=runtime.strategy)
    assert "# Usage Report" in report
    assert "| Cycle |" in report and "**total**" in report
    assert "$0." in report  # a dollar figure, because Pricing was declared

    offline = Runtime(name="Unpriced").reasoning_log.usage_report()
    assert "—" in offline or "total" in offline  # no pricing -> no invented dollars


# -- the native MCP client ---------------------------------------------------

FAKE_MCP_SERVER = str(Path(__file__).resolve().parent / "fixtures" / "fake_mcp_server.py")


def mcp_command() -> str:
    import sys

    return f"{sys.executable} {FAKE_MCP_SERVER}"


def test_native_mcp_client_handshakes_lists_and_calls_over_stdio():
    from ear.mcp_client import McpClient, McpError, command_words

    with McpClient(command_words(mcp_command())) as client:
        tools = client.list_tools()
        assert [(tool.name, tool.parameters) for tool in tools] == [("add", ["a", "b"])]
        assert client.call_tool("add", {"a": 3, "b": 4}) == "sum is 7"
        with pytest.raises(McpError, match="error"):
            client.call_tool("nonexistent", {})


def test_mcp_client_fails_loudly_on_a_bad_command():
    from ear.mcp_client import McpClient, McpError

    with pytest.raises(McpError, match="could not launch"):
        McpClient(["definitely-not-a-real-binary-xyz"]).connect()


def test_mcp_client_timeout_does_not_spawn_a_stray_reader():
    """A client-side timeout used to spawn a new reader thread per call and
    leave the timed-out call's thread alive, still blocked on stdout,
    racing every later call's own newly-spawned reader for the same bytes
    -- occasionally stealing and discarding a later call's genuine
    response. The fix is structural: one persistent reader for the whole
    connection, so there is never a second thread to race. Proving the
    invariant (thread count never grows) is more reliable than trying to
    reproduce the race's nondeterministic symptom directly."""
    import threading

    from ear.mcp_client import McpClient, McpError, command_words

    client = McpClient(command_words(mcp_command()), timeout=0.05)
    client.connect()
    try:
        baseline = threading.active_count()
        with pytest.raises(McpError, match="did not answer"):
            client.call_tool("sleep", {"seconds": 1.0})
        assert threading.active_count() == baseline

        # The fake server is single-threaded and still busy sleeping from
        # the call above; raise the timeout back up so the next call can
        # wait that out, then confirm it gets its own, correct response --
        # not a stray thread's leftovers.
        client.timeout = 5.0
        assert client.call_tool("add", {"a": 10, "b": 5}) == "sum is 15"
    finally:
        client.close()


def test_declared_mcp_server_binds_tools_into_a_cycle_on_the_record(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = f"## MCP\n\n- calc: arithmetic over stdio `{mcp_command()}`\n"
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    assert [server.name for server in runtime.strategy.mcp_servers] == ["calc"]

    bound = runtime.connect_mcp()
    try:
        assert [(tool.name, tool.parameters) for tool in bound] == [("calc: add", ["a", "b"])]
        # The bound MCP tool joins the cycle's toolset and runs through the
        # same logged handler as any native tool.
        tool = runtime.tool_binder.bound_tools(runtime)[0]
        assert tool.name == "calc: add"
        result = runtime.tool_binder.logged_handler(runtime, tool)(a=8, b=9)
        assert result == "sum is 17"
        record = runtime.reasoning_log.for_stage("tool")[-1]
        assert record.inputs["tool"] == "calc: add" and record.output == "sum is 17"
    finally:
        runtime.disconnect_mcp()
    assert runtime.tool_binder.mcp_tools == []


def test_tool_scoped_policy_governs_an_mcp_call(tmp_path):
    from ear import Policy

    stack = dict(MINIMAL_STACK)
    stack["memory"] = f"## MCP\n\n- calc: arithmetic over stdio `{mcp_command()}`\n"
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    # A tool-scoped policy that forbids the MCP tool by name.
    runtime.tool_policies.append(
        Policy(name="No Adding", statement="never", fallback_expression="tool != 'calc: add'")
    )
    bound = runtime.connect_mcp()
    try:
        blocked = runtime.tool_binder.logged_handler(runtime, bound[0])(a=1, b=2)
        assert "blocked by policy 'No Adding'" in blocked
    finally:
        runtime.disconnect_mcp()


# ---------------------------------------------------------------------------
# The runtime dashboard: a self-contained HTML view of the reasoning trail,
# the TensorBoard-equivalent, rendered from the log with zero dependencies.
# ---------------------------------------------------------------------------


def test_dashboard_renders_self_contained_html_from_a_runtime(tmp_path):
    from ear import Dashboard

    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    runtime.reason(Intent(text="Underwrite a consumer loan", context={"loan_amount": 5000}))
    runtime.reason(Intent(text="Underwrite another loan", context={"loan_amount": 6000}))

    html = Dashboard().render(runtime)
    # Self-contained: a real document, no external fetches of any kind.
    assert html.startswith("<!doctype html>")
    assert "<style>" in html and "<svg" in html
    assert "http://" not in html and "https://" not in html
    assert "cdn" not in html.lower() and "<script src" not in html
    # It shows the runtime and its cycles.
    assert "Credit Risk Runtime" in html
    assert html.count('class="cycle"') == 2
    assert 'class="tile"' in html and "Cycles" in html
    # And it writes the same document to disk.
    path = tmp_path / "board.html"
    written = Dashboard().write(runtime, path)
    assert path.read_text(encoding="utf-8") == written


def test_dashboard_integrity_badge_reflects_the_hash_chain(tmp_path):
    from ear import Dashboard

    trail = tmp_path / "trail.jsonl"
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    runtime.reasoning_log.path = str(trail)
    runtime.reason(Intent(text="Underwrite a loan", context={"loan_amount": 5000}))

    assert "badge-good" in Dashboard().render(runtime)  # chain intact

    tampered = trail.read_text(encoding="utf-8").replace("Underwrite", "Tampered", 1)
    trail.write_text(tampered, encoding="utf-8")
    assert "badge-bad" in Dashboard().render(runtime)  # verify catches the edit


def test_dashboard_surfaces_governance_and_tool_calls(tmp_path):
    from ear import Dashboard

    stack = dict(MINIMAL_STACK)
    stack["memory"] = "## Tools\n\n- credit lookup: fetch the bureau score\n"
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    runtime.tool_binder.bind("credit lookup", lambda applicant: "score 700")

    from ear.tool_binder import BoundTool

    runtime.tool_binder.logged_handler(
        runtime, BoundTool(name="credit lookup", description="x", handler=lambda applicant: "score 700")
    )(applicant="a1")
    runtime.reason(Intent(text="Underwrite a loan", context={"loan_amount": 5000}))
    with pytest.raises(PermissionError):
        runtime.reason(Intent(text="Underwrite oversized", context={"loan_amount": 90000}))

    html = Dashboard().render(runtime)
    assert "Governance" in html and "Loan Amount Cap" in html
    assert "Tool calls" in html and "credit lookup" in html
    assert "v-bad" in html  # the blocked policy is flagged


def test_dashboard_rebuilds_from_a_persisted_jsonl_trail(tmp_path):
    from ear import Dashboard
    from ear.reasoning_log import ReasoningLog

    trail = tmp_path / "trail.jsonl"
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    runtime.reasoning_log.path = str(trail)
    runtime.reason(Intent(text="Underwrite a loan", context={"loan_amount": 5000}))
    original = len(runtime.reasoning_log.records)

    # Lossless reconstruction from disk, and a dashboard straight from the path.
    rebuilt = ReasoningLog.from_trail(str(trail))
    assert len(rebuilt.records) == original
    assert rebuilt.records[0].stage == "intent"
    html = Dashboard().render(str(trail))
    assert html.startswith("<!doctype html>") and 'class="cycle"' in html


def test_fleet_dashboard_shows_every_runtime_with_health_and_progress(tmp_path):
    from ear import Dashboard

    # alpha: healthy, two clean cycles.
    alpha = load_runtime(write_stack(tmp_path / "alpha", **MINIMAL_STACK))
    alpha.reason(Intent(text="Underwrite one", context={"loan_amount": 5000}))
    alpha.reason(Intent(text="Underwrite two", context={"loan_amount": 6000}))

    # beta: an approval gate parks a cycle -> needs attention (pending).
    beta_stack = dict(MINIMAL_STACK)
    beta_stack["policy"] = (
        "# Policies\n\n## Loan Amount Cap\nUnder 75k.\n\n"
        "Fallback: loan_amount <= 75000\nApplies to: Underwriting Workflow\n\n"
        "## Big Loan\nOver 50k needs a human.\n\n"
        "Fallback: loan_amount <= 50000\nApplies to: Underwriting Workflow\nApproval: required\n"
    )
    beta = load_runtime(write_stack(tmp_path / "beta", **beta_stack))
    with pytest.raises(PermissionError):
        beta.reason(Intent(text="Underwrite big", context={"loan_amount": 60000}))

    # gamma: a persisted trail, hand-edited -> broken chain.
    gamma = load_runtime(write_stack(tmp_path / "gamma", **MINIMAL_STACK))
    gamma.reasoning_log.path = str(tmp_path / "gamma.jsonl")
    gamma.reason(Intent(text="Underwrite", context={"loan_amount": 5000}))
    trail = tmp_path / "gamma.jsonl"
    trail.write_text(trail.read_text(encoding="utf-8").replace("Underwrite", "Tampered", 1), encoding="utf-8")

    html = Dashboard().render_fleet({"alpha": alpha, "beta": beta, "gamma": gamma})

    # Self-contained, one page, one card per runtime.
    assert html.startswith("<!doctype html>")
    assert "http://" not in html and "https://" not in html
    assert html.count('class="run"') == 3
    assert "alpha" in html and "beta" in html and "gamma" in html
    # The fleet overview counts health, and the worst status wins the badge.
    assert "Healthy" in html and "Attention" in html and "Broken" in html
    assert "badge-bad" in html  # gamma's broken chain drives the fleet badge
    # Progress is visible per runtime: sparklines and cross-run charts.
    assert html.count('class="spark') == 3
    assert "Compare runtimes" in html
    # Each runtime drills down into its full board (its cycles are embedded).
    assert 'class="run-body"' in html and html.count('class="cycle"') >= 3


def test_fleet_classifies_health_broken_beats_attention_beats_healthy(tmp_path):
    from ear.dashboard import _classify

    assert _classify((False, "broken chain"), [], 0, 0, 0)[0] == "broken"
    assert _classify((True, "ok"), ["exporter down"], 0, 0, 0)[0] == "attention"
    assert _classify((True, "ok"), [], 1, 0, 0)[0] == "attention"  # a failed cycle
    assert _classify((True, "ok"), [], 0, 1, 0)[0] == "attention"  # awaiting approval
    healthy, reason = _classify((True, "ok"), [], 0, 0, 2)  # policy blocks are governance working
    assert healthy == "healthy" and "governance working" in reason
    assert _classify(None, [], 0, 0, 0) == ("healthy", "all clear")


def test_fleet_dashboard_from_a_directory_of_trails(tmp_path):
    from ear import Dashboard
    from ear.reasoning_log import ReasoningLog

    trails = tmp_path / "trails"
    trails.mkdir()
    for name in ("payments", "lending"):
        runtime = load_runtime(write_stack(tmp_path / name, **MINIMAL_STACK))
        runtime.reasoning_log.path = str(trails / f"{name}.jsonl")
        runtime.reason(Intent(text=f"a {name} cycle", context={"loan_amount": 5000}))

    html = Dashboard().render_fleet(str(trails))  # one run per JSONL file, rebuilt from disk
    assert html.count('class="run"') == 2
    assert "payments" in html and "lending" in html
    # A chain badge per rebuilt run, verified from the file.
    assert "✓ chain" in html


def test_gantt_lays_cycles_on_a_time_axis_with_status_and_a_now_marker(tmp_path):
    from ear import Dashboard

    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    runtime.reason(Intent(text="Underwrite one", context={"loan_amount": 5000}))
    runtime.reason(Intent(text="Underwrite two", context={"loan_amount": 6000}))
    with pytest.raises(PermissionError):
        runtime.reason(Intent(text="oversized", context={"loan_amount": 90000}))

    html = Dashboard().render_gantt(runtime)
    assert html.startswith("<!doctype html>")
    assert "http://" not in html and "https://" not in html.replace("http-equiv", "")
    # A bar per cycle, on a wall-clock axis, with a "now" marker.
    assert 'class="gantt"' in html
    assert html.count('class="gbar') == 3
    assert 'class="gbar bad"' in html  # the blocked cycle is red
    assert "nowline" in html and ">now<" in html
    # Time ticks on the axis (HH:MM:SS labels).
    import re

    assert re.search(r">\d{2}:\d{2}:\d{2}<", html)


def test_gantt_tickle_injects_an_auto_refresh_and_off_by_default(tmp_path):
    from ear import Dashboard

    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    runtime.reason(Intent(text="Underwrite", context={"loan_amount": 5000}))

    ticking = Dashboard().render_gantt(runtime, refresh=3)
    assert '<meta http-equiv="refresh" content="3">' in ticking  # the tickle
    still = Dashboard().render_gantt(runtime)  # no refresh by default
    assert "http-equiv" not in still


def test_fleet_gantt_has_one_lane_per_runtime(tmp_path):
    from ear import Dashboard

    alpha = load_runtime(write_stack(tmp_path / "alpha", **MINIMAL_STACK))
    alpha.reason(Intent(text="a1", context={"loan_amount": 5000}))
    alpha.reason(Intent(text="a2", context={"loan_amount": 6000}))
    beta = load_runtime(write_stack(tmp_path / "beta", **MINIMAL_STACK))
    beta.reason(Intent(text="b1", context={"loan_amount": 7000}))

    html = Dashboard().render_gantt({"alpha": alpha, "beta": beta}, refresh=5)
    assert '<meta http-equiv="refresh" content="5">' in html
    assert 'class="gantt"' in html
    assert "alpha" in html and "beta" in html
    # alpha ran two cycles, beta one -> three bars across two lanes.
    assert html.count('class="gbar') == 3


def test_freshness_heartbeat_distinguishes_active_idle_stale():
    from datetime import datetime, timedelta, timezone

    from ear.dashboard import _freshness

    now = datetime.now(timezone.utc)
    assert _freshness(now - timedelta(seconds=3), now) == "active"
    assert _freshness(now - timedelta(minutes=10), now) == "idle"
    assert _freshness(now - timedelta(hours=3), now) == "stale"
    assert _freshness(None, now) == "idle"


def test_fleet_card_shows_a_freshness_heartbeat(tmp_path):
    from datetime import datetime, timezone

    from ear import Dashboard

    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    runtime.reason(Intent(text="Underwrite", context={"loan_amount": 5000}))
    # Render "as of now" -> the just-run cycle reads active.
    html = Dashboard().render_fleet({"only": runtime}, now=datetime.now(timezone.utc))
    assert "fresh-active" in html and ">active<" in html


# ---------------------------------------------------------------------------
# The live dashboard server: reachable for as long as a human wants it, not
# just until the driving script's own work finishes -- create_server() is a
# thread-friendly split of the old blocking serve(), plus two new routes:
# a download link for whatever lands in the sandbox's outputs/ (confined,
# so a crafted path can never escape it) and a page-button POST /shutdown
# that stops the server without touching a single file on disk.
# ---------------------------------------------------------------------------


def test_outputs_section_lists_files_confined_to_the_sandbox(tmp_path):
    from ear.dashboard import _human_size, _list_outputs, _outputs_section

    outputs = tmp_path / "outputs"
    outputs.mkdir()
    (outputs / "dashboard.xlsx").write_bytes(b"x" * 2048)
    (outputs / "validation_log.md").write_text("# log\n")

    files = _list_outputs(outputs)
    assert dict(files) == {"dashboard.xlsx": 2048, "validation_log.md": 6}
    assert _human_size(2048) == "2.0 KB"
    assert _human_size(500) == "500 B"

    html_text = _outputs_section(outputs)
    assert "dashboard.xlsx" in html_text and "/download/dashboard.xlsx" in html_text
    assert "No outputs yet" not in html_text

    assert "No outputs yet" in _outputs_section(None)
    assert "No outputs yet" in _outputs_section(tmp_path / "does-not-exist")


def test_live_dashboard_server_downloads_confines_paths_and_shuts_down(tmp_path):
    """End to end against a real HTTPServer on an OS-assigned free port:
    the main page is reachable and lists an output, a download returns the
    exact bytes with an attachment header, a path-traversal attempt 404s
    rather than escaping outputs/, and POSTing /shutdown stops the server
    -- and only the server; the file on disk survives untouched."""
    import threading
    import urllib.error
    import urllib.request

    from ear import Sandbox
    from ear.dashboard import create_server
    from ear.reasoning_log import ReasoningLog

    class _FakeRuntime:
        pass

    sandbox = Sandbox.create(root=str(tmp_path / "box"), name="LiveDashboardTest")
    outputs = sandbox.resolve("outputs")
    outputs.mkdir(parents=True, exist_ok=True)
    (outputs / "completed_dashboard.xlsx").write_bytes(b"fake completed workbook bytes")

    runtime = _FakeRuntime()
    runtime.sandbox = sandbox
    runtime.reasoning_log = ReasoningLog(path=None)
    runtime.strategy = None
    runtime.name = "LiveDashboardTest"

    server = create_server(runtime, port=0, host="127.0.0.1")
    port = server.server_address[1]
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        body = urllib.request.urlopen(f"http://127.0.0.1:{port}/", timeout=5).read().decode()
        assert "completed_dashboard.xlsx" in body
        assert 'action="/shutdown"' in body

        response = urllib.request.urlopen(f"http://127.0.0.1:{port}/download/completed_dashboard.xlsx", timeout=5)
        assert response.read() == b"fake completed workbook bytes"
        assert "attachment" in response.headers.get("Content-Disposition", "")

        with pytest.raises(urllib.error.HTTPError) as excinfo:
            urllib.request.urlopen(f"http://127.0.0.1:{port}/download/..%2F..%2Fsecrets.txt", timeout=5)
        assert excinfo.value.code == 404

        request = urllib.request.Request(f"http://127.0.0.1:{port}/shutdown", method="POST", data=b"")
        shutdown_response = urllib.request.urlopen(request, timeout=5)
        assert shutdown_response.status == 200
        assert b"stopped" in shutdown_response.read()
    finally:
        thread.join(timeout=5)

    assert not thread.is_alive()
    with pytest.raises(Exception):
        urllib.request.urlopen(f"http://127.0.0.1:{port}/", timeout=1)
    # The file on disk is exactly what shutdown promised it would leave alone.
    assert (outputs / "completed_dashboard.xlsx").read_bytes() == b"fake completed workbook bytes"


# ---------------------------------------------------------------------------
# Sandbox: each runtime instance in its own filesystem-confined, resource-
# limited workspace -- native, stdlib-only, on the record.
# ---------------------------------------------------------------------------


def test_sandbox_confines_the_filesystem_and_refuses_escapes(tmp_path):
    from ear import Sandbox, SandboxViolation

    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    assert box.list() == ["outputs/", "uploads/", "workspace/"]  # scaffold
    box.write_text("workspace/notes.md", "hello")
    assert box.read_text("workspace/notes.md") == "hello"

    for escape in ("../escape.txt", "/etc/passwd", "workspace/../../x"):
        with pytest.raises(SandboxViolation):
            box.resolve(escape)
    with pytest.raises(SandboxViolation):
        box.write_text("../outside.txt", "nope")


def test_sandbox_run_strips_secrets_caps_time_and_captures_output(tmp_path, monkeypatch):
    import sys

    from ear import Sandbox

    box = Sandbox.create(root=str(tmp_path / "box"), name="t", timeout=2.0)

    # A command runs, its output and exit code captured.
    ran = box.run([sys.executable, "-c", "print('hi from the box')"])
    assert ran.ok and "hi from the box" in ran.stdout and ran.returncode == 0

    # The ambient environment's secrets never reach the child.
    monkeypatch.setenv("EAR_SECRET_TEST", "topsecret")
    leaked = box.run([sys.executable, "-c", "import os; print(os.environ.get('EAR_SECRET_TEST', 'STRIPPED'))"])
    assert "STRIPPED" in leaked.stdout and "topsecret" not in leaked.stdout
    # But an explicitly passed variable is available.
    passed = box.run([sys.executable, "-c", "import os; print(os.environ.get('FOO', 'none'))"], env={"FOO": "bar"})
    assert passed.stdout.strip() == "bar"

    # A command over the wall-clock limit is killed and marked.
    slow = box.run([sys.executable, "-c", "import time; time.sleep(5)"], timeout=1)
    assert slow.timed_out is True


def test_shell_syntax_token_flags_operators_never_quoted_content():
    """The detector checked against whole argv tokens, never substrings --
    the precision that keeps a legitimate quoted argument (a grep pattern
    with a semicolon in it, say) from ever being a false positive."""
    from ear.sandbox import _shell_syntax_token

    assert _shell_syntax_token(["ls", "-la", "workspace/", "&&", "cat", "f"]) == "&&"
    assert _shell_syntax_token(["python3", "a.py", ";", "echo", "done"]) == ";"
    assert _shell_syntax_token(["cmd", "|", "tee", "out.txt"]) == "|"
    assert _shell_syntax_token(["python3", "a.py", ">", "out.txt"]) == ">"
    assert _shell_syntax_token(["python3", "a.py", "2>&1"]) == "2>&1"
    assert _shell_syntax_token(["echo", "$(date)"]) == "$(date)"
    assert _shell_syntax_token(["echo", "`date`"]) == "`date`"
    # A single token that merely *contains* a shell character, quoted as
    # part of legitimate content, is not a false positive.
    assert _shell_syntax_token(["grep", "a;b", "file.txt"]) is None
    assert _shell_syntax_token(["python3", "-c", "print(1 if x>1 else 0)"]) is None
    assert _shell_syntax_token(["ls", "-la"]) is None


def test_sandbox_run_rejects_shell_syntax_with_an_actionable_error(tmp_path):
    """A command string with shell operators fails fast and clearly --
    exit 126, an actionable stderr message -- instead of silently passing
    the operator as a literal argument and returning a confusing 'exit 0,
    no output' that used to cost several wasted turns figuring out (see
    examples/sales_mis_stack/logs/09-*.log for the real-world case this
    closes). Nothing is ever actually executed when this fires -- proven
    here by checking the redirect target was never created."""
    from ear import Sandbox

    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    result = box.run("ls -la workspace/ 2>&1 | tee out.txt; echo done")
    assert result.returncode == 126
    assert not result.ok
    assert "not a shell" in result.stderr or "literal argument" in result.stderr
    assert "&&" in result.stderr or "was passed as a literal argument" in result.stderr
    assert not (box.root / "out.txt").exists()  # the redirect never ran

    redirect = box.run("echo hi > workspace/should-not-exist.txt")
    assert redirect.returncode == 126
    assert not (box.root / "workspace" / "should-not-exist.txt").exists()


def test_sandbox_run_still_runs_a_plain_command_with_shell_looking_quoted_args(tmp_path):
    """The rejection is precise, not paranoid: a legitimate single command
    -- even given as one string, shlex-split same as always -- whose
    *argument* happens to contain a shell character (quoted, part of the
    content) still runs normally, in both the list and string call forms."""
    import sys

    from ear import Sandbox

    box = Sandbox.create(root=str(tmp_path / "box"), name="t")

    listed = box.run([sys.executable, "-c", "print('a;b>c')"])
    assert listed.ok and "a;b>c" in listed.stdout

    stringed = box.run(f'{sys.executable} -c "print(1 if 2>1 else 0)"')
    assert stringed.ok and stringed.stdout.strip() == "1"


def test_sandbox_tools_run_through_the_logged_handler_and_escapes_return_as_text(tmp_path):
    from ear import Runtime, Sandbox

    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    runtime = Runtime(name="Boxed")
    runtime.tool_binder.sandbox_tools = box.as_tools()
    tools = {tool.name: tool for tool in runtime.tool_binder.bound_tools(runtime)}
    assert set(tools) == {"read_file", "write_file", "list_files", "run_shell"}

    write = runtime.tool_binder.logged_handler(runtime, tools["write_file"])
    assert "wrote 4 characters" in write(path="outputs/o.txt", content="data")
    assert box.read_text("outputs/o.txt") == "data"

    # A path escape is a tool failure returned to the model, on the record.
    escaped = runtime.tool_binder.logged_handler(runtime, tools["read_file"])(path="../../secret")
    assert "escapes the sandbox" in escaped
    record = runtime.reasoning_log.for_stage("tool")[-1]
    assert record.inputs["tool"] == "read_file"


def test_sandbox_capabilities_checks_live_never_assumes(tmp_path):
    """'Fully capable' is a fact this establishes by actually running
    `--version` through the confined executor -- not a promise inherited
    from whatever the image was supposed to contain."""
    from ear import Sandbox

    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    report = box.capabilities(("python3", "definitely_not_a_real_binary_xyz"))

    assert report["python3"]["available"] is True
    assert report["python3"]["path"]
    assert "Python" in report["python3"]["version"] or report["python3"]["version"]

    assert report["definitely_not_a_real_binary_xyz"] == {"available": False}


def test_check_environment_tool_reports_python_and_node_live(tmp_path):
    """The environment_admin toolset's own descriptor promise ('inspect
    the system environment') -- exercised as a bound tool, not just the
    underlying Sandbox method."""
    stack = dict(MINIMAL_STACK)
    stack["memory"] = "# Memory & Strategy\n\n## Sandbox\n\nIsolate each runtime under `.ear/box`.\n"
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    bound = {tool.name: tool for tool in runtime.tool_binder.bound_tools(runtime)}
    assert "check_environment" in bound  # environment_admin defaults on

    report = bound["check_environment"].handler("python3,node")
    assert "python3: available" in report
    assert "node: available" in report


def test_strategy_reads_the_sandbox_section():
    strategy = Strategy.from_markdown(
        "## Sandbox\n\nIsolate each runtime under `.ear/box`. Shell commands time out after 20 seconds; "
        "limit memory to 256 MB. Expose file and shell tools.\n"
    )
    assert strategy.sandbox_enabled is True
    assert strategy.sandbox_root == ".ear/box"
    assert strategy.sandbox_timeout == 20.0
    assert strategy.sandbox_memory_mb == 256
    assert strategy.sandbox_tools is True

    off = Strategy.from_markdown("## Sandbox\n\nNo sandbox; run directly on the host.\n")
    assert off.sandbox_enabled is False


def test_loader_opens_a_sandbox_per_instance_and_records_it(tmp_path):
    stack = dict(MINIMAL_STACK)
    stack["memory"] = (
        "## Sandbox\n\nEach runtime runs in an isolated workspace under `.ear/box`; "
        "commands time out after 15 seconds. Expose a shell.\n"
    )
    directory = write_stack(tmp_path / "stack", **stack)
    runtime = load_runtime(directory)

    assert runtime.sandbox is not None
    assert Path(runtime.sandbox.root).exists()
    assert runtime.sandbox.timeout == 15.0
    assert [tool.name for tool in runtime.tool_binder.sandbox_tools] == [
        "read_file",
        "write_file",
        "list_files",
        "run_shell",
    ]
    opened = runtime.reasoning_log.for_stage("sandbox")[0]
    assert "opened at" in opened.output and str(runtime.sandbox.root) in opened.output

    # No Sandbox section -> no sandbox, host filesystem as before.
    plain = load_runtime(write_stack(tmp_path / "plain", **MINIMAL_STACK))
    assert plain.sandbox is None


def test_spawned_subagents_nest_their_own_child_sandbox(tmp_path):
    from ear import Persona

    stack = dict(MINIMAL_STACK)
    stack["memory"] = (
        "## Sandbox\n\nIsolate each runtime under `.ear/box`. Expose a shell.\n\n"
        "## Subagents\n\nSubagents may be spawned, up to 3.\n"
    )
    runtime = load_runtime(write_stack(tmp_path / "stack", **stack))
    runtime.spawner.spawn(runtime, Persona(name="Analyst"), Intent(text="do a thing"))

    child = runtime.spawner.spawned[-1]
    assert child.sandbox is not None
    assert child.sandbox.root != runtime.sandbox.root
    assert "subagents" in str(child.sandbox.root)
    # The child's box is nested under the parent's root.
    assert str(runtime.sandbox.root) in str(child.sandbox.root)
    assert [tool.name for tool in child.tool_binder.sandbox_tools]  # tools nested too


# ---------------------------------------------------------------------------
# Session Goals: a completion condition that drives the runtime forward with
# typed blockers, bounded autonomous continuation, on the trail.
# ---------------------------------------------------------------------------


def _goal_runtime(goal_replies, decisions=None):
    """A minimal runtime whose reason() is stubbed, so the only model calls
    are the goal judgments -- letting a scripted LM drive the loop
    deterministically."""
    from types import SimpleNamespace

    from ear.reasoning_log import ReasoningLog

    calls = {"n": 0}
    decisions = decisions or []
    runtime = SimpleNamespace(
        name="Goal-Fake",
        reasoning_log=ReasoningLog(),
        memory=None,
        model_binding=SimpleNamespace(lm=ScriptedLM(list(goal_replies)), model_id="anthropic/test"),
    )

    def reason(intent, approval=None):
        index = calls["n"]
        calls["n"] += 1
        return decisions[index] if index < len(decisions) else f"decision {index}"

    runtime.reason = reason
    runtime.calls = calls
    return runtime


def _goal_reply(satisfied, blocker="", evidence="e", next_step=""):
    return (
        f"## satisfied\n\n{satisfied}\n\n## blocker\n\n{blocker}\n\n"
        f"## evidence\n\n{evidence}\n\n## next_step\n\n{next_step}\n"
    )


def test_goal_satisfied_stops_immediately_on_the_record():
    from ear import GoalKeeper

    runtime = _goal_runtime([_goal_reply("yes", evidence="all criteria met")])
    outcome = GoalKeeper().pursue(runtime, "finish the analysis", Intent(text="start"))

    assert outcome.satisfied and outcome.status == "satisfied"
    assert outcome.continuations == 0 and runtime.calls["n"] == 1
    record = runtime.reasoning_log.for_stage("goal")[-1]
    assert record.output == "satisfied" and record.rationale == "all criteria met"


def test_goal_not_met_drives_a_bounded_autonomous_continuation():
    from ear import GoalKeeper

    runtime = _goal_runtime(
        [_goal_reply("no", "goal_not_met_yet", "still missing X", "compute X"), _goal_reply("yes", evidence="X done")]
    )
    outcome = GoalKeeper().pursue(runtime, "produce X", Intent(text="start"))

    # One continuation cycle, then satisfied -- two cycles total.
    assert outcome.status == "satisfied"
    assert outcome.continuations == 1 and runtime.calls["n"] == 2
    assert [e.blocker for e in outcome.history] == ["goal_not_met_yet", "satisfied"]


def test_goal_typed_blockers_stop_and_surface_the_reason():
    from ear import GoalKeeper

    for blocker in ("needs_user_input", "external_wait", "missing_evidence", "run_failed"):
        runtime = _goal_runtime([_goal_reply("no", blocker, f"blocked: {blocker}")])
        outcome = GoalKeeper().pursue(runtime, "g", Intent(text="start"))
        assert outcome.status == "blocked" and outcome.blocker == blocker
        assert runtime.calls["n"] == 1  # a stop blocker never continues


def test_goal_no_progress_breaker_stops_a_stuck_loop():
    from ear import GoalKeeper

    # The same non-progress verdict, over and over.
    runtime = _goal_runtime([_goal_reply("no", "goal_not_met_yet", "same", "same")] * 5)
    outcome = GoalKeeper().pursue(runtime, "g", Intent(text="start"))
    assert outcome.status == "exhausted" and outcome.blocker == "no_progress"
    assert outcome.continuations == 2  # default no-progress cap


def test_goal_continuation_budget_is_enforced_in_code():
    from ear import Goal, GoalKeeper

    # Distinct next steps each time (so the no-progress breaker never fires),
    # but the continuation budget of 2 stops it.
    runtime = _goal_runtime(
        [
            _goal_reply("no", "goal_not_met_yet", "e1", "s1"),
            _goal_reply("no", "goal_not_met_yet", "e2", "s2"),
            _goal_reply("no", "goal_not_met_yet", "e3", "s3"),
        ]
    )
    outcome = GoalKeeper().pursue(runtime, Goal(condition="g", max_continuations=2), Intent(text="start"))
    assert outcome.status == "exhausted" and outcome.blocker == "max_continuations"
    assert outcome.continuations == 2


def test_goal_maps_an_approval_gate_to_needs_user_input():
    from ear import GoalKeeper
    from ear.approval import ApprovalRequired

    runtime = _goal_runtime([])

    def park(intent, approval=None):
        raise ApprovalRequired([type("P", (), {"name": "Large Loan Approval"})()])

    runtime.reason = park
    outcome = GoalKeeper().pursue(runtime, "underwrite the loan", Intent(text="start"))
    assert outcome.status == "blocked" and outcome.blocker == "needs_user_input"


def test_goal_is_ungraded_offline_and_never_fabricates_a_continuation(tmp_path):
    runtime = load_runtime(write_stack(tmp_path / "stack", **MINIMAL_STACK))
    outcome = runtime.pursue("underwrite it", Intent(text="Underwrite", context={"loan_amount": 5000}))

    # No model -> the goal cannot be judged; stop honestly after one cycle.
    assert outcome.status == "ungraded" and outcome.blocker == "ungraded"
    assert outcome.continuations == 0
    record = runtime.reasoning_log.for_stage("goal")[-1]
    assert record.model == "" and "no model bound" in record.rationale


def test_goal_reads_from_markdown():
    from ear import Goal

    assert Goal.from_markdown("# Task\n\n## Goal\n\nFinish and pass all tests.\n").condition == "Finish and pass all tests."
    assert Goal.from_markdown("Just do the thing.").condition == "Just do the thing."


@requires_anthropic_key
def test_goal_pursued_live_reaches_a_terminal_outcome(tmp_path):
    directory = tmp_path / "credit_risk_stack"
    shutil.copytree(EXAMPLE_STACK, directory)
    memory = (directory / "memory.md").read_text(encoding="utf-8")
    (directory / "memory.md").write_text(
        memory.replace("anthropic/claude-opus-4-8", "anthropic/claude-haiku-4-5"), encoding="utf-8"
    )
    runtime = load_runtime(directory)
    outcome = runtime.pursue(
        "Reach a clear approve-or-decline decision for the applicant, with the risk grade stated.",
        Intent(
            text="Underwrite a $15,000 consumer loan",
            context={"loan_amount": 15000, "debt_to_income": 0.30, "credit_score": 720},
        ),
    )
    # It terminates cleanly, within the budget, and records each judgment.
    assert outcome.status in ("satisfied", "blocked", "exhausted")
    assert outcome.continuations <= 8
    assert runtime.reasoning_log.for_stage("goal")
    assert all(blocker in {
        "goal_not_met_yet", "needs_user_input", "external_wait", "missing_evidence", "run_failed", "satisfied"
    } for blocker in [e.blocker for e in outcome.history])


# ---------------------------------------------------------------------------
# Strict tool-call recovery: the native loop corrects a hallucinated or
# empty turn instead of silently abandoning it, bounded and on the record.
# ---------------------------------------------------------------------------


def _tool_action(tool="", args="", decision=""):
    return f"## tool\n\n{tool}\n\n## arguments\n\n{args}\n\n## decision\n\n{decision}\n"


def test_tool_loop_recovers_from_a_hallucinated_tool_name():
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    tool = BoundTool(name="calc", description="add one to x", handler=lambda x: f"result {int(x) + 1}")
    runtime = Runtime(name="Recover")
    lm = ScriptedLM(
        [
            _tool_action(tool="nonexistent_tool"),  # a tool that does not exist -> recover
            _tool_action(tool="calc", args="- x: 5"),  # the real call
            _tool_action(decision="Final answer is 6"),  # decide
        ]
    )
    decision = Reasoner._reason_with_tools(
        Intent(text="go"), runtime, lm, context={}, capabilities="none", tools=[tool], max_iterations=6
    )
    assert decision == "Final answer is 6"

    records = runtime.reasoning_log.for_stage("tool")
    # The hallucinated call is recorded as a recovery, then the real call runs.
    assert records[0].inputs["recovery"] is True and records[0].output.startswith("RECOVERED")
    assert "no tool named 'nonexistent_tool'" in records[0].output and "calc" in records[0].output
    assert records[1].inputs["tool"] == "calc" and "result 6" in records[1].output


def test_tool_loop_recovers_from_an_empty_turn():
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    tool = BoundTool(name="calc", description="x", handler=lambda x: str(x))
    runtime = Runtime(name="Recover2")
    lm = ScriptedLM([_tool_action(), _tool_action(decision="Decided after the nudge")])
    decision = Reasoner._reason_with_tools(
        Intent(text="go"), runtime, lm, context={}, capabilities="none", tools=[tool], max_iterations=6
    )
    assert decision == "Decided after the nudge"
    recovery = runtime.reasoning_log.for_stage("tool")[0]
    assert recovery.inputs["recovery"] is True and "no tool call and no decision" in recovery.output


def test_tool_loop_recovery_budget_is_bounded_then_concludes():
    from ear.reasoner import Reasoner
    from ear.reasoner import _MAX_TOOL_RECOVERIES
    from ear.tool_binder import BoundTool

    tool = BoundTool(name="calc", description="x", handler=lambda x: str(x))
    runtime = Runtime(name="Recover3")
    # Two recoveries, then a third bad turn trips the break; the last reply
    # concludes via the plain ReasonAboutIntent fallback.
    lm = ScriptedLM([_tool_action(tool="bad")] * 3 + ["## decision\n\nConcluded with what I have\n"])
    decision = Reasoner._reason_with_tools(
        Intent(text="go"), runtime, lm, context={}, capabilities="none", tools=[tool], max_iterations=8
    )
    assert decision == "Concluded with what I have"
    recoveries = [r for r in runtime.reasoning_log.for_stage("tool") if r.inputs.get("recovery")]
    assert len(recoveries) == _MAX_TOOL_RECOVERIES  # never unbounded


def test_tool_loop_refuses_an_identical_failed_retry_until_calibrated():
    """Attempt discipline: a failed call (tool + identical arguments) may
    fail only once. The unchanged retry is a blind retry by definition and
    is refused with calibration guidance, on the record. A *different*
    intervening call (the calibration) resets the failed-call guard."""
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    calls = []

    def run(cmd):
        calls.append(cmd)
        if "--fixed" in cmd:
            return "ok [exit 0, 10 ms]"
        return "boom [exit 1, 10 ms]"

    tool = BoundTool(name="run", description="run it", handler=run)
    runtime = Runtime(name="Discipline")
    same = _tool_action(tool="run", args="- cmd: python3 job.py")
    lm = ScriptedLM(
        [
            same,  # attempt 1 -- runs
            same,  # unchanged failed retry -- REFUSED, never executed
            _tool_action(tool="run", args="- cmd: python3 job.py --fixed"),  # calibrated -- runs
            _tool_action(tool="run", args="- cmd: python3 job.py --fixed"),  # successful repeat -- runs
            _tool_action(decision="Done after calibration"),
        ]
    )
    decision = Reasoner._reason_with_tools(
        Intent(text="go"), runtime, lm, context={}, capabilities="none", tools=[tool], max_iterations=10
    )
    assert decision == "Done after calibration"
    # The unchanged failed retry never reached the handler.
    assert calls == ["python3 job.py", "python3 job.py --fixed", "python3 job.py --fixed"]
    refusals = [r for r in runtime.reasoning_log.for_stage("tool") if "may fail only once" in str(r.output)]
    assert len(refusals) == 1
    assert "Calibrate first" in refusals[0].output


def test_tool_loop_streak_resets_after_an_intervening_calibration_call():
    """fix -> rerun -> fix -> rerun never trips the guard: rewriting the
    script between identical run commands is exactly the calibration the
    discipline demands."""
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    runs = []
    write = BoundTool(name="write_file", description="w", handler=lambda path, content: "wrote")
    run = BoundTool(name="run_shell", description="r", handler=lambda command: (runs.append(command), "exit 1")[1])
    runtime = Runtime(name="Discipline2")
    rerun = _tool_action(tool="run_shell", args="- command: python3 v.py")
    fix = _tool_action(tool="write_file", args="- path: v.py\n- content: print(1)")
    lm = ScriptedLM([rerun, fix, rerun, fix, rerun, _tool_action(decision="ok")])
    decision = Reasoner._reason_with_tools(
        Intent(text="go"), runtime, lm, context={}, capabilities="none", tools=[write, run], max_iterations=10
    )
    assert decision == "ok"
    assert len(runs) == 3  # every rerun followed a calibration; none refused
    assert not [r for r in runtime.reasoning_log.for_stage("tool") if "blind call" in str(r.output)]


def test_tool_loop_failed_calibration_does_not_unlock_original_failed_call():
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    runs = []
    writes = []
    run = BoundTool(name="run_shell", description="r", handler=lambda command: (runs.append(command), "[exit 1, 1 ms]")[1])
    write = BoundTool(
        name="write_file",
        description="w",
        handler=lambda path, content: (writes.append(path), "Tool 'write_file' failed: disk full")[1],
    )
    runtime = Runtime(name="Discipline3")
    rerun = _tool_action(tool="run_shell", args="- command: python3 v.py")
    failed_fix = _tool_action(tool="write_file", args="- path: v.py\n- content: print(1)")
    lm = ScriptedLM([rerun, failed_fix, rerun, _tool_action(decision="blocked")])

    decision = Reasoner._reason_with_tools(
        Intent(text="go"), runtime, lm, context={}, capabilities="none", tools=[write, run], max_iterations=10
    )

    assert decision == "blocked"
    assert runs == ["python3 v.py"]  # the original failing command was not allowed to fail twice
    assert writes == ["v.py"]
    refusals = [r for r in runtime.reasoning_log.for_stage("tool") if "may fail only once" in str(r.output)]
    assert len(refusals) == 1


# ---------------------------------------------------------------------------
# argument_blocks: a tool call's arguments can mix short '- name: value'
# bullets with a 'name:' + blockquote form for a value that needs more than
# one line -- a script's source, a whole file -- which a single bullet line
# can never carry. This is the fix for the native tool loop silently
# truncating a multi-line write_file/run_shell argument to whatever fit on
# one line.
# ---------------------------------------------------------------------------


def test_argument_blocks_keeps_short_bullets_working_unchanged():
    from ear.section import argument_blocks

    blocks = argument_blocks(["- applicant_id: 42", "- date: today"])
    # Names survive verbatim (never case/underscore-folded by `normalize`):
    # a tool argument becomes a Python keyword argument, so 'applicant_id'
    # must stay 'applicant_id', not fold into 'applicant id'.
    assert blocks == {"applicant_id": "42", "date": "today"}


def test_argument_blocks_carries_a_genuinely_multiline_value_intact():
    from ear.section import argument_blocks

    lines = [
        "- path: workspace/script.py",
        "content:",
        "> import openpyxl",
        ">",
        "> wb = openpyxl.load_workbook('x.xlsx')",
        "> for name in wb.sheetnames:",
        ">     if name:",
        ">         print(name)",
        ">     else:",
        ">         pass",
    ]
    blocks = argument_blocks(lines)
    assert blocks["path"] == "workspace/script.py"
    assert blocks["content"] == (
        "import openpyxl\n"
        "\n"
        "wb = openpyxl.load_workbook('x.xlsx')\n"
        "for name in wb.sheetnames:\n"
        "    if name:\n"
        "        print(name)\n"
        "    else:\n"
        "        pass"
    )


def test_argument_blocks_does_not_end_a_block_on_an_unquoted_blank_line():
    from ear.section import argument_blocks

    # The model forgot to prefix the blank line with '>' -- the value must
    # still survive whole, not truncate at the first line.
    lines = ["content:", "> line one", "", "> line two"]
    blocks = argument_blocks(lines)
    assert blocks["content"] == "line one\n\nline two"


def test_tool_loop_passes_a_multiline_script_through_write_file_intact():
    """End to end through the real ChooseToolAction wire format: a
    write_file-shaped tool call whose 'content' argument is a multi-line
    Python script must reach the handler byte-for-byte, not truncated to
    whatever fit on one bullet line."""
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    script = (
        "import openpyxl\n"
        "\n"
        "wb = openpyxl.load_workbook('uploads/data.xlsx')\n"
        "for name in wb.sheetnames:\n"
        "    print(name)\n"
    )
    received = {}

    def write_file(path, content):
        received["path"], received["content"] = path, content
        return f"wrote {len(content)} characters to {path}"

    tool = BoundTool(name="write_file", description="write a file", handler=write_file)
    runtime = Runtime(name="MultilineArgs")
    call = (
        "## tool\n\nwrite_file\n\n"
        "## arguments\n\n"
        "- path: workspace/script.py\n"
        "content:\n"
        + "\n".join(f"> {line}" if line else ">" for line in script.splitlines())
        + "\n\n## decision\n\n"
    )
    lm = ScriptedLM([call, _tool_action(decision="Script written and run.")])
    decision = Reasoner._reason_with_tools(
        Intent(text="write a script"), runtime, lm, context={}, capabilities="none", tools=[tool], max_iterations=6
    )
    assert decision == "Script written and run."
    assert received["path"] == "workspace/script.py"
    # A trailing blank line is inherently unrecoverable through blockquoting
    # (the same limitation `quote`/`unquote` already have everywhere else in
    # this codec) -- every other character, including internal blank lines
    # and indentation, must survive exactly.
    assert received["content"] == script.rstrip("\n")


def test_a_failed_tool_call_reaches_the_model_on_the_next_prompt():
    """Proof, not assertion: a tool call that raises must show up -- error
    text and all -- in the exact prompt the model sees on its *next* turn,
    not just in the internal `gathered` list. `ToolBinder._logged` turns the
    exception into 'Tool <name> failed: <error>' and the loop appends that
    into `gathered`, which is one of `ChooseToolAction`'s own input fields,
    so `render_prompt` places it in the prompt under its own heading. This
    test reads the *rendered prompt string itself* (`lm.prompts[1]`), not
    the intermediate variable, so it can't be fooled by a wiring bug between
    'the result was computed' and 'the result was actually put in front of
    the model'."""
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    attempts = {"count": 0}

    def flaky(path):
        attempts["count"] += 1
        if attempts["count"] == 1:
            raise RuntimeError("boom: no such file")
        return f"read {path} fine"

    tool = BoundTool(name="read_file", description="read a file", handler=flaky)
    runtime = Runtime(name="FailureFeedback")
    lm = ScriptedLM([
        _tool_action(tool="read_file", args="- path: missing.xlsx"),  # fails
        _tool_action(tool="read_file", args="- path: missing.xlsx"),  # unchanged retry refused
        _tool_action(decision="Read the file after one failed attempt."),
    ])
    decision = Reasoner._reason_with_tools(
        Intent(text="read a file"), runtime, lm, context={}, capabilities="none", tools=[tool], max_iterations=6
    )
    assert decision == "Read the file after one failed attempt."
    assert attempts["count"] == 1

    # The second prompt (the one the model saw right after the failure) must
    # literally contain the failure text -- this is the actual string handed
    # to the model, read back from the stub LM's own recorded history.
    second_prompt = lm.prompts[1]
    assert "Tool 'read_file' failed: boom: no such file" in second_prompt
    assert "## gathered" in second_prompt

    # The unchanged retry is refused before the handler runs, so the tool
    # only fails once in this cycle.
    refusals = [r for r in runtime.reasoning_log.for_stage("tool") if "may fail only once" in str(r.output)]
    assert len(refusals) == 1

    # And the trail's own tool-stage record shows the same failure, on the
    # audit trail an investigator would actually read.
    tool_records = runtime.reasoning_log.for_stage("tool")
    assert "FAILED" in tool_records[0].output and "boom: no such file" in tool_records[0].output


# ---------------------------------------------------------------------------
# ear.caveman: the deterministic, zero-dependency prose compressor that
# always runs on a tool result before it re-enters `gathered` -- ported from
# the MIT-licensed caveman-shrink MCP middleware. It can only delete matched
# filler words via `re.sub`, never generate replacement text, so it cannot
# hallucinate or garble a fact -- unlike an earlier LLM-based version of this
# feature, which once summarized a 907-row file as "1000 rows" and caused a
# downstream cycle to block on a discrepancy that was never real.
# ---------------------------------------------------------------------------


def test_caveman_drops_articles_filler_pleasantries_and_hedging():
    from ear.caveman import compress

    assert compress("The user is the owner of an account").text == "User is owner of account"

    result = compress("Sure, this just basically returns the value").text.lower()
    assert "sure" not in result and "just" not in result and "basically" not in result

    result = compress("I will perhaps connect to the database").text.lower()
    assert "perhaps" not in result and not result.startswith("i will")


def test_caveman_never_touches_protected_spans():
    from ear.caveman import compress

    # Fenced code, verbatim -- including filler words inside it.
    text = "See ```\nthe just sure return 1;\n``` for reference"
    assert "the just sure return 1;" in compress(text).text

    # Inline code, verbatim.
    assert "`the just basically API`" in compress("Call `the just basically API` now").text

    # URLs, verbatim.
    url = "https://example.com/the/just/api"
    assert url in compress(f"Fetch {url} please").text

    # Filesystem paths, verbatim.
    path = "/tmp/the/just/file.txt"
    assert path in compress(f"Wrote to {path} successfully").text

    # CONST_CASE identifiers and dotted calls, verbatim.
    assert "API_KEY_VALUE" in compress("Set the API_KEY_VALUE please").text
    assert "config.api.endpoint()" in compress("Just call config.api.endpoint() basically").text


def test_caveman_never_touches_a_bare_number():
    """The exact real-world failure this exists to prevent: a row count (or
    any other bare number) must survive compression byte-for-byte, because
    nothing in the compression rule set matches digits at all -- there is no
    generation step that could invent or round a number the way an LLM
    summarizer can."""
    from ear.caveman import compress

    text = (
        "The staged dataset was actually written successfully with 907 data "
        "rows and 1 header row, basically 908 lines total."
    )
    result = compress(text).text
    assert "907" in result and "908" in result and "1" in result


def test_caveman_degrades_safely_on_empty_or_non_string_input():
    from ear.caveman import compress

    assert compress("").text == ""
    assert compress("").saved_pct == 0.0


def test_caveman_returns_source_code_byte_for_byte():
    """The real-world failure this guards: a read_file of a healthy script
    fed back through prose compression lost `for a in anomalies:`'s loop
    variable to the article rule and its indentation to whitespace
    collapse -- so the next cycle 'repaired' a script that was never
    broken, on the clock and on the bill. Text that is mostly code must
    pass through untouched."""
    from ear.caveman import compress

    script = (
        "import openpyxl\n"
        "anomalies = []\n"
        "for a in anomalies:\n"
        "    if a is None:\n"
        "        continue\n"
        "    total = sum(1 for v in seen.values() if v > 1)\n"
    )
    assert compress(script).text == script


def test_caveman_preserves_indented_lines_and_restores_nested_sentinels_in_mixed_text():
    """Prose around code still compresses, but an indented line is
    structure, not prose -- and a protected span claimed inside another
    protected span must be spliced back fully, never left as a stray
    sentinel index the model reads as a phantom number."""
    from ear.caveman import compress

    code_line = "    total = sum(1 for v in seen.values() if v > 1)"
    text = f"The run just finished. Output was:\n{code_line}\nIt seems the check basically passed."
    result = compress(text).text
    assert code_line in result
    assert "\x00" not in result
    assert "basically" not in result.lower()


# ---------------------------------------------------------------------------
# Auxiliary Model: a second, cheaper ModelBinding memory.md may optionally
# declare to compress a tool result *further*, on top of the always-on
# ear.caveman pass above -- never a replacement for it. Declaring none (the
# default) still applies deterministic compression; declaring one must never
# shrink what actually lands on the audit trail, only what feeds the next
# prompt.
# ---------------------------------------------------------------------------


def test_strategy_reads_the_auxiliary_model_section_without_clobbering_the_primary():
    from ear import Strategy

    strategy = Strategy.from_markdown(
        "## Model Selection\n\nReason with anthropic/claude-opus-4-8, reading the credential from "
        "ANTHROPIC_API_KEY.\n\n"
        "## Auxiliary Model\n\nCompress tool results with anthropic/claude-haiku-4-5, reading the "
        "credential from ANTHROPIC_API_KEY.\n"
    )
    assert strategy.provider == "anthropic" and strategy.model == "anthropic/claude-opus-4-8"
    assert strategy.auxiliary_provider == "anthropic"
    assert strategy.auxiliary_model == "anthropic/claude-haiku-4-5"
    # Both share the same env var name here, but they're read into distinct
    # fields, not one shared one -- proven by the models themselves differing.
    assert strategy.api_key_env_var == "ANTHROPIC_API_KEY"
    assert strategy.auxiliary_api_key_env_var == "ANTHROPIC_API_KEY"


def test_strategy_reads_a_declared_max_output_tokens_and_wires_it_into_the_binding():
    """The native tool loop can need a reply large enough to hold a whole
    script's source (see the truncation this caught in the wild: a 4000+
    character write cut off mid-identifier because the default 2048-token
    ceiling was never overridable from memory.md prose). 'Allow up to N
    tokens per reply' must reach the ModelBinding's request params, exactly
    like temperature does."""
    from ear import Strategy

    strategy = Strategy.from_markdown(
        "## Model Selection\n\nReason with anthropic/claude-opus-4-8, reading the credential from "
        "ANTHROPIC_API_KEY. Allow up to 8,000 tokens per reply.\n\n"
        "## Auxiliary Model\n\nCompress tool results with anthropic/claude-haiku-4-5, reading the "
        "credential from ANTHROPIC_API_KEY. max_tokens: 512.\n"
    )
    assert strategy.max_output_tokens == 8000
    assert strategy.auxiliary_max_output_tokens == 512

    binding = strategy.model_binding()
    assert binding.params["max_tokens"] == 8000
    aux_binding = strategy.auxiliary_model_binding()
    assert aux_binding.params["max_tokens"] == 512

    # A stack that never mentions a token ceiling gets none -- ear/llm.py's
    # own DEFAULT_MAX_TOKENS keeps applying, unchanged, exactly as before
    # this setting existed.
    plain = Strategy.from_markdown(
        "## Model Selection\n\nReason with anthropic/claude-opus-4-8, reading the credential from "
        "ANTHROPIC_API_KEY.\n"
    )
    assert plain.max_output_tokens is None
    assert "max_tokens" not in plain.model_binding().params


def test_no_auxiliary_model_still_applies_deterministic_compression():
    """No Auxiliary Model declared -- the default -- does not mean "no
    compression at all" any more: `ear.caveman`'s deterministic pass always
    runs, needs no credential, and cannot hallucinate (it can only delete
    matched filler words, never generate replacement text). It is the safe
    always-on default; the LLM layer below is the opt-in extra squeeze."""
    from ear.reasoner import Reasoner

    runtime = Runtime(name="NoAux")  # auxiliary_model_binding defaults to None
    fed_back = Reasoner._compress_tool_result(runtime, "calc", {"x": 5}, "The result is really just 6")
    assert "6" in fed_back
    assert "really" not in fed_back.lower() and "just" not in fed_back.lower()


def test_auxiliary_model_squeezes_only_a_still_large_result():
    from ear import ModelBinding
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    aux = ModelBinding(provider="anthropic", model="test-haiku")
    aux.lm = ScriptedLM(["## summary\n\nfile read ok, 907 rows\n"])
    runtime = Runtime(name="WithAux", auxiliary_model_binding=aux)

    # Only a result still large after the deterministic pass (over
    # _SUMMARIZE_ABOVE_CHARS) earns the Auxiliary Model call.
    long_result = "read uploads/data.xlsx: 907 rows, 17 columns. " + ("segment " * 5000)
    tool = BoundTool(name="read_file", description="read", handler=lambda path: long_result)
    invoke = runtime.tool_binder.logged_handler(runtime, tool)
    raw = invoke(path="uploads/data.xlsx")
    assert raw == long_result  # the handler's own return is never touched

    fed_back = Reasoner._compress_tool_result(runtime, "read_file", {"path": "uploads/data.xlsx"}, raw)
    assert fed_back == "file read ok, 907 rows"  # the compressed text, not the raw one

    # The full, uncompressed result is what actually landed on the trail --
    # compression only touched the copy handed back into `gathered`.
    tool_record = runtime.reasoning_log.for_stage("tool")[0]
    assert tool_record.output == long_result

    summarize_record = runtime.reasoning_log.for_stage("summarize")[0]
    assert summarize_record.output == "file read ok, 907 rows"
    assert summarize_record.model == "anthropic/test-haiku"


def test_auxiliary_model_failure_falls_back_to_the_deterministic_result():
    """When the (opt-in) LLM squeeze fails, the fallback is the
    deterministically-compressed text, not the fully raw one -- the
    deterministic pass already ran first and never fails, so there is
    nothing to lose by keeping its output rather than reverting further."""
    from ear import ModelBinding
    from ear.reasoner import Reasoner

    class BrokenLM:
        history = []

        def complete(self, prompt, system=""):
            raise RuntimeError("summarizer is down")

    aux = ModelBinding(provider="anthropic", model="test-haiku")
    aux.lm = BrokenLM()
    runtime = Runtime(name="AuxBroken", auxiliary_model_binding=aux)

    # A large result reaches the (now broken) Auxiliary Model; the fallback is
    # the deterministic text, and the cycle is never blocked.
    fed_back = Reasoner._compress_tool_result(runtime, "read_file", {"path": "x"}, "the real result " + ("segment " * 5000))
    assert "real result" in fed_back.lower()  # substance never lost, never blocked the cycle
    record = runtime.reasoning_log.for_stage("summarize")[0]
    assert "FAILED" in record.output and "summarizer is down" in record.output


def test_auxiliary_model_is_not_called_for_a_small_result():
    """The change that removes the per-tool-call summarize: a result the
    deterministic pass already handles never reaches the Auxiliary Model, so
    no summarize call is billed and no summarize record is written."""
    from ear import ModelBinding
    from ear.reasoner import Reasoner

    class CountingLM:
        history: list = []

        def __init__(self):
            self.calls = 0

        def complete(self, prompt, system=""):
            self.calls += 1
            return "## summary\n\nshould not have been called\n"

    lm = CountingLM()
    aux = ModelBinding(provider="anthropic", model="test-haiku")
    aux.lm = lm
    runtime = Runtime(name="SmallResult", auxiliary_model_binding=aux)

    fed_back = Reasoner._compress_tool_result(runtime, "calc", {"x": 5}, "The result is really just 6")
    assert "6" in fed_back
    assert "really" not in fed_back.lower() and "just" not in fed_back.lower()
    assert lm.calls == 0  # no model call for a small result
    assert len(runtime.reasoning_log.for_stage("summarize")) == 0


def test_tool_loop_end_to_end_with_auxiliary_model_compression():
    """The full native tool loop, with a real (scripted) Auxiliary Model in
    play: the *next* prompt the primary model sees must carry the
    compressed summary, not the tool's full raw output."""
    from ear import ModelBinding
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    raw_output = "wrote 4096 characters to workspace/generate_dashboard.py. " + ("segment " * 5000)
    tool = BoundTool(name="write_file", description="write", handler=lambda path, content: raw_output)

    aux = ModelBinding(provider="anthropic", model="test-haiku")
    aux.lm = ScriptedLM(["## summary\n\nfile wrote ok\n"])
    runtime = Runtime(name="EndToEndAux", auxiliary_model_binding=aux)

    primary = ScriptedLM([
        (
            "## tool\n\nwrite_file\n\n## arguments\n\n- path: workspace/generate_dashboard.py\n"
            "content:\n> print('x')\n\n## decision\n\n"
        ),
        _tool_action(decision="Dashboard generated."),
    ])
    decision = Reasoner._reason_with_tools(
        Intent(text="generate"), runtime, primary, context={}, capabilities="none", tools=[tool], max_iterations=6
    )
    assert decision == "Dashboard generated."

    second_prompt = primary.prompts[1]
    assert "file wrote ok" in second_prompt
    assert raw_output not in second_prompt  # the raw text never re-enters the prompt

    # But the raw text is exactly what an investigator finds on the trail.
    tool_record = runtime.reasoning_log.for_stage("tool")[0]
    assert tool_record.output == raw_output


# ---------------------------------------------------------------------------
# Context checkpoint: every _CONTEXT_CHECKPOINT_EVERY (3) tool calls, the
# native tool loop consolidates everything gathered so far into one verified
# statement -- via the Auxiliary Model -- so key facts stay retained rather
# than diluting across a lengthening list of independently-compressed
# entries. No Auxiliary Model declared is a no-op, same as compression.
# ---------------------------------------------------------------------------


def test_no_auxiliary_model_checkpoint_is_a_noop():
    from ear.reasoner import Reasoner

    runtime = Runtime(name="NoAuxCheckpoint")  # auxiliary_model_binding defaults to None
    gathered = ["tool_a({}) -> did a thing", "tool_b({}) -> did another"]
    assert Reasoner._checkpoint_gathered_context(runtime, gathered) == "\n".join(gathered)


def test_checkpoint_consolidates_after_every_third_tool_call():
    """The 4th tool call's prompt must see the checkpoint's consolidated
    text, not the three raw gathered entries it replaced."""
    from ear import ModelBinding
    from ear.reasoner import Reasoner
    from ear.tool_binder import BoundTool

    calls = {"count": 0}

    def probe(path):
        calls["count"] += 1
        return f"probed {path}, attempt {calls['count']}"

    tool = BoundTool(name="probe", description="probe a path", handler=probe)

    # The per-call probe results are small, so none reaches the Auxiliary
    # Model now -- only the every-3-calls checkpoint does. The script supplies
    # just that one consolidation.
    aux = ModelBinding(provider="anthropic", model="test-sonnet")
    aux.lm = ScriptedLM([
        "## checkpoint\n\nChecked a, b, c so far -- all probed successfully, 3 attempts total.\n",
    ])
    runtime = Runtime(name="CheckpointFlow", auxiliary_model_binding=aux)

    primary = ScriptedLM([
        _tool_action(tool="probe", args="- path: a"),
        _tool_action(tool="probe", args="- path: b"),
        _tool_action(tool="probe", args="- path: c"),
        _tool_action(tool="probe", args="- path: d"),
        _tool_action(decision="Done."),
    ])
    decision = Reasoner._reason_with_tools(
        Intent(text="probe things"), runtime, primary, context={}, capabilities="none",
        tools=[tool], max_iterations=8,
    )
    assert decision == "Done."

    # The prompt for the 4th tool call (index 3) must carry the checkpoint,
    # not the three individual entries it consolidated.
    fourth_call_prompt = primary.prompts[3]
    assert "Checked a, b, c so far" in fourth_call_prompt
    assert "probed a, attempt 1" not in fourth_call_prompt

    checkpoint_record = runtime.reasoning_log.for_stage("checkpoint")[0]
    assert checkpoint_record.inputs["entries"] == 3
    assert "Checked a, b, c so far" in checkpoint_record.output
    assert checkpoint_record.model == "anthropic/test-sonnet"


def test_checkpoint_failure_falls_back_to_joined_entries():
    from ear import ModelBinding
    from ear.reasoner import Reasoner

    class BrokenLM:
        history = []

        def complete(self, prompt, system=""):
            raise RuntimeError("checkpoint model is down")

    aux = ModelBinding(provider="anthropic", model="test-sonnet")
    aux.lm = BrokenLM()
    runtime = Runtime(name="CheckpointBroken", auxiliary_model_binding=aux)

    gathered = ["tool_a({}) -> fact one", "tool_b({}) -> fact two"]
    result = Reasoner._checkpoint_gathered_context(runtime, gathered)
    assert "fact one" in result and "fact two" in result  # nothing lost on failure
    record = runtime.reasoning_log.for_stage("checkpoint")[0]
    assert "FAILED" in record.output and "checkpoint model is down" in record.output


def test_a_failure_is_fed_back_verbatim_up_to_the_raised_bound():
    """A failure is never summarized -- the next turn needs its exact error
    text. The verbatim bound is now the same size a success would be
    summarized at, so a mid-size error that the old 8000-char bound would have
    truncated is fed back whole."""
    from ear.reasoner import Reasoner

    err = "SyntaxError: invalid syntax (line 4)\n" + ("context line\n" * 800)  # ~10K: over old 8000, under 30000
    assert 8000 < len(err) < 30000
    fed = Reasoner._failure_feedback(err)
    assert fed == f"FAILED: {err.strip()}"   # whole error (trailing whitespace stripped), no truncation
    assert "truncated" not in fed


def test_a_giant_failure_is_truncated_head_and_tail_not_summarized():
    """A failure larger than the bound is truncated deterministically, keeping
    the head (first error line) and tail (final error) where the signal lives,
    dropping the middle -- no model ever paraphrases it."""
    from ear.reasoner import Reasoner

    huge = "SyntaxError at line 4\n" + ("noise line\n" * 4000) + "FinalError: boom"
    assert len(huge) > 30000
    fed = Reasoner._failure_feedback(huge)
    assert fed.startswith("FAILED:")
    assert "SyntaxError at line 4" in fed   # head kept
    assert "FinalError: boom" in fed        # tail kept
    assert "truncated" in fed               # middle dropped
    assert len(fed) < len(huge)


def test_prune_recovered_failures_collapses_failed_entries_to_markers():
    """Once a call succeeds, earlier failures' verbose bodies collapse to a
    one-line marker (fact + short note kept, stack trace dropped); non-failure
    entries pass through untouched."""
    from ear.reasoner import Reasoner

    gathered = [
        "read_file({'path': 'x'}) -> file contents ok",
        "run_cmd({'cmd': 'python bad.py'}) -> FAILED: SyntaxError line 4 -- " + ("noise " * 500),
        "run_cmd({'cmd': 'python good.py'}) -> exit 0, ran fine",
    ]
    pruned = Reasoner._prune_recovered_failures(gathered)
    assert pruned[0] == gathered[0]                    # success untouched
    assert pruned[2] == gathered[2]                    # success untouched
    assert "earlier failure, recovered" in pruned[1]   # failure collapsed
    assert "SyntaxError line 4" in pruned[1]           # the short note survives
    assert "full text on the trail" in pruned[1]
    assert len(pruned[1]) < len(gathered[1])           # verbose body dropped


# ---------------------------------------------------------------------------
# Provider-agnostic prompt caching: a signature declares its one volatile
# input as the cache_boundary; that input renders last so everything before it
# is a stable, byte-identical prefix across the tool loop's iterations. The LM
# seam marks that prefix (Anthropic cache_control) or ignores it (OpenAI-
# compatible auto-caches) and captures cache-read/write tokens either way.
# ---------------------------------------------------------------------------


def test_render_without_a_boundary_is_unchanged_and_emits_no_prefix():
    from ear.judgment import Field, Judgment

    j = Judgment(instruction="do", inputs=[Field("a", "first"), Field("b", "second")])
    prompt, prefix = j._render({"a": "x", "b": "y"})
    assert prefix == ""                                  # nothing to cache
    assert prompt == j.render_prompt({"a": "x", "b": "y"})  # text identical to before


def test_choose_tool_action_emits_a_stable_cache_prefix_before_gathered():
    from ear.signatures import ChooseToolAction

    stable = {"intent": "do it", "context": "ctx", "capabilities": "caps", "tools": "read(x): reads"}
    prompt1, prefix1 = ChooseToolAction._render({**stable, "gathered": "none yet"})
    prompt2, prefix2 = ChooseToolAction._render({**stable, "gathered": "read(x) -> a lot more gathered now"})

    assert prefix1 and prompt1.startswith(prefix1)   # a real byte-prefix of the prompt
    assert prompt2.startswith(prefix1)               # the growing call still starts with it
    assert prefix1 == prefix2                         # byte-stable across a changing `gathered`
    assert "none yet" not in prefix1                  # the volatile value sits past the boundary
    assert "gathered" in prefix1                      # the boundary heading is in the cached span


def test_anthropic_wire_caches_the_declared_prefix_and_parses_cache_tokens():
    from ear.llm import LM

    lm = LM(model="anthropic/claude-sonnet-5", api_key="k")

    _, _, body, parse = lm._anthropic("STABLE-part|volatile-part", "sys", cache_prefix="STABLE-part|")
    content = body["messages"][0]["content"]
    assert isinstance(content, list) and len(content) == 2
    assert content[0]["cache_control"] == {"type": "ephemeral"}
    assert content[0]["text"] == "STABLE-part|" and content[1]["text"] == "volatile-part"

    # No boundary -> the content stays a plain string, byte-identical to the
    # uncached request (so caching is inert until a caller opts in).
    _, _, body2, _ = lm._anthropic("STABLE-part|volatile-part", "sys")
    assert body2["messages"][0]["content"] == "STABLE-part|volatile-part"

    _, usage = parse({
        "content": [{"type": "text", "text": "ok"}],
        "usage": {"input_tokens": 10, "output_tokens": 2,
                  "cache_read_input_tokens": 7, "cache_creation_input_tokens": 3},
    })
    assert usage["cache_read_tokens"] == 7 and usage["cache_write_tokens"] == 3


# ---------------------------------------------------------------------------
# The Monitor: a premium live TUI of the whole fleet as a factory assembly
# line, rendered from the trail with zero dependencies.
# ---------------------------------------------------------------------------


def _strip_ansi(text):
    import re

    return re.sub(r"\x1b\[[0-9;?]*[a-zA-Z]", "", text)


def test_monitor_renders_an_assembly_line_frame_from_the_fleet(tmp_path):
    from datetime import datetime, timezone

    from ear import Monitor

    a = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    for i in range(4):
        a.reason(Intent(text=f"c{i}", context={"loan_amount": 5000 + i}))
    b = load_runtime(write_stack(tmp_path / "b", **MINIMAL_STACK))
    b.reason(Intent(text="one", context={"loan_amount": 5000}))

    frame = Monitor().render_frame(
        {"consumer-lending": a, "smb-lending": b}, width=100, frame=2, now=datetime.now(timezone.utc)
    )
    plain = _strip_ansi(frame)

    # It's a truecolor terminal frame...
    assert "\x1b[38;2;" in frame
    # ...laid out as the assembly line, with both instances as lanes...
    assert "ASSEMBLY LINE" in plain
    assert "consumer-lending" in plain and "smb-lending" in plain
    # ...the pipeline stations and KPI tiles...
    assert "GOV" in plain and "DLB" in plain and "LRN" in plain
    assert "INSTANCES" in plain and "HEALTHY" in plain and "CYCLES" in plain
    # ...and every rendered row padded to the width (no ragged edges).
    for line in frame.split("\n"):
        assert _visible_width(line) == 100


def test_monitor_frame_reflects_health_and_animates_between_frames(tmp_path):
    from datetime import datetime, timezone

    from ear import Monitor

    healthy = load_runtime(write_stack(tmp_path / "h", **MINIMAL_STACK))
    healthy.reason(Intent(text="ok", context={"loan_amount": 5000}))

    # A broken audit chain -> the fleet shows a broken instance.
    broken = load_runtime(write_stack(tmp_path / "k", **MINIMAL_STACK))
    broken.reasoning_log.path = str(tmp_path / "k.jsonl")
    broken.reason(Intent(text="ok", context={"loan_amount": 5000}))
    trail = tmp_path / "k.jsonl"
    trail.write_text(trail.read_text(encoding="utf-8").replace("ok", "hax", 1), encoding="utf-8")

    now = datetime(2026, 7, 3, 12, 0, 0, tzinfo=timezone.utc)
    monitor = Monitor()
    frame0 = monitor.render_frame({"good-one": healthy, "bad-one": broken}, width=100, frame=0, now=now)
    frame1 = monitor.render_frame({"good-one": healthy, "bad-one": broken}, width=100, frame=1, now=now)

    plain0 = _strip_ansi(frame0)
    assert "BROKEN" in plain0 and "✗" in plain0  # the broken instance is surfaced
    # The animation phase differs frame to frame (spinner/sweep/shimmer).
    assert frame0 != frame1


def test_monitor_handles_an_empty_fleet_without_crashing():
    from ear import Monitor

    frame = Monitor().render_frame({}, width=80, frame=0)
    assert "no runtime instances" in _strip_ansi(frame)


def _visible_width(text):
    import re

    return len(re.sub(r"\x1b\[[0-9;?]*[a-zA-Z]", "", text))


# ---------------------------------------------------------------------------
# The Kernel: EAR as an OS-style scheduler -- run work when there is work,
# sleep until an interrupt when there is not.
# ---------------------------------------------------------------------------


def test_kernel_is_idle_until_work_is_submitted(tmp_path):
    from ear import Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", runtime)

    # No work -> the loop reports idle, never blocks in tick().
    assert kernel.tick() is None
    assert kernel.idle_waits == 1 and kernel.pending == 0

    kernel.submit("lending", Intent(text="Underwrite", context={"loan_amount": 5000}))
    assert kernel.pending == 1
    dispatch = kernel.tick()
    assert dispatch is not None and dispatch.status == "ran" and dispatch.instance == "lending"
    assert kernel.tick() is None  # back to idle


def test_kernel_dispatches_to_the_named_instance_and_records_history(tmp_path):
    from ear import Kernel

    a = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    b = load_runtime(write_stack(tmp_path / "b", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", a)
    kernel.register("mortgage", b)

    kernel.submit("lending", Intent(text="one", context={"loan_amount": 5000}))
    kernel.submit("mortgage", Intent(text="two", context={"loan_amount": 6000}))
    kernel.submit("lending", Intent(text="three", context={"loan_amount": 7000}))
    ran = kernel.drain()

    assert [d.instance for d in ran] == ["lending", "mortgage", "lending"]
    assert all(d.status == "ran" for d in ran)
    # The work actually landed on each instance's own trail.
    assert len({r.cycle for r in a.reasoning_log.records if r.stage == "intent"}) == 2
    assert len({r.cycle for r in b.reasoning_log.records if r.stage == "intent"}) == 1
    assert len(kernel.history) == 3


def test_kernel_parks_a_governance_stop_and_keeps_running(tmp_path):
    from ear import Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", runtime)

    kernel.submit("lending", Intent(text="oversized", context={"loan_amount": 90000}))  # blocked by policy
    kernel.submit("lending", Intent(text="fine", context={"loan_amount": 5000}))  # should still run
    ran = kernel.drain()

    assert ran[0].status == "blocked" and "Loan Amount Cap" in ran[0].summary
    assert ran[1].status == "ran"  # one task's stop never takes the kernel down


def test_kernel_fails_a_task_for_an_unknown_instance_without_crashing():
    from ear import Kernel

    kernel = Kernel()
    kernel.submit("ghost", Intent(text="x"))
    dispatch = kernel.drain()[0]
    assert dispatch.status == "failed" and "no such instance" in dispatch.summary


def test_kernel_runs_a_goal_task_through_pursue(tmp_path):
    from ear import Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", runtime)

    kernel.submit("lending", Intent(text="Underwrite", context={"loan_amount": 5000}), goal="reach a decision")
    dispatch = kernel.drain()[0]
    # Offline the goal is ungraded, but it ran through pursue and is recorded.
    assert dispatch.status == "ran" and "goal ungraded" in dispatch.summary
    assert runtime.reasoning_log.for_stage("goal")


def test_kernel_recurring_task_reschedules_forward_and_one_shots_clear(tmp_path):
    import time as _time

    from ear import Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", runtime)

    one_shot = kernel.submit("lending", Intent(text="once", context={"loan_amount": 5000}))
    recurring = kernel.submit(
        "lending", Intent(text="recur", context={"loan_amount": 6000}), every=100, delay=0
    )
    kernel.drain()

    # The one-shot is gone; the recurring task remains, its next firing in
    # the future, having run once.
    assert one_shot not in kernel.queue
    assert recurring in kernel.queue
    assert recurring.runs == 1 and recurring.due > _time.monotonic()
    assert kernel.tick() is None  # not due again yet -> idle


def test_kernel_cancel_removes_a_queued_task(tmp_path):
    from ear import Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", runtime)
    task = kernel.submit("lending", Intent(text="x", context={"loan_amount": 5000}), delay=100)
    assert kernel.pending == 1
    assert kernel.cancel(task) is True and kernel.pending == 0


def test_kernel_blocks_a_task_carrying_a_claim_unauthorized_for_the_instance(tmp_path):
    from ear import Claim, Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    assert runtime.tenant.org_id == "default"
    kernel = Kernel()
    kernel.register("lending", runtime)

    claim = Claim(subject="bob", org_ids=("org_other",))
    kernel.submit("lending", Intent(text="x", context={"loan_amount": 5000}), claim=claim)
    dispatch = kernel.drain()[0]
    assert dispatch.status == "blocked" and "bob" in dispatch.summary


def test_kernel_runs_a_task_carrying_a_claim_authorized_for_the_instance(tmp_path):
    from ear import Claim, Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", runtime)

    claim = Claim(subject="alice", org_ids=("default",))
    kernel.submit("lending", Intent(text="x", context={"loan_amount": 5000}), claim=claim)
    dispatch = kernel.drain()[0]
    assert dispatch.status == "ran"


def test_kernel_background_loop_wakes_on_submit_and_stops(tmp_path):
    import time as _time

    from ear import Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", runtime)
    kernel.start()
    try:
        kernel.submit("lending", Intent(text="bg", context={"loan_amount": 5000}))
        deadline = _time.time() + 3
        while _time.time() < deadline and not kernel.history:
            _time.sleep(0.02)
        assert kernel.history and kernel.history[-1].status == "ran"
    finally:
        kernel.stop()
    assert kernel.running is False


def test_monitor_shows_a_station_legend_tooltip(tmp_path):
    from datetime import datetime, timezone

    from ear import Monitor

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    runtime.reason(Intent(text="c", context={"loan_amount": 5000}))
    frame = Monitor().render_frame({"lending": runtime}, width=100, frame=0, now=datetime(2026, 7, 3, tzinfo=timezone.utc))
    plain = _strip_ansi(frame)

    # The on-screen key expands every abbreviated station code.
    assert "STATIONS" in plain
    for code, name in [("GOV", "govern"), ("DLB", "deliberate"), ("AUD", "audit"), ("LRN", "learn")]:
        assert code in plain and name in plain
    # Wrapped to fit -- the width invariant still holds on every row.
    for line in frame.split("\n"):
        assert _visible_width(line) == 100


# ---------------------------------------------------------------------------
# The Server: EAR as a control-plane HTTP service over the Kernel. The whole
# API is a pure function, so it's tested without opening a socket -- plus one
# real-socket check for the auth layer.
# ---------------------------------------------------------------------------


def test_server_health_and_instance_lifecycle(tmp_path):
    from ear import Server

    stacks = tmp_path / "stacks"
    stacks.mkdir()
    write_stack(stacks / "lending", **MINIMAL_STACK)
    server = Server(stacks_root=stacks)

    status, health = server.handle("GET", "/health")
    assert status == 200 and health["status"] == "ok" and health["instances"] == 0

    # Create from a stack under the root.
    status, created = server.handle("POST", "/instances", {"name": "lending", "stack": "lending"})
    assert status == 201 and created["from_stack"] is True
    assert server.handle("GET", "/instances")[1]["instances"][0]["instance"] == "lending"

    # Duplicates, missing fields and path escapes are client errors, not crashes.
    assert server.handle("POST", "/instances", {"name": "lending"})[0] == 409
    assert server.handle("POST", "/instances", {})[0] == 400
    assert server.handle("POST", "/instances", {"name": "x", "stack": "../../etc"})[0] == 400
    # A bare instance needs no stack.
    assert server.handle("POST", "/instances", {"name": "bare"})[0] == 201

    assert server.handle("DELETE", "/instances/bare")[0] == 200
    assert server.handle("GET", "/instances/bare/status")[0] == 404


def test_server_submits_work_and_reports_status(tmp_path):
    from ear import Server

    stacks = tmp_path / "stacks"
    stacks.mkdir()
    write_stack(stacks / "lending", **MINIMAL_STACK)
    server = Server(stacks_root=stacks)
    server.handle("POST", "/instances", {"name": "lending", "stack": "lending"})

    status, accepted = server.handle(
        "POST", "/instances/lending/submit", {"intent": "Underwrite", "context": {"loan_amount": 5000}}
    )
    assert status == 202 and accepted["task_id"] and accepted["recurring"] is False
    assert server.handle("POST", "/instances/ghost/submit", {"intent": "x"})[0] == 404
    assert server.handle("POST", "/instances/lending/submit", {})[0] == 400  # intent required

    server.kernel.drain()  # run the queued work

    reported = server.handle("GET", "/instances/lending/status")[1]
    assert reported["cycles"] == 1 and reported["status"] == "healthy"
    assert "resolved intent" in server.handle("GET", "/instances/lending/decision")[1]["decision"]
    assert len(server.handle("GET", "/instances/lending/trail", {"limit": 5})[1]["records"]) == 5
    assert server.handle("GET", "/kernel")[1]["dispatched"] == 1


def test_server_submit_credentials_construct_a_binding_the_stack_never_declared(tmp_path):
    """A stack with no ## Model section leaves runtime.model_binding as None
    (loader.py's single-tenant-safe default) -- fine when the caller can
    fall back to os.environ, not when its key arrives per-submission from a
    multi-tenant server instead. Naming provider/model explicitly also
    sidesteps memory.md's prose-guessing for providers EAR's own heuristics
    don't recognise (e.g. "openrouter")."""
    from ear import Server

    stacks = tmp_path / "stacks"
    stacks.mkdir()
    write_stack(stacks / "lending", **MINIMAL_STACK)
    server = Server(stacks_root=stacks)
    server.handle("POST", "/instances", {"name": "lending", "stack": "lending"})

    runtime = server.kernel.instances["lending"]
    assert runtime.model_binding is None

    server.handle(
        "POST",
        "/instances/lending/submit",
        {
            "intent": "Underwrite",
            "context": {"loan_amount": 5000},
            "credentials": {"provider": "openrouter", "model": "anthropic/claude-sonnet-5", "api_key": "sk-tenant-1"},
        },
    )

    assert runtime.model_binding is not None
    assert runtime.model_binding.provider == "openrouter"
    assert runtime.model_binding.model == "anthropic/claude-sonnet-5"
    assert runtime.model_binding.resolve_api_key() == "sk-tenant-1"


def test_server_creates_an_instance_from_inline_files(tmp_path):
    """A caller in a different process (e.g. a LENS server generating a
    persona's stack on the fly) has no filesystem to share with this one --
    `files` lets it hand over file contents directly instead of requiring a
    pre-populated stacks_root directory."""
    from ear import Server

    stacks = tmp_path / "stacks"
    stacks.mkdir()
    server = Server(stacks_root=stacks)

    status, created = server.handle(
        "POST",
        "/instances",
        {
            "name": "inline-lending",
            "files": {
                "skills.md": MINIMAL_STACK["skills"],
                "persona.md": MINIMAL_STACK["persona"],
                "workflow.md": MINIMAL_STACK["workflow"],
                "process.md": MINIMAL_STACK["process"],
                "policy.md": MINIMAL_STACK["policy"],
            },
        },
    )
    assert status == 201 and created["from_stack"] is True
    assert server.handle("GET", "/instances")[1]["instances"][0]["instance"] == "inline-lending"

    server.handle(
        "POST", "/instances/inline-lending/submit", {"intent": "Underwrite", "context": {"loan_amount": 5000}}
    )
    server.kernel.drain()
    assert server.kernel.history[-1].status == "ran"

    # An unknown filename or non-string content is a client error, not a crash.
    assert server.handle("POST", "/instances", {"name": "x", "files": {"evil.md": "hi"}})[0] == 400
    assert server.handle("POST", "/instances", {"name": "y", "files": {"skills.md": 123}})[0] == 400
    # An empty files object is indistinguishable from "no files" -- a bare instance, not an error.
    assert server.handle("POST", "/instances", {"name": "z", "files": {}})[0] == 201


def test_server_trail_reads_limit_from_the_query_string(tmp_path):
    """GET requests carry no body over most HTTP clients (Node's fetch
    refuses one outright), so a caller must be able to pass `limit` as
    `?limit=N` -- not only as a JSON body, which only a from-scratch client
    like EAR's own test harness would ever send on a GET."""
    from ear import Server

    stacks = tmp_path / "stacks"
    stacks.mkdir()
    write_stack(stacks / "lending", **MINIMAL_STACK)
    server = Server(stacks_root=stacks)
    server.handle("POST", "/instances", {"name": "lending", "stack": "lending"})
    server.handle("POST", "/instances/lending/submit", {"intent": "Underwrite", "context": {"loan_amount": 5000}})
    server.kernel.drain()

    status, trail = server.handle("GET", "/instances/lending/trail?limit=2")
    assert status == 200 and len(trail["records"]) == 2


def test_server_approve_resubmits_a_parked_intent_over_the_wire(tmp_path):
    """`Exchange`'s approval.md file-drop assumes a shared filesystem --
    not true when the caller is a separate process over the network. The
    server's `/approve` endpoint is the same release without one: it
    remembers the last submitted intent and resubmits it with a verdict."""
    from ear import Server

    stack = dict(
        MINIMAL_STACK,
        policy=(
            "# Policies\n\n## Loan Amount Cap\nThe loan must not exceed $50,000.\n\n"
            "Fallback: loan_amount <= 50000\nApproval: required\nApplies to: runtime\n"
        ),
    )
    stacks = tmp_path / "stacks"
    stacks.mkdir()
    write_stack(stacks / "lending", **stack)
    server = Server(stacks_root=stacks)
    server.handle("POST", "/instances", {"name": "lending", "stack": "lending"})

    # No pending intent yet -- approving before ever submitting is a client error.
    assert server.handle("POST", "/instances/lending/approve", {"verdict": "approved"})[0] == 409

    server.handle(
        "POST", "/instances/lending/submit", {"intent": "Underwrite", "context": {"loan_amount": 60000}}
    )
    server.kernel.drain()
    assert server.kernel.history[-1].status == "blocked"
    trail = server.handle("GET", "/instances/lending/trail", {"limit": 20})[1]["records"]
    assert any(r["stage"] == "approval" and "PENDING" in r["output"] for r in trail)

    status, approved = server.handle(
        "POST", "/instances/lending/approve",
        {"verdict": "approved", "approver": "senior@bank.com", "note": "exception reviewed"},
    )
    assert status == 202 and approved["verdict"] is True

    server.kernel.drain()
    assert server.kernel.history[-1].status == "ran"


def test_server_unknown_route_is_a_404():
    from ear import Server

    server = Server()
    assert server.handle("GET", "/nope")[0] == 404
    assert server.handle("PUT", "/instances")[0] == 404


def test_server_bearer_token_auth_over_a_real_socket(tmp_path):
    import urllib.error
    import urllib.request

    from ear import Server

    server = Server(host="127.0.0.1", port=0, token="s3cret")
    server.start()
    try:
        host, port = server.address

        def get(path, token=None):
            request = urllib.request.Request(f"http://{host}:{port}{path}")
            if token:
                request.add_header("Authorization", f"Bearer {token}")
            try:
                with urllib.request.urlopen(request, timeout=3) as response:
                    return response.status
            except urllib.error.HTTPError as error:
                return error.code

        assert get("/health") == 401  # no token
        assert get("/health", "wrong") == 401  # bad token
        assert get("/health", "s3cret") == 200  # right token
    finally:
        server.stop()


def test_server_bridges_a_declared_tool_to_a_remote_http_system(tmp_path):
    """Declaring `- check_inventory: ...` in memory.md's Tools section gives
    the model something to see, not something to run (tool.py) -- nothing
    auto-executes a `command` string. When the caller creating the instance
    is itself a remote system, bridge_url/bridge_token wire every declared
    but otherwise-unbound tool to a generic HTTP forward instead."""
    import json as _json
    import threading as _threading
    from http.server import BaseHTTPRequestHandler, HTTPServer

    from ear import Server

    received = []

    class _BridgeHandler(BaseHTTPRequestHandler):
        def do_POST(self):
            length = int(self.headers.get("Content-Length") or 0)
            payload = _json.loads(self.rfile.read(length))
            received.append((self.headers.get("Authorization"), payload))
            body = _json.dumps({"ok": True, "output": {"level": 42}}).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def log_message(self, *args):
            return

    bridge = HTTPServer(("127.0.0.1", 0), _BridgeHandler)
    thread = _threading.Thread(target=bridge.serve_forever, daemon=True)
    thread.start()
    try:
        bridge_url = f"http://127.0.0.1:{bridge.server_address[1]}/execute"
        stack = dict(
            MINIMAL_STACK,
            memory="# Memory\n\n## Tools\n\n- check_inventory: looks up stock for a SKU\n",
        )
        stacks = tmp_path / "stacks"
        stacks.mkdir()
        write_stack(stacks / "lending", **stack)
        server = Server(stacks_root=stacks, bridge_url=bridge_url, bridge_token="bridge-secret")
        server.handle(
            "POST",
            "/instances",
            {"name": "lending", "stack": "lending", "bridge_context": {"orgId": "org-1", "taskId": "task-1"}},
        )

        runtime = server.kernel.instances["lending"]
        tools = runtime.tool_binder.bound_tools(runtime)
        tool = next(t for t in tools if t.name == "check_inventory")
        result = tool.handler(sku="ABC-1")

        assert result == '{"level": 42}'
        auth_header, payload = received[0]
        assert auth_header == "Bearer bridge-secret"
        assert payload == {
            "tool": "check_inventory",
            "input": {"sku": "ABC-1"},
            "context": {"orgId": "org-1", "taskId": "task-1"},
        }
    finally:
        bridge.shutdown()
        thread.join(timeout=2)


# ---------------------------------------------------------------------------
# Kubernetes: EAR instances as Jobs and CronJobs, spoken natively over the
# K8s REST API. Unit-tested against a faithful fake transport -- not a live
# cluster (none is available from here); the tests hold it to the API shape.
# ---------------------------------------------------------------------------


class _FakeKube:
    """Records calls and returns canned Kubernetes API responses."""

    def __init__(self, status=201):
        self.status = status
        self.calls = []

    def __call__(self, method, url, headers, body):
        self.calls.append((method, url, body))
        name = (body or {}).get("metadata", {}).get("name", "obj")
        if self.status >= 400:
            return self.status, {"message": "AlreadyExists"}
        return self.status, {"metadata": {"name": name}, "status": {}}


def test_run_entrypoint_runs_one_cycle_from_the_environment(tmp_path, monkeypatch):
    import ear.run as run

    stack = write_stack(tmp_path / "stack", **MINIMAL_STACK)
    monkeypatch.setenv("EAR_INTENT", "Underwrite a loan")
    monkeypatch.setenv("EAR_CONTEXT", '{"loan_amount": 5000}')
    monkeypatch.setenv("EAR_DECISION_PATH", str(tmp_path / "decision.md"))
    assert run.main([str(stack)]) == 0
    assert (tmp_path / "decision.md").exists()

    # A governance block exits 2 (a refusal is a valid outcome, not a crash).
    monkeypatch.setenv("EAR_CONTEXT", '{"loan_amount": 90000}')
    assert run.main([str(stack)]) == 2
    # A missing intent exits 2.
    monkeypatch.delenv("EAR_INTENT")
    assert run.main([str(stack)]) == 2


def test_k8s_manifests_have_the_right_shape():
    from ear.k8s import cronjob_manifest, container_spec, job_manifest

    job = job_manifest(
        "Consumer Lending",
        "ear:1.0",
        command=["python", "-m", "ear.run", "/stack"],
        env={"EAR_INTENT": "x"},
        namespace="tenants",
        cpu="500m",
        memory="512Mi",
    )
    assert job["kind"] == "Job" and job["apiVersion"] == "batch/v1"
    assert job["metadata"]["name"] == "consumer-lending"  # RFC-1123 safe
    assert job["metadata"]["namespace"] == "tenants"
    pod = job["spec"]["template"]["spec"]
    assert pod["restartPolicy"] == "Never"
    assert pod["containers"][0]["command"] == ["python", "-m", "ear.run", "/stack"]
    assert pod["containers"][0]["resources"]["limits"] == {"cpu": "500m", "memory": "512Mi"}

    cron = cronjob_manifest("nightly", "ear:1.0", "0 0 * * *", namespace="tenants")
    assert cron["kind"] == "CronJob" and cron["spec"]["schedule"] == "0 0 * * *"
    assert cron["spec"]["jobTemplate"]["spec"]["template"]["spec"]["restartPolicy"] == "Never"


def test_every_to_cron_maps_periods_and_refuses_sub_minute():
    from ear.k8s import KubeError, every_to_cron

    assert every_to_cron(60) == "*/1 * * * *"
    assert every_to_cron(300) == "*/5 * * * *"
    assert every_to_cron(3600) == "0 */1 * * *"
    assert every_to_cron(86400) == "0 0 * * *"
    with pytest.raises(KubeError, match="sub-minute"):
        every_to_cron(30)


def test_kube_config_never_shows_the_token_in_repr():
    from ear import KubeConfig

    config = KubeConfig(api_server="https://k8s:443", token="super-secret-token")
    assert "super-secret-token" not in repr(config)
    assert config.token == "super-secret-token"  # suppressed from repr, not from use


def test_kube_client_lists_jobs_with_an_encoded_label_selector():
    from ear import KubeClient, KubeConfig

    fake = _FakeKube()
    client = KubeClient(KubeConfig(api_server="https://k8s:443", token="t", namespace="tenants"), transport=fake)

    client.list_jobs(label_selector="app=lending,tier=prod")
    method, url, _body = fake.calls[-1]
    assert method == "GET"
    assert url.endswith("/apis/batch/v1/namespaces/tenants/jobs?labelSelector=app%3Dlending%2Ctier%3Dprod")

    client.list_jobs()  # no selector -> no query string at all
    _method, bare_url, _body = fake.calls[-1]
    assert bare_url.endswith("/jobs")


def test_kube_client_and_provider_speak_the_api(tmp_path):
    from ear import Intent, KubeClient, KubeConfig, KubeProvider

    fake = _FakeKube()
    client = KubeClient(KubeConfig(api_server="https://k8s:443", token="t", namespace="tenants"), transport=fake)
    provider = KubeProvider(client=client, image="ear:1.0", cpu="500m", memory="256Mi")

    created = provider.run("lending", Intent(text="Underwrite", context={"loan_amount": 5000}))
    method, url, body = fake.calls[-1]
    assert method == "POST" and url.endswith("/apis/batch/v1/namespaces/tenants/jobs")
    assert created["metadata"]["name"].startswith("lending-")
    env = {e["name"]: e["value"] for e in body["spec"]["template"]["spec"]["containers"][0]["env"]}
    assert env["EAR_INTENT"] == "Underwrite" and json.loads(env["EAR_CONTEXT"]) == {"loan_amount": 5000}

    provider.schedule("lending", Intent(text="daily"), every=86400)
    _, cron_url, cron_body = fake.calls[-1]
    assert cron_url.endswith("/cronjobs") and cron_body["spec"]["schedule"] == "0 0 * * *"


def test_job_manifest_omits_volumes_when_none_are_given():
    """The pre-existing shape is untouched when no shared volume is
    configured -- additive, never a behavior change for a caller who
    never asked for one."""
    from ear.k8s import job_manifest

    job = job_manifest("lending", "ear:1.0")
    pod = job["spec"]["template"]["spec"]
    assert "volumes" not in pod
    assert "volumeMounts" not in pod["containers"][0]


def test_host_path_and_pvc_volume_manifests_have_the_right_shape():
    from ear.k8s import host_path_volume, job_manifest, pvc_volume, volume_mount

    hostpath = host_path_volume("stack", "/srv/ear/lending")
    assert hostpath == {"name": "stack", "hostPath": {"path": "/srv/ear/lending", "type": "DirectoryOrCreate"}}

    claim = pvc_volume("stack", "ear-stacks-pvc")
    assert claim == {"name": "stack", "persistentVolumeClaim": {"claimName": "ear-stacks-pvc"}}

    mount = volume_mount("stack", "/stack")
    assert mount == {"name": "stack", "mountPath": "/stack"}

    job = job_manifest("lending", "ear:1.0", volumes=[hostpath], volume_mounts=[mount])
    pod = job["spec"]["template"]["spec"]
    assert pod["volumes"] == [hostpath]
    assert pod["containers"][0]["volumeMounts"] == [mount]


def test_kube_provider_mounts_a_host_path_at_stack_mount_for_every_job_and_cronjob():
    """The swift host<->container artifact path this session asked for:
    configure `host_path` once on the provider, every Job and CronJob it
    creates mounts it at `stack_mount` automatically -- the sandbox's
    uploads/outputs inside that mount are then plain host files, not
    something fetched through the Kubernetes API after the fact."""
    from ear import Intent, KubeClient, KubeConfig, KubeProvider

    fake = _FakeKube()
    client = KubeClient(KubeConfig(api_server="https://k8s:443", token="t", namespace="tenants"), transport=fake)
    provider = KubeProvider(client=client, image="ear:1.0", host_path="/srv/ear/lending")

    provider.run("lending", Intent(text="Underwrite"))
    _, _, job_body = fake.calls[-1]
    pod = job_body["spec"]["template"]["spec"]
    assert pod["volumes"] == [{"name": "stack", "hostPath": {"path": "/srv/ear/lending", "type": "DirectoryOrCreate"}}]
    assert pod["containers"][0]["volumeMounts"] == [{"name": "stack", "mountPath": "/stack"}]

    provider.schedule("lending", Intent(text="daily"), every=86400)
    _, _, cron_body = fake.calls[-1]
    cron_pod = cron_body["spec"]["jobTemplate"]["spec"]["template"]["spec"]
    assert cron_pod["volumes"] == [{"name": "stack", "hostPath": {"path": "/srv/ear/lending", "type": "DirectoryOrCreate"}}]


def test_kube_provider_prefers_pvc_over_host_path_when_both_are_set():
    from ear import Intent, KubeClient, KubeConfig, KubeProvider

    fake = _FakeKube()
    client = KubeClient(KubeConfig(api_server="https://k8s:443", token="t"), transport=fake)
    provider = KubeProvider(client=client, image="ear:1.0", host_path="/srv/ear/lending", pvc_claim="ear-stacks-pvc")

    provider.run("lending", Intent(text="Underwrite"))
    _, _, job_body = fake.calls[-1]
    pod = job_body["spec"]["template"]["spec"]
    assert pod["volumes"] == [{"name": "stack", "persistentVolumeClaim": {"claimName": "ear-stacks-pvc"}}]


def test_kube_client_raises_on_api_errors():
    from ear import KubeClient, KubeConfig, KubeError

    client = KubeClient(KubeConfig(api_server="https://k8s", token="t"), transport=_FakeKube(status=409))
    with pytest.raises(KubeError, match="409"):
        client.create_job({"metadata": {"name": "dup"}})


def test_kube_config_reads_in_cluster_service_account(tmp_path, monkeypatch):
    from ear import KubeConfig

    base = tmp_path / "sa"
    base.mkdir()
    (base / "token").write_text("tok-123", encoding="utf-8")
    (base / "namespace").write_text("prod", encoding="utf-8")
    (base / "ca.crt").write_text("CA", encoding="utf-8")
    monkeypatch.setenv("KUBERNETES_SERVICE_HOST", "10.0.0.1")
    monkeypatch.setenv("KUBERNETES_SERVICE_PORT", "443")

    config = KubeConfig.in_cluster(base=str(base))
    assert config.api_server == "https://10.0.0.1:443"
    assert config.token == "tok-123" and config.namespace == "prod" and config.ca_cert

    monkeypatch.delenv("KUBERNETES_SERVICE_HOST")
    from ear import KubeError

    with pytest.raises(KubeError, match="not running in a cluster"):
        KubeConfig.in_cluster(base=str(base))


def test_kernel_dispatcher_seam_runs_work_on_the_provider(tmp_path):
    from ear import Kernel

    runtime = load_runtime(write_stack(tmp_path / "a", **MINIMAL_STACK))
    kernel = Kernel()
    kernel.register("lending", runtime)

    seen = []

    def dispatcher(task, rt):
        seen.append((task.instance, rt))
        return "dispatched", f"job for {task.instance} created"

    kernel.dispatcher = dispatcher
    kernel.submit("lending", Intent(text="go", context={"loan_amount": 5000}))
    dispatch = kernel.drain()[0]

    # The work went to the dispatcher (the cluster), not the in-process cycle.
    assert dispatch.status == "dispatched" and "job for lending" in dispatch.summary
    assert seen == [("lending", runtime)]
    assert not runtime.reasoning_log.records  # reason() was never called in-process
