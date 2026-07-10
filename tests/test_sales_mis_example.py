"""Offline preflight for the Sales MIS Guru example.

Every check here reproduces, with no API key and no model call, a failure
class that previously took a paid live run to discover (see
examples/sales_mis_stack/logs/): the sandbox environment silently unable to
run the model's openpyxl scripts, the runner feeding a cycle a false
premise, stale artifacts masquerading as progress, or the dashboard
narrating instead of proving. A live run should only ever be spent on the
one thing this file cannot test -- the model's own reasoning.
"""

from __future__ import annotations

import importlib.util
import shutil
import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parent.parent
EXAMPLES = REPO / "examples"
STACK = EXAMPLES / "sales_mis_stack"
RAW_SOURCE = STACK / "Sales MIS" / "bank_daily_sales_2025.xlsx"
DASHBOARD_TEMPLATE = STACK / "Sales MIS" / "bank_daily_sales_dashboard_2025.xlsx"

STACK_DOCS = ("persona.md", "skills.md", "workflow.md", "process.md", "policy.md", "memory.md")


def _runner():
    """Import examples/sales_mis_guru.py as a module without running main()."""
    spec = importlib.util.spec_from_file_location("sales_mis_guru", EXAMPLES / "sales_mis_guru.py")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


@pytest.fixture()
def stack_copy(tmp_path, monkeypatch):
    """A private copy of the stack's markdown -- loading it opens a sandbox
    under the copy's .ear/box, so tests never touch the real example's box
    (which a live run may be using at this very moment)."""
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    target = tmp_path / "stack"
    target.mkdir()
    for name in STACK_DOCS:
        shutil.copy(STACK / name, target / name)
    shutil.copytree(STACK / "knowledge", target / "knowledge")
    return target


def _sandbox_python_has_openpyxl(sandbox) -> bool:
    return sandbox.run(["python3", "-c", "import openpyxl"], timeout=30).ok


# -- the stack itself loads and binds everything it declares -------------------


def test_stack_loads_offline_with_sandbox_strategy_and_excel_skills(stack_copy):
    sys.path.insert(0, str(REPO))
    from ear import load_runtime

    runtime = load_runtime(stack_copy)
    assert runtime.sandbox is not None, "memory.md's Sandbox section must open a sandbox"
    assert runtime.sandbox.timeout == 90.0, "memory.md declares a 90-second shell timeout"
    assert runtime.model_binding is None, "without a credential the runtime must stay on the fallback"

    personas = {}
    for process in runtime.processes:
        for workflow in process.workflows:
            for step in workflow.steps:
                if step.persona is not None:
                    personas[step.persona.name] = {skill.name for skill in step.persona.skills}
    assert "Sales MIS Guru" in personas and "Dashboard Analyst" in personas and "MIS Controller" in personas
    assert "read_excel" in personas["Sales MIS Guru"]
    assert {"read_excel", "write_excel"} <= personas["Dashboard Analyst"]
    assert "read_excel" in personas["MIS Controller"]


# -- the sandbox environment can actually do the work the model will ask of it --


def test_sandbox_runs_an_openpyxl_read_script_on_the_real_workbook(stack_copy):
    """The exact path the model takes -- write_file a script, run_shell it --
    against the real raw workbook, inside the stack's own env allowlist,
    rlimits and timeout. If this fails, a live run burns money to discover
    the same thing."""
    sys.path.insert(0, str(REPO))
    from ear import load_runtime

    runtime = load_runtime(stack_copy)
    box = runtime.sandbox
    if not _sandbox_python_has_openpyxl(box):
        pytest.skip("sandbox python3 has no openpyxl -- install it before any live run")

    shutil.copy(RAW_SOURCE, box.resolve("uploads") / RAW_SOURCE.name)
    box.write_text(
        "workspace/probe.py",
        "import openpyxl\n"
        f"wb = openpyxl.load_workbook('uploads/{RAW_SOURCE.name}', data_only=False)\n"
        "print('sheets:', wb.sheetnames)\n"
        "ws = wb['Daily Sales Data']\n"
        "rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]\n"
        "print('data rows:', len(rows))\n",
    )
    result = box.run("python3 workspace/probe.py")
    assert result.ok, f"sandboxed openpyxl read failed: {result.render()}"
    assert "data rows: 907" in result.stdout, "the probe must observe the real row count"
    assert "Daily Sales Data" in result.stdout


def test_sandbox_runs_an_openpyxl_write_script_with_read_back_verification(stack_copy):
    """The write path write_excel demands: fill the real template, save to
    outputs/, re-open the saved copy and read back what was written."""
    sys.path.insert(0, str(REPO))
    from ear import load_runtime

    runtime = load_runtime(stack_copy)
    box = runtime.sandbox
    if not _sandbox_python_has_openpyxl(box):
        pytest.skip("sandbox python3 has no openpyxl -- install it before any live run")

    shutil.copy(DASHBOARD_TEMPLATE, box.resolve("uploads") / DASHBOARD_TEMPLATE.name)
    box.write_text(
        "workspace/fill.py",
        "import openpyxl\n"
        f"wb = openpyxl.load_workbook('uploads/{DASHBOARD_TEMPLATE.name}')\n"
        "ws = wb['Monthly Trend']\n"
        "ws.cell(row=3, column=2).value = 1042\n"
        "wb.save('outputs/completed.xlsx')\n"
        "back = openpyxl.load_workbook('outputs/completed.xlsx')\n"
        "assert back['Monthly Trend'].cell(row=3, column=2).value == 1042\n"
        "print('read-back verified; charts on Dashboard:', len(back['Dashboard']._charts))\n",
    )
    result = box.run("python3 workspace/fill.py")
    assert result.ok, f"sandboxed openpyxl write failed: {result.render()}"
    assert "read-back verified" in result.stdout
    assert (box.root / "outputs" / "completed.xlsx").exists()


def test_read_file_on_a_binary_workbook_returns_guidance_not_a_codec_error(tmp_path):
    """Live-trail failure shape: read_file on a .xlsx surfaced a bare
    'utf-8 codec can't decode' exception. The tool now names the situation
    and the way forward."""
    sys.path.insert(0, str(REPO))
    from ear import Sandbox

    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    shutil.copy(RAW_SOURCE, box.resolve("uploads") / RAW_SOURCE.name)
    read_file = next(tool for tool in box.as_tools() if tool.name == "read_file")
    message = read_file.handler(f"uploads/{RAW_SOURCE.name}")
    assert "binary file" in message and "write_file" in message and "run_shell" in message


# -- the runner never lies to a cycle and never trusts prose over the disk ------


def test_gate_reports_exactly_the_missing_handoffs(tmp_path):
    sys.path.insert(0, str(REPO))
    from ear import Sandbox

    runner = _runner()
    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    step2 = runner.STEPS[1]
    assert runner._missing_inputs(box, step2) == [runner.STAGED_DATASET]
    (box.root / runner.STAGED_DATASET).parent.mkdir(parents=True, exist_ok=True)
    (box.root / runner.STAGED_DATASET).write_text("Date,Branch\n", encoding="utf-8")
    assert runner._missing_inputs(box, step2) == []


def test_intent_text_carries_only_disk_verified_facts(tmp_path):
    sys.path.insert(0, str(REPO))
    from ear import Sandbox

    runner = _runner()
    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    staged = box.root / runner.STAGED_DATASET
    staged.parent.mkdir(parents=True, exist_ok=True)
    staged.write_text("x" * 1234, encoding="utf-8")
    text = runner._intent_text(box, runner.STEPS[1])
    assert "verified present, 1,234 bytes" in text
    assert f"output_expected: {runner.CLEAN_DATASET}" in text
    # An intent must never claim a prior step succeeded -- only file facts.
    assert "already ran" not in text and "clean result" not in text


def test_step_board_calls_out_outputs_missing_despite_decision(tmp_path):
    sys.path.insert(0, str(REPO))
    from ear import Sandbox

    runner = _runner()
    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    step1 = runner.STEPS[0]
    board = runner._step_board(box, runner.STEPS, {step1["name"]: {"status": "validated"}})
    assert "OUTPUTS MISSING despite decision" in board, "a decision claiming success with no artifact must be flagged"
    assert "not reached" in board  # later steps stay honest

    for rel in step1["produces"]:
        target = box.root / rel
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text("data", encoding="utf-8")
    board = runner._step_board(box, runner.STEPS, {step1["name"]: {"status": "validated"}})
    assert "outputs verified on disk" in board
    assert "-- present, 4 bytes" in board


def test_fresh_slate_clears_stale_artifacts_and_exchange_documents(tmp_path, capsys):
    sys.path.insert(0, str(REPO))
    from ear import Sandbox

    runner = _runner()
    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    (box.root / "workspace" / "staged_daily_sales.csv").write_text("stale", encoding="utf-8")
    (box.root / "outputs" / "old_dashboard.xlsx").write_text("stale", encoding="utf-8")
    stack = tmp_path / "stack"
    (stack / "intents").mkdir(parents=True)
    (stack / "decisions").mkdir(parents=True)
    for step in runner.STEPS:
        (stack / "intents" / step["intent_file"]).write_text("old intent", encoding="utf-8")
        (stack / "decisions" / step["intent_file"]).write_text("old decision", encoding="utf-8")

    runner._fresh_slate(box, runner.STEPS, stack=stack)

    assert not (box.root / "workspace" / "staged_daily_sales.csv").exists()
    assert not (box.root / "outputs" / "old_dashboard.xlsx").exists()
    for step in runner.STEPS:
        assert not (stack / "intents" / step["intent_file"]).exists()
        assert not (stack / "decisions" / step["intent_file"]).exists()
    assert "stale artifact" in capsys.readouterr().out


def test_decision_status_prefers_data_then_completion_then_headline():
    runner = _runner()

    # The declared Data field wins when present.
    data_doc = "# Decision\n\n## Data\n\n- status: validated\n- anomalies found: 0\n"
    assert runner._decision_status(data_doc) == "validated"

    # No Data section (the model omitted it), but the Completion summary's
    # always-written Status line carries the real status -- the board must
    # report it, not "(no status field)".
    completion_doc = (
        "# Decision -- Step 1\n\nStatus: decided\n\n## Completion summary\n\n"
        "- Status: decided (decision reports: validated)\n- Tool output summary:\n"
    )
    assert runner._decision_status(completion_doc) == "decided (decision reports: validated)"

    # Blocked/parked cycles have neither Data nor Completion summary, but the
    # document headline always states the status.
    blocked_doc = "# Decision -- Step 3\n\nStatus: BLOCKED\n\n## Decision\n\n> credit balance too low\n"
    assert runner._decision_status(blocked_doc) == "BLOCKED"

    # Only a document with no status anywhere reports the gap.
    assert runner._decision_status("# Decision\n\nno status anywhere\n") == "(no status field)"


def test_provider_failure_detection_matches_only_infrastructure_blocks():
    runner = _runner()
    billing = (
        "# Decision -- x\n\nStatus: BLOCKED\n\n## Decision\n\n"
        "> LLM call to https://api.anthropic.com/v1/messages failed (400): credit balance too low\n"
    )
    governance = "# Decision -- x\n\nStatus: BLOCKED\n\n## Decision\n\n> Loan Amount Cap VIOLATED\n"
    decided = "# Decision -- x\n\nStatus: decided\n\n## Decision\n\n> all good\n"
    assert runner._provider_failure(billing) is True
    assert runner._provider_failure(governance) is False
    assert runner._provider_failure(decided) is False


def test_pause_marker_names_the_step_reason_and_resume_command(tmp_path):
    runner = _runner()
    marker = runner._pause(tmp_path, runner.STEPS[1], "the provider call failed mid-cycle")
    text = marker.read_text(encoding="utf-8")
    assert marker == tmp_path / ".ear" / "paused.md"
    assert "PAUSED" in text and "Step 2 -- Sanity Check" in text
    assert "the provider call failed mid-cycle" in text
    assert "--resume" in text


def test_step_complete_is_filesystem_fact(tmp_path):
    sys.path.insert(0, str(REPO))
    from ear import Sandbox

    runner = _runner()
    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    step1 = runner.STEPS[0]
    assert runner._step_complete(box, step1) is False
    for rel in step1["produces"]:
        target = box.root / rel
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text("data", encoding="utf-8")
    assert runner._step_complete(box, step1) is True


# -- the live dashboard proves progress from the filesystem ---------------------


def test_dashboard_artifacts_panel_lists_every_area_with_sizes(tmp_path):
    sys.path.insert(0, str(REPO))
    from ear import Sandbox
    from ear.dashboard import _artifacts_section

    box = Sandbox.create(root=str(tmp_path / "box"), name="t")
    (box.root / "uploads" / "raw.xlsx").write_bytes(b"x" * 2048)
    (box.root / "workspace" / "staged.csv").write_text("a,b\n", encoding="utf-8")
    page = _artifacts_section(box)
    assert "inputs staged" in page and "work in progress" in page and "outputs produced" in page
    assert "raw.xlsx" in page and "staged.csv" in page
    assert "2.0 KB" in page
    assert "(empty)" in page  # outputs/ has nothing yet and says so

    assert _artifacts_section(None) == ""
